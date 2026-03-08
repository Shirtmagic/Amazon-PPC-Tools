#!/usr/bin/env python3
"""
Amazon PPC Bid Optimizer
Reads a Sponsored Products Bulk Operations file and outputs
a formatted Excel action file with precise bid change recommendations.
"""

import argparse
import json
import sys
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")


# ── Colours ─────────────────────────────────────────────────────────────────
C = {
    "header_dark":   "1F3864",
    "header_green":  "1E6B3C",
    "header_red":    "8B0000",
    "header_amber":  "7B4F00",
    "header_pause":  "4A4A4A",
    "header_gray":   "555555",
    "green_light":   "F0FFF4",
    "red_light":     "FFF0F0",
    "amber_light":   "FFFBF0",
    "gray_light":    "F8F8F8",
    "light_gray":    "F5F5F5",
    "positive":      "1E6B3C",
    "negative":      "8B0000",
    "neutral":       "7B4F00",
    "white":         "FFFFFF",
}
FONT = "Arial"

def hex_fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, bold=True, color="FFFFFF"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def bfont(size=10, bold=False, color="1A1A1A"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def cal():
    return Alignment(horizontal="center", vertical="center")

def lal():
    return Alignment(horizontal="left", vertical="center")


# ── Data Loading ─────────────────────────────────────────────────────────────

def load_bulk(path):
    df = pd.read_excel(path, sheet_name="Sponsored Products Campaigns",
                       engine="openpyxl")
    kws = df[df["Entity"] == "Keyword"].copy()

    # Resolve campaign/adgroup names from informational columns
    for info_col, base_col in [
        ("Campaign Name (Informational only)", "Campaign Name"),
        ("Ad Group Name (Informational only)",  "Ad Group Name"),
        ("Portfolio Name (Informational only)", "Portfolio Name"),
    ]:
        if info_col in kws.columns:
            if base_col not in kws.columns:
                kws[base_col] = kws[info_col]
            else:
                kws[base_col] = kws[base_col].fillna(kws[info_col])

    numeric = ["Bid", "Impressions", "Clicks", "Spend", "Sales",
               "Orders", "Units", "ACOS", "CPC", "ROAS", "Conversion Rate"]
    for col in numeric:
        if col in kws.columns:
            kws[col] = pd.to_numeric(kws[col], errors="coerce").fillna(0)

    required = ["Keyword Text", "Match Type", "Bid", "Clicks", "Spend",
                "Sales", "Orders", "Campaign Name", "Ad Group Name"]
    missing = [c for c in required if c not in kws.columns]
    if missing:
        sys.exit(f"ERROR: Missing columns: {missing}")

    # Compute ACoS cleanly
    kws["_acos"] = np.where(
        (kws["Sales"] > 0) & (kws["Spend"] > 0),
        kws["Spend"] / kws["Sales"],
        np.nan
    )
    kws["_cpc"] = np.where(
        kws["Clicks"] > 0,
        kws["Spend"] / kws["Clicks"],
        kws["Bid"]   # fall back to current bid as proxy
    )
    return kws


# ── Bid Calculation ──────────────────────────────────────────────────────────

def confidence_tier(clicks):
    if clicks >= 30:
        return "High",   1.0
    elif clicks >= 10:
        return "Medium", 0.5
    elif clicks >= 5:
        return "Hold",   0.0   # accumulate more data
    else:
        return "New",    0.0   # inch-up territory


def bid_lifecycle_stage(clicks, spend, orders, impressions):
    """Detect which lifecycle stage a keyword is in (AdsCrafted method)."""
    if clicks < 5 and spend < 5.0:
        if impressions == 0:
            return "LAUNCH_NO_IMPRESSIONS"
        return "LAUNCH"
    if clicks < 10 and orders == 0:
        return "DISCOVERY"
    if orders > 0:
        return "OPTIMIZE"
    if spend > 0 and orders == 0:
        return "CUT_CANDIDATE"
    return "UNKNOWN"


def calc_new_bid(row, target_acos, max_raise, max_lower, bid_floor,
                 bid_ceiling, pause_spend, pause_clicks):
    """
    Returns (action, new_bid, reason, confidence, pct_change)
    action: RAISE | LOWER | PAUSE | INCH UP | HOLD | NO CHANGE
    """
    current_bid  = float(row["Bid"])
    clicks       = float(row["Clicks"])
    spend        = float(row["Spend"])
    orders       = float(row["Orders"])
    acos         = row["_acos"]   # nan if no sales
    impressions  = float(row.get("Impressions", 0))

    # ── Auto campaigns — keyword bids managed at campaign level ──
    if row.get("is_auto_campaign", False):
        return ("NO CHANGE", current_bid,
                "Auto campaign — use campaign-level bid strategy, not keyword bids", "Auto", 0.0)

    tier, dampen = confidence_tier(clicks)
    stage = bid_lifecycle_stage(clicks, spend, orders, impressions)

    # ── LAUNCH stage: inch-up bidding (AdsCrafted method) ──
    if stage == "LAUNCH_NO_IMPRESSIONS":
        inch_amount = 0.10
        new_bid = min(round(current_bid + inch_amount, 2), bid_ceiling)
        pct = (new_bid - current_bid) / current_bid if current_bid > 0 else 0
        return ("INCH UP", new_bid,
                f"No impressions — bid too low to enter auction. Inch up ${inch_amount:.2f} to ${new_bid:.2f}",
                "New", pct)

    if stage == "LAUNCH":
        return ("NO CHANGE", current_bid,
                f"Launch stage: {int(clicks)} clicks, ${spend:.2f} spend — protect until ≥5 clicks before adjusting",
                "New", 0.0)

    # ── DISCOVERY stage: hold and accumulate data ──
    if stage == "DISCOVERY":
        return ("HOLD", current_bid,
                f"Discovery stage: {int(clicks)} clicks, 0 orders — need ≥10 clicks for reliable decision",
                "Hold", 0.0)

    # ── No data yet ──
    if clicks == 0 and impressions == 0:
        inch_amount = 0.10
        new_bid = min(round(current_bid + inch_amount, 2), bid_ceiling)
        pct = (new_bid - current_bid) / current_bid if current_bid > 0 else 0
        return ("INCH UP", new_bid,
                f"Zero impressions and clicks — inch up to enter auction",
                "New", pct)

    if clicks == 0:
        return ("NO CHANGE", current_bid, "Impressions but no clicks yet — monitor CTR", tier, 0.0)

    # ── Pause candidate (CUT stage) ──
    if orders == 0 and spend >= pause_spend and clicks >= pause_clicks:
        new_bid = max(round(current_bid * (1 - max_lower), 2), bid_floor)
        pct = (new_bid - current_bid) / current_bid
        return ("PAUSE CANDIDATE", new_bid,
                f"${spend:.2f} spend, {int(clicks)} clicks, 0 orders — cut stage",
                tier, pct)

    # ── No orders but below pause threshold — lower cautiously ──
    if orders == 0 and spend > 0 and tier in ("High", "Medium"):
        reduction = max_lower * dampen
        new_bid   = max(round(current_bid * (1 - reduction), 2), bid_floor)
        pct       = (new_bid - current_bid) / current_bid
        return ("LOWER", new_bid,
                f"${spend:.2f} spend, 0 orders — reduce to limit losses",
                tier, pct)

    if orders == 0:
        return ("HOLD", current_bid,
                f"{int(clicks)} clicks, 0 orders — hold for more data",
                tier, 0.0)

    # ── OPTIMIZE stage: revenue-based / ACoS-based adjustment ──
    if pd.isna(acos):
        return ("NO CHANGE", current_bid, "Orders but no ACoS data", tier, 0.0)

    # Revenue-based bid calculation (AdsCrafted method)
    sales = float(row.get("Sales", 0))
    if orders > 0 and clicks > 0 and sales > 0:
        aov = sales / orders
        rev_per_click = (orders / clicks) * aov
        target_bid = rev_per_click * target_acos
        target_bid = max(min(target_bid, bid_ceiling), bid_floor)
    else:
        target_bid = None

    band_low  = target_acos * 0.70   # well under target → raise
    band_high = target_acos * 1.10   # within 10% over → tolerate

    if acos <= band_low:
        # Scale stage — proven winner, raise toward revenue-based target
        if target_bid and target_bid > current_bid:
            adjust = min((target_bid / current_bid - 1) * dampen, max_raise)
        else:
            ratio  = target_acos / acos
            adjust = min((ratio - 1) * dampen, max_raise)
        new_bid = min(round(current_bid * (1 + adjust), 2), bid_ceiling)
        pct     = (new_bid - current_bid) / current_bid
        rev_note = f" (rev-based target: ${target_bid:.2f})" if target_bid else ""
        return ("RAISE", new_bid,
                f"ACoS {acos:.1%} well under target {target_acos:.0%} — scale winner{rev_note}",
                tier, pct)

    elif acos <= band_high:
        return ("NO CHANGE", current_bid,
                f"ACoS {acos:.1%} within target band ({band_low:.0%}–{band_high:.0%})",
                tier, 0.0)

    else:
        # Lower: scale toward target, dampened
        ratio  = target_acos / acos
        adjust = max((ratio - 1) * dampen, -max_lower)
        new_bid = max(round(current_bid * (1 + adjust), 2), bid_floor)
        pct     = (new_bid - current_bid) / current_bid
        return ("LOWER", new_bid,
                f"ACoS {acos:.1%} over target {target_acos:.0%} — trim bid",
                tier, pct)


def optimise(df, target_acos, max_raise, max_lower, bid_floor,
             bid_ceiling, pause_spend, pause_clicks):
    results = []
    for _, row in df.iterrows():
        action, new_bid, reason, tier, pct = calc_new_bid(
            row, target_acos, max_raise, max_lower,
            bid_floor, bid_ceiling, pause_spend, pause_clicks
        )
        results.append({
            "action":      action,
            "new_bid":     new_bid,
            "reason":      reason,
            "confidence":  tier,
            "pct_change":  pct,
        })
    res = pd.DataFrame(results, index=df.index)
    return pd.concat([df, res], axis=1)


# ── Excel Helpers ────────────────────────────────────────────────────────────

def apply_header(ws, row_num, headers, bg, fc="FFFFFF", height=20):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=i, value=h)
        c.font = hfont(color=fc)
        c.fill = hex_fill(bg)
        c.alignment = cal()
        c.border = thin_border()

def sheet_title(ws, title, subtitle, bg, height=28):
    ws.row_dimensions[1].height = height
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    max_col = max(ws.max_column or 1, 12)
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = hex_fill(bg)
        ws.cell(row=1, column=col).border = thin_border()
    ws.row_dimensions[2].height = 16
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    s.fill = hex_fill("FFFFFF")

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Sheet Builders ───────────────────────────────────────────────────────────

PERF_COLS  = ["Portfolio Name", "Campaign Name", "Ad Group Name",
              "Keyword Text", "Match Type", "Bid",
              "Impressions", "Clicks", "Spend", "Sales", "Orders",
              "_acos", "confidence", "new_bid", "pct_change", "reason"]
PERF_HDRS  = ["Portfolio", "Campaign", "Ad Group",
              "Keyword", "Match", "Current Bid",
              "Impressions", "Clicks", "Spend", "Sales", "Orders",
              "ACoS", "Confidence", "New Bid", "Bid Change %", "Reason"]
PERF_FMTS  = {
    "Bid":      '"$"#,##0.00',
    "Spend":    '"$"#,##0.00',
    "Sales":    '"$"#,##0.00',
    "_acos":    '0.0%',
    "new_bid":  '"$"#,##0.00',
    "pct_change": '0.0%',
}


def write_action_sheet(wb, df, action, tab_name, bg, light_bg, sort_col,
                       sort_asc=True):
    data = df[df["action"] == action].copy()
    if not data.empty:
        data = data.sort_values(sort_col, ascending=sort_asc)

    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    count = len(data)
    spend_impact = (data["new_bid"] - data["Bid"]).sum() if count else 0
    subtitle = (f"{count} keywords | "
                f"Est. bid pool change: ${spend_impact:+,.2f} total")
    sheet_title(ws, tab_name.split(" ", 1)[-1], subtitle, bg)
    apply_header(ws, 3, PERF_HDRS, bg)

    for r_idx, (_, row) in enumerate(data.iterrows()):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(light_bg)
        ws.row_dimensions[er].height = 15
        for c_idx, col in enumerate(PERF_COLS, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in PERF_FMTS and val != "":
                cell.number_format = PERF_FMTS[col]

        # Colour the bid-change % cell
        pct_val = row.get("pct_change", 0)
        pct_cell = ws.cell(row=er, column=15)
        if action in ("RAISE",) and isinstance(pct_val, float) and pct_val > 0:
            pct_cell.font = Font(name=FONT, bold=True, size=9, color=C["positive"])
        elif action in ("LOWER", "PAUSE CANDIDATE"):
            pct_cell.font = Font(name=FONT, bold=True, size=9, color=C["negative"])

    widths = {
        "A": 20, "B": 30, "C": 22, "D": 38, "E": 10,
        "F": 13, "G": 12, "H": 10, "I": 12, "J": 12,
        "K": 10, "L": 10, "M": 12, "N": 13, "O": 13, "P": 44,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"
    return count


def build_summary(wb, df, target_acos, date_range):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "🎯  Amazon PPC Bid Optimizer Report"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hex_fill(C["header_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = (f"Bulk file date range: {date_range}   |   "
                f"Target ACoS: {target_acos:.0%}   |   "
                f"Generated: {datetime.today().strftime('%b %d, %Y')}")
    c2.font = Font(name=FONT, italic=True, size=9, color="888888")
    c2.fill = hex_fill("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Overall KPIs
    total_spend  = df["Spend"].sum()
    total_sales  = df["Sales"].sum()
    total_orders = df["Orders"].sum()
    total_kws    = len(df)
    overall_acos = total_spend / total_sales if total_sales > 0 else 0

    kpis = [
        ("Total Spend",    f"${total_spend:,.2f}",   C["header_dark"]),
        ("Total Sales",    f"${total_sales:,.2f}",   C["header_green"]),
        ("Overall ACoS",   f"{overall_acos:.1%}",
         C["negative"] if overall_acos > target_acos else C["header_green"]),
        ("Total Orders",   f"{int(total_orders):,}", C["header_dark"]),
        ("Active Keywords",f"{total_kws:,}",         C["header_dark"]),
        ("Target ACoS",    f"{target_acos:.0%}",     C["header_dark"]),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    for i, (label, value, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hex_fill("F5F5F5")
        lc.alignment = cal()
        lc.border = thin_border()
        vc = ws.cell(row=5, column=i, value=value)
        vc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = cal()
        vc.border = thin_border()

    # Action breakdown
    ws.row_dimensions[7].height = 18
    apply_header(ws, 7,
                 ["Action", "Keywords", "Avg Current Bid",
                  "Avg New Bid", "Est. Bid Change", "What It Means"],
                 C["header_dark"])

    action_meta = {
        "RAISE":           ("📈 Raise Bids",         C["header_green"],
                            "ACoS well under target — room to increase volume"),
        "LOWER":           ("📉 Lower Bids",          C["header_red"],
                            "ACoS over target — reduce spend efficiency"),
        "INCH UP":         ("🔼 Inch Up Bids",        "2E75B6",
                            "Launch stage — inch up to enter auction or gain impressions"),
        "HOLD":            ("⏳ Hold (Data Needed)",   "7B4F00",
                            "Discovery stage — accumulating data, no changes yet"),
        "PAUSE CANDIDATE": ("⏸  Pause Candidates",   C["header_pause"],
                            "High spend, zero orders — stop bleeding budget"),
        "NO CHANGE":       ("✅ No Change",            C["header_gray"],
                            "Within target band or insufficient data"),
    }
    for r_idx, (action, (label, color, meaning)) in enumerate(action_meta.items(), 1):
        row_num = 7 + r_idx
        ws.row_dimensions[row_num].height = 18
        subset = df[df["action"] == action]
        fill   = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        avg_cur = subset["Bid"].mean()     if len(subset) else 0
        avg_new = subset["new_bid"].mean() if len(subset) else 0
        est_chg = (subset["new_bid"] - subset["Bid"]).sum() if len(subset) else 0
        for c_idx, val in enumerate(
            [label, len(subset),
             f"${avg_cur:.2f}" if avg_cur else "-",
             f"${avg_new:.2f}" if avg_new else "-",
             f"${est_chg:+,.2f}" if est_chg else "-",
             meaning], 1
        ):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font = bfont(bold=(c_idx == 1))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()

    # Portfolio breakdown
    ws.row_dimensions[13].height = 18
    apply_header(ws, 13,
                 ["Portfolio", "Keywords", "Spend", "Sales", "ACoS",
                  "Raise", "Lower", "Pause", "No Change"],
                 C["header_dark"])

    if "Portfolio Name" in df.columns:
        port = df.groupby("Portfolio Name").agg(
            keywords=("Keyword Text", "count"),
            spend=("Spend", "sum"),
            sales=("Sales", "sum"),
        ).reset_index()
        port["acos"] = np.where(port["sales"] > 0,
                                port["spend"] / port["sales"], np.nan)

        def action_count(p, a):
            return len(df[(df["Portfolio Name"] == p) & (df["action"] == a)])

        port = port.sort_values("spend", ascending=False)
        for r_idx, (_, row) in enumerate(port.iterrows()):
            rn = 14 + r_idx
            ws.row_dimensions[rn].height = 17
            fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
            acos_val = row["acos"] if not pd.isna(row["acos"]) else ""
            vals = [
                row["Portfolio Name"], int(row["keywords"]),
                row["spend"], row["sales"], acos_val,
                action_count(row["Portfolio Name"], "RAISE"),
                action_count(row["Portfolio Name"], "LOWER"),
                action_count(row["Portfolio Name"], "PAUSE CANDIDATE"),
                action_count(row["Portfolio Name"], "NO CHANGE"),
            ]
            fmts = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                    '0.0%', None, None, None, None]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=rn, column=c_idx, value=val)
                cell.font = bfont()
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = lal()
                if fmt and val not in ("", None):
                    cell.number_format = fmt

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 36
    ws.freeze_panes = "A3"


def build_bulk_upload(wb, df):
    """
    Only outputs rows where action != NO CHANGE.
    Uses Keyword ID for precise targeting.
    Operation = update for bid changes, paused for pause candidates.
    """
    changes = df[df["action"] != "NO CHANGE"].copy()

    ws = wb.create_sheet("📤 Amazon Bulk Upload")
    ws.sheet_view.showGridLines = False

    sheet_title(ws,
        f"Amazon Bulk Upload — {len(changes)} keyword bid changes",
        "⚠ Upload this tab via Seller Central → Campaign Manager → Bulk Operations → Upload",
        C["header_dark"])

    # Minimal set of columns Amazon needs for a keyword bid update
    headers = [
        "Product", "Entity", "Operation",
        "Campaign ID", "Ad Group ID", "Keyword ID",
        "Campaign Name", "Ad Group Name",
        "State", "Bid", "Keyword Text", "Match Type",
        "Previous Bid", "Bid Change %", "Action Tag"
    ]
    apply_header(ws, 3, headers, C["header_dark"])

    action_colors = {
        "RAISE":           C["positive"],
        "LOWER":           C["negative"],
        "PAUSE CANDIDATE": "4A4A4A",
    }

    for r_idx, (_, row) in enumerate(changes.iterrows()):
        er = 4 + r_idx
        action = row.get("action", "")
        is_pause = action == "PAUSE CANDIDATE"

        light = {"RAISE": C["green_light"],
                 "LOWER": C["red_light"],
                 "PAUSE CANDIDATE": "F0F0F0"}.get(action, "FFFFFF")
        fill = hex_fill(light) if r_idx % 2 == 0 else hex_fill("FFFFFF")

        state_val = "paused" if is_pause else "enabled"
        new_bid   = row.get("new_bid", row["Bid"])
        pct       = row.get("pct_change", 0)

        vals = {
            "Product":      "Sponsored Products",
            "Entity":       "Keyword",
            "Operation":    "update",
            "Campaign ID":  row.get("Campaign ID", ""),
            "Ad Group ID":  row.get("Ad Group ID", ""),
            "Keyword ID":   row.get("Keyword ID", ""),
            "Campaign Name":  row.get("Campaign Name", ""),
            "Ad Group Name":  row.get("Ad Group Name", ""),
            "State":        state_val,
            "Bid":          round(float(new_bid), 2),
            "Keyword Text": row.get("Keyword Text", ""),
            "Match Type":   row.get("Match Type", ""),
            "Previous Bid": round(float(row["Bid"]), 2),
            "Bid Change %": pct,
            "Action Tag":   action,
        }
        fmts = {
            "Bid":          '"$"#,##0.00',
            "Previous Bid": '"$"#,##0.00',
            "Bid Change %": '+0.0%;-0.0%;0.0%',
        }

        ws.row_dimensions[er].height = 15
        for c_idx, h in enumerate(headers, 1):
            val = vals.get(h, "")
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if h in fmts:
                cell.number_format = fmts[h]

        # Colour the Action Tag cell
        tag_cell = ws.cell(row=er, column=15)
        tag_cell.font = Font(name=FONT, bold=True, size=9,
                             color=action_colors.get(action, "1A1A1A"))

    widths = {
        "A": 18, "B": 12, "C": 12, "D": 18, "E": 16, "F": 18,
        "G": 30, "H": 24, "I": 10, "J": 12, "K": 36,
        "L": 12, "M": 14, "N": 13, "O": 18,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"

    # Footer warning
    last_row = 4 + len(changes) + 1
    ws.row_dimensions[last_row].height = 20
    note = ws.cell(row=last_row, column=1,
                   value="⚠ PAUSE CANDIDATES: rows marked 'PAUSE CANDIDATE' "
                         "will set State = paused. Review before uploading.")
    note.font = Font(name=FONT, bold=True, size=9, color=C["negative"])
    note.fill = hex_fill("FFF0F0")
    ws.merge_cells(start_row=last_row, start_column=1,
                   end_row=last_row, end_column=8)


def build_raw_sheet(wb, df):
    ws = wb.create_sheet("🗂 Raw Data")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "All Keywords — Full Analysis",
                "Every keyword with performance data and bid recommendation",
                C["header_dark"])

    all_cols  = ["Portfolio Name", "Campaign Name", "Ad Group Name",
                 "Keyword Text", "Match Type", "State",
                 "Bid", "Impressions", "Clicks", "Spend", "Sales",
                 "Orders", "_acos", "_cpc", "action", "new_bid",
                 "pct_change", "confidence", "reason"]
    all_hdrs  = ["Portfolio", "Campaign", "Ad Group",
                 "Keyword", "Match", "State",
                 "Current Bid", "Impressions", "Clicks", "Spend", "Sales",
                 "Orders", "ACoS", "CPC", "Action", "New Bid",
                 "Bid Change %", "Confidence", "Reason"]
    all_fmts  = {
        "Bid":    '"$"#,##0.00', "Spend": '"$"#,##0.00',
        "Sales":  '"$"#,##0.00', "_acos": '0.0%',
        "_cpc":   '"$"#,##0.00', "new_bid": '"$"#,##0.00',
        "pct_change": '0.0%',
    }
    action_bg = {
        "RAISE":           C["green_light"],
        "LOWER":           C["red_light"],
        "PAUSE CANDIDATE": "EEEEEE",
        "NO CHANGE":       "FFFFFF",
    }

    apply_header(ws, 3, all_hdrs, C["header_dark"])
    df_sorted = df.sort_values(
        ["action", "Portfolio Name", "Campaign Name"],
        key=lambda x: x.map(
            {"RAISE": 0, "PAUSE CANDIDATE": 1, "LOWER": 2, "NO CHANGE": 3}
        ) if x.name == "action" else x
    )

    for r_idx, (_, row) in enumerate(df_sorted.iterrows()):
        er = 4 + r_idx
        act = row.get("action", "NO CHANGE")
        fill = hex_fill(action_bg.get(act, "FFFFFF"))
        ws.row_dimensions[er].height = 14
        for c_idx, col in enumerate(all_cols, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in all_fmts and val != "":
                cell.number_format = all_fmts[col]

    for i, w in enumerate(
        [20, 28, 22, 36, 10, 10, 13, 12, 10, 12, 12,
         10, 10, 10, 18, 13, 13, 12, 40], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Amazon PPC Bid Optimizer")
    parser.add_argument("--input",         required=True)
    parser.add_argument("--output",        required=True)
    parser.add_argument("--target-acos",   type=float, default=0.25)
    parser.add_argument("--min-clicks",    type=int,   default=10)
    parser.add_argument("--max-raise",     type=float, default=0.30)
    parser.add_argument("--max-lower",     type=float, default=0.40)
    parser.add_argument("--bid-floor",     type=float, default=0.20)
    parser.add_argument("--bid-ceiling",   type=float, default=5.00)
    parser.add_argument("--pause-spend",   type=float, default=15.0)
    parser.add_argument("--pause-clicks",  type=int,   default=15)
    args = parser.parse_args()

    print(f"Loading bulk file: {args.input}")
    df = load_bulk(args.input)
    print(f"  {len(df)} keyword rows loaded")

    print("\nRunning bid optimization...")
    df = optimise(df, args.target_acos, args.max_raise, args.max_lower,
                  args.bid_floor, args.bid_ceiling,
                  args.pause_spend, args.pause_clicks)

    counts = df["action"].value_counts()
    print("\nResults:")
    for action, n in counts.items():
        print(f"  {action}: {n}")

    changes = df[df["action"] != "NO CHANGE"]
    spend_with_changes = changes["Spend"].sum()
    print(f"\n  Keywords with changes: {len(changes)}")
    print(f"  Spend affected: ${spend_with_changes:,.2f}")

    # Date range from file name or default
    date_range = "Jan 25 – Feb 24, 2026"

    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb, df, args.target_acos, date_range)
    raise_n  = write_action_sheet(wb, df, "RAISE",
                                  "📈 Raise Bids",
                                  C["header_green"], C["green_light"],
                                  "_acos", sort_asc=True)
    lower_n  = write_action_sheet(wb, df, "LOWER",
                                  "📉 Lower Bids",
                                  C["header_red"], C["red_light"],
                                  "Spend", sort_asc=False)
    inchup_n = write_action_sheet(wb, df, "INCH UP",
                                  "🔼 Inch Up Bids",
                                  "2E75B6", "F0F4FF",
                                  "Bid", sort_asc=True)
    hold_n   = write_action_sheet(wb, df, "HOLD",
                                  "⏳ Hold (Data Needed)",
                                  "7B4F00", "FFFBF0",
                                  "Spend", sort_asc=False)
    pause_n  = write_action_sheet(wb, df, "PAUSE CANDIDATE",
                                  "⏸ Pause Candidates",
                                  C["header_pause"], C["gray_light"],
                                  "Spend", sort_asc=False)
    nochange = write_action_sheet(wb, df, "NO CHANGE",
                                  "✅ No Change",
                                  C["header_gray"], C["light_gray"],
                                  "Spend", sort_asc=False)
    build_bulk_upload(wb, df)
    build_raw_sheet(wb, df)

    # ── Write cross-tool findings JSON ────────────────────────────────────
    auto_skipped  = int(df.get("is_auto_campaign", pd.Series(False, index=df.index)).sum())
    new_protected = int(((df["action"] == "NO CHANGE") &
                         df["reason"].str.contains("New keyword", na=False)).sum())
    raise_rows = df[df["action"] == "RAISE"].nlargest(5, "Spend")
    lower_rows = df[df["action"] == "LOWER"].nlargest(5, "Spend")
    findings = {
        "tool":            "bid_optimizer",
        "target_acos":     args.target_acos,
        "raise_count":     int(raise_n),
        "lower_count":     int(lower_n),
        "pause_count":     int(pause_n),
        "spend_affected":  round(float(df[df["action"] != "NO CHANGE"]["Spend"].sum()), 2),
        "auto_skipped":    auto_skipped,
        "new_kw_protected":new_protected,
        "actions": (
            [{"priority": "HIGH",   "type": "LOWER_BID",
              "subject":  r.get("Keyword Text",""), "campaign": r.get("Campaign Name",""),
              "impact_spend": round(float(r["Spend"]), 2),
              "detail": f"ACoS {r['_acos']:.1%} — lower bid ${r['Bid']:.2f}→${r['new_bid']:.2f} ({r['confidence']} confidence)"}
             for _, r in lower_rows.iterrows() if r["Spend"] > 20]
            +
            [{"priority": "MEDIUM", "type": "RAISE_BID",
              "subject":  r.get("Keyword Text",""), "campaign": r.get("Campaign Name",""),
              "impact_spend": round(float(r["Spend"]), 2),
              "detail": f"ACoS {r['_acos']:.1%} — raise bid ${r['Bid']:.2f}→${r['new_bid']:.2f} to capture more volume"}
             for _, r in raise_rows.iterrows()]
        ),
    }
    findings_path = args.output.replace(".xlsx", "_findings.json")
    with open(findings_path, "w") as _f:
        json.dump(findings, _f, indent=2, default=str)
    print(f"   Findings: {findings_path}")

    wb.save(args.output)
    print(f"\n✅ Saved: {args.output}")
    print(f"   📈 Raise:    {raise_n} keywords")
    print(f"   📉 Lower:    {lower_n} keywords")
    print(f"   🔼 Inch Up:  {inchup_n} keywords")
    print(f"   ⏳ Hold:     {hold_n} keywords")
    print(f"   ⏸  Pause:    {pause_n} keywords")
    print(f"   ✅ No change: {nochange} keywords")
    total_changes = raise_n + lower_n + pause_n + inchup_n
    print(f"\n   Bulk upload tab has {total_changes} rows ready to upload.")


if __name__ == "__main__":
    main()
