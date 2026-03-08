#!/usr/bin/env python3
"""
Amazon PPC Rank Tracker
Reads a Sponsored Products Bulk Operations file and outputs
a formatted Excel report with keyword ranking scores, bait & switch
readiness, day parting recommendations, and budget adequacy analysis.
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")


# -- Colours -----------------------------------------------------------------
C = {
    "header_dark":   "1F3864",
    "header_green":  "1E6B3C",
    "header_red":    "8B0000",
    "header_amber":  "7B4F00",
    "header_pause":  "4A4A4A",
    "header_gray":   "555555",
    "header_purple": "4A1A7A",
    "header_blue":   "1A5276",
    "green_light":   "F0FFF4",
    "red_light":     "FFF0F0",
    "amber_light":   "FFFBF0",
    "purple_light":  "F8F0FF",
    "blue_light":    "F0F8FF",
    "gray_light":    "F8F8F8",
    "light_gray":    "F5F5F5",
    "positive":      "1E6B3C",
    "negative":      "8B0000",
    "neutral":       "7B4F00",
    "white":         "FFFFFF",
}
FONT = "Arial"


def hex_fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, bold=True, color="FFFFFF"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def bfont(size=10, bold=False, color="1A1A1A"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def cal():
    return Alignment(horizontal="center", vertical="center")

def lal():
    return Alignment(horizontal="left", vertical="center")


# -- Data Loading ------------------------------------------------------------

def load_bulk(path):
    df = pd.read_excel(path, sheet_name="Sponsored Products Campaigns",
                       engine="openpyxl")

    # Resolve informational columns
    for info_col, base_col in [
        ("Campaign Name (Informational only)", "Campaign Name"),
        ("Ad Group Name (Informational only)",  "Ad Group Name"),
        ("Portfolio Name (Informational only)", "Portfolio Name"),
    ]:
        if info_col in df.columns:
            if base_col not in df.columns:
                df[base_col] = df[info_col]
            else:
                df[base_col] = df[base_col].fillna(df[info_col])

    return df


def extract_campaigns(df):
    """Pull campaign-level rows for budget and placement data."""
    camps = df[df["Entity"] == "Campaign"].copy()
    numeric = ["Budget", "Impressions", "Clicks", "Spend", "Sales", "Orders"]
    for col in numeric:
        if col in camps.columns:
            camps[col] = pd.to_numeric(camps[col], errors="coerce").fillna(0)
    return camps


def extract_keywords(df):
    """Pull keyword rows with numeric coercion."""
    kws = df[df["Entity"] == "Keyword"].copy()
    numeric = ["Bid", "Impressions", "Clicks", "Spend", "Sales",
               "Orders", "Units", "ACOS", "CPC", "ROAS",
               "Conversion Rate", "Click-through Rate"]
    for col in numeric:
        if col in kws.columns:
            kws[col] = pd.to_numeric(kws[col], errors="coerce").fillna(0)

    # Top of Search Impression Share
    tos_col = None
    for candidate in ["Top of Search Impression Share",
                      "Top-of-search Impression Share",
                      "Impressions Top of Search %"]:
        if candidate in kws.columns:
            tos_col = candidate
            break
    if tos_col:
        kws["_tos_is"] = pd.to_numeric(
            kws[tos_col].astype(str).str.replace("%", ""),
            errors="coerce").fillna(0) / 100.0
    else:
        kws["_tos_is"] = 0.0

    # Compute derived metrics
    kws["_acos"] = np.where(
        (kws["Sales"] > 0) & (kws["Spend"] > 0),
        kws["Spend"] / kws["Sales"], np.nan)
    kws["_ctr"] = np.where(
        kws["Impressions"] > 0,
        kws["Clicks"] / kws["Impressions"], 0.0)
    kws["_cvr"] = np.where(
        kws["Clicks"] > 0,
        kws["Orders"] / kws["Clicks"], 0.0)

    required = ["Keyword Text", "Match Type", "Bid", "Clicks", "Spend",
                "Sales", "Orders", "Campaign Name", "Ad Group Name"]
    missing = [c for c in required if c not in kws.columns]
    if missing:
        sys.exit(f"ERROR: Missing columns: {missing}")

    return kws


def identify_ranking_campaigns(kws, camps_df, brand):
    """
    Identify ranking campaigns by:
    1. Campaign name contains rank/launch/SKC/skc
    2. Exact match with Top of Search placement modifier > 50%
    3. Campaigns with a single keyword per ad group
    """
    ranking_flags = []
    ranking_patterns = re.compile(r"(rank|launch|skc)", re.IGNORECASE)

    # Build placement lookup from campaign rows
    placement_map = {}
    if "Bidding Strategy" in camps_df.columns:
        for _, row in camps_df.iterrows():
            cname = row.get("Campaign Name", "")
            # Top of Search placement modifier
            tos_mod = 0
            for col in camps_df.columns:
                if "Top of Search" in col and "Modifier" in col:
                    tos_mod = pd.to_numeric(row.get(col, 0), errors="coerce")
                    if pd.isna(tos_mod):
                        tos_mod = 0
                    break
            placement_map[cname] = tos_mod

    # Count keywords per ad group
    kw_per_ag = kws.groupby(["Campaign Name", "Ad Group Name"]).size()

    for idx, row in kws.iterrows():
        cname = str(row.get("Campaign Name", ""))
        ag_name = str(row.get("Ad Group Name", ""))
        match_type = str(row.get("Match Type", "")).lower()
        reasons = []

        # Check 1: name patterns
        if ranking_patterns.search(cname):
            reasons.append("Campaign name contains rank/launch/SKC")

        # Check 2: exact match + high ToS placement
        tos_mod = placement_map.get(cname, 0)
        if match_type == "exact" and tos_mod > 50:
            reasons.append(f"Exact match + ToS modifier {tos_mod}%")

        # Check 3: single keyword ad group
        ag_key = (cname, ag_name)
        if ag_key in kw_per_ag.index and kw_per_ag[ag_key] == 1:
            reasons.append("Single keyword ad group")

        is_ranking = len(reasons) > 0
        ranking_flags.append({
            "is_ranking":       is_ranking,
            "ranking_reason":   " | ".join(reasons) if reasons else "",
            "_tos_modifier":    tos_mod,
        })

    flags_df = pd.DataFrame(ranking_flags, index=kws.index)
    return pd.concat([kws, flags_df], axis=1)


def load_sqp(path):
    """Load Search Query Performance data."""
    try:
        sqp = pd.read_csv(path)
    except Exception:
        sqp = pd.read_excel(path, engine="openpyxl")

    # Normalise column names
    col_map = {}
    for col in sqp.columns:
        low = col.lower().strip()
        if "search query" in low or "search term" in low:
            col_map[col] = "search_query"
        elif "impression" in low and "share" in low and "organic" in low:
            col_map[col] = "organic_impression_share"
        elif "click" in low and "share" in low and "organic" in low:
            col_map[col] = "organic_click_share"
        elif "conversion" in low and "share" in low:
            col_map[col] = "conversion_share"
        elif "search query volume" in low or "search frequency rank" in low:
            col_map[col] = "search_volume_rank"

    sqp = sqp.rename(columns=col_map)
    for num_col in ["organic_impression_share", "organic_click_share",
                    "conversion_share"]:
        if num_col in sqp.columns:
            sqp[num_col] = pd.to_numeric(
                sqp[num_col].astype(str).str.replace("%", ""),
                errors="coerce").fillna(0)
    return sqp


def merge_sqp(kws, sqp):
    """Merge SQP metrics into keyword data by keyword text match."""
    if "search_query" not in sqp.columns:
        return kws

    sqp_lookup = sqp.set_index(
        sqp["search_query"].str.lower().str.strip())

    sqp_cols = ["organic_impression_share", "organic_click_share",
                "conversion_share"]
    for col in sqp_cols:
        kws[f"_sqp_{col}"] = 0.0

    for idx, row in kws.iterrows():
        kw = str(row.get("Keyword Text", "")).lower().strip()
        if kw in sqp_lookup.index:
            sqp_row = sqp_lookup.loc[kw]
            if isinstance(sqp_row, pd.DataFrame):
                sqp_row = sqp_row.iloc[0]
            for col in sqp_cols:
                if col in sqp_row.index:
                    kws.at[idx, f"_sqp_{col}"] = float(sqp_row[col])

    kws["_has_sqp"] = (kws["_sqp_organic_impression_share"] > 0) | \
                      (kws["_sqp_organic_click_share"] > 0)
    return kws


# -- Ranking Score -----------------------------------------------------------

def calc_ranking_score(row, camps_budget_map):
    """
    Compute ranking score 0-100:
      CVR 30% | CTR 20% | ToS IS 20% | Sales consistency 15% | Budget util 15%
    """
    score = 0
    breakdown = {}

    # 1. Conversion Rate (30% weight)
    cvr = float(row.get("_cvr", 0))
    if cvr >= 0.15:
        pts = 30
    elif cvr >= 0.10:
        pts = 20
    elif cvr >= 0.05:
        pts = 10
    else:
        pts = 0
    score += pts
    breakdown["cvr"] = (cvr, pts)

    # 2. CTR (20% weight)
    ctr = float(row.get("_ctr", 0))
    if ctr >= 0.005:
        pts = 20
    elif ctr >= 0.003:
        pts = 15
    elif ctr >= 0.001:
        pts = 10
    else:
        pts = 0
    score += pts
    breakdown["ctr"] = (ctr, pts)

    # 3. Top of Search Impression Share (20% weight)
    tos = float(row.get("_tos_is", 0))
    if tos >= 0.10:
        pts = 20
    elif tos >= 0.05:
        pts = 15
    elif tos >= 0.01:
        pts = 10
    else:
        pts = 0
    score += pts
    breakdown["tos_is"] = (tos, pts)

    # 4. Sales Consistency (15% weight) - orders / assumed 14-day window
    orders = float(row.get("Orders", 0))
    daily_orders = orders / 14.0  # assume 14-day reporting window
    if daily_orders >= 2.0:
        pts = 15
    elif daily_orders >= 1.0:
        pts = 12
    elif daily_orders >= 0.5:
        pts = 8
    elif daily_orders > 0:
        pts = 4
    else:
        pts = 0
    score += pts
    breakdown["sales_consistency"] = (daily_orders, pts)

    # 5. Budget Utilization (15% weight)
    cname = str(row.get("Campaign Name", ""))
    budget = camps_budget_map.get(cname, 0)
    spend = float(row.get("Spend", 0))
    if budget > 0:
        util = spend / budget
    else:
        util = 0
    if 0.60 <= util <= 0.90:
        pts = 15
    elif util > 0.90:
        pts = 10  # throttled
    elif util > 0:
        pts = 5
    else:
        pts = 0
    score += pts
    breakdown["budget_util"] = (util, pts)

    return score, breakdown


def assign_status(score):
    if score >= 70:
        return "RANKING READY"
    elif score >= 50:
        return "PROGRESSING"
    elif score >= 30:
        return "STRUGGLING"
    else:
        return "NOT RANKING"


def status_color(status):
    return {
        "RANKING READY": C["positive"],
        "PROGRESSING":   C["neutral"],
        "STRUGGLING":    C["header_red"],
        "NOT RANKING":   C["header_pause"],
    }.get(status, C["header_gray"])


def check_bait_switch(row):
    """Check if keyword is ready for bait & switch test."""
    score = float(row.get("_rank_score", 0))
    orders = float(row.get("Orders", 0))
    tos = float(row.get("_tos_is", 0))

    ready = score >= 70 and orders >= 14 and tos > 0.05
    reasons = []
    if score < 70:
        reasons.append(f"Score {score:.0f} < 70")
    if orders < 14:
        reasons.append(f"Orders {orders:.0f} < 14")
    if tos <= 0.05:
        reasons.append(f"ToS IS {tos:.1%} <= 5%")

    if ready:
        instruction = ("BAIT_SWITCH_READY: Pause ranking campaign for 48-72 hours. "
                        "Monitor organic rank via Brand Analytics. "
                        "If organic rank holds, keyword is ranked. "
                        "If rank drops, resume campaign immediately.")
    else:
        instruction = "NOT READY: " + "; ".join(reasons)

    return ready, instruction


def daypart_recommendation(util):
    """Return day parting recommendation based on budget utilization."""
    if util > 0.90:
        return ("NEEDS DAY PARTING",
                "6AM-12PM: HIGH | 12PM-6PM: MEDIUM | 6PM-10PM: HIGH | 10PM-6AM: LOW")
    elif util > 0.70:
        return ("MONITOR",
                "Budget moderately utilised. Consider day parting if throttling starts.")
    else:
        return ("OK",
                "Budget headroom available. No day parting needed.")


def budget_adequacy(spend, budget):
    """Analyse budget adequacy and recommend adjustments."""
    if budget <= 0:
        return "UNKNOWN", 0, "No budget data available"
    util = spend / budget
    if util > 0.90:
        rec_budget = round((spend / util) * 1.3, 2)
        return "THROTTLED", rec_budget, \
            f"Spend/budget={util:.0%}. Campaign throttled. Increase to ${rec_budget:.2f}/day"
    elif util < 0.30:
        rec_budget = round(spend * 1.5, 2) if spend > 0 else budget
        return "UNDERSPENDING", rec_budget, \
            f"Spend/budget={util:.0%}. Check bids — may be too low to win auctions"
    else:
        return "ADEQUATE", budget, f"Spend/budget={util:.0%}. Budget is adequate"


def cluster_keywords(keywords):
    """Group keywords by shared root words (simple word overlap)."""
    groups = defaultdict(list)
    processed = set()

    kw_list = list(set(keywords))
    kw_words = {kw: set(kw.lower().split()) for kw in kw_list}

    for i, kw1 in enumerate(kw_list):
        if kw1 in processed:
            continue
        cluster = [kw1]
        processed.add(kw1)
        words1 = kw_words[kw1]

        for j, kw2 in enumerate(kw_list):
            if j <= i or kw2 in processed:
                continue
            words2 = kw_words[kw2]
            overlap = words1 & words2
            # Require at least one meaningful shared word (len > 2)
            meaningful = [w for w in overlap if len(w) > 2]
            if meaningful:
                cluster.append(kw2)
                processed.add(kw2)

        root = " + ".join(sorted(
            {w for w in kw_words[kw1] if len(w) > 2}
        )[:3]) or kw1
        groups[root] = cluster

    return dict(groups)


# -- Analysis Pipeline -------------------------------------------------------

def analyse(kws, camps_df, brand, sqp_df=None):
    """Run the full ranking analysis pipeline."""
    # Build campaign budget map
    camps_budget_map = {}
    for _, row in camps_df.iterrows():
        cname = str(row.get("Campaign Name", ""))
        budget = float(row.get("Budget", 0)) if "Budget" in camps_df.columns else 0
        camps_budget_map[cname] = budget

    # Identify ranking campaigns
    kws = identify_ranking_campaigns(kws, camps_df, brand)

    # Merge SQP data if available
    if sqp_df is not None:
        kws = merge_sqp(kws, sqp_df)
    else:
        kws["_sqp_organic_impression_share"] = 0.0
        kws["_sqp_organic_click_share"] = 0.0
        kws["_sqp_conversion_share"] = 0.0
        kws["_has_sqp"] = False

    # Filter to ranking keywords only
    ranking = kws[kws["is_ranking"]].copy()
    if ranking.empty:
        print("  WARNING: No ranking campaigns identified. Scoring all exact match keywords.")
        ranking = kws[kws["Match Type"].str.lower() == "exact"].copy()
    if ranking.empty:
        print("  WARNING: No exact match keywords found. Scoring all keywords.")
        ranking = kws.copy()

    # Calculate ranking scores
    scores = []
    breakdowns = []
    for _, row in ranking.iterrows():
        score, bd = calc_ranking_score(row, camps_budget_map)
        # SQP bonus: add up to 5 pts if organic signals present
        if row.get("_has_sqp", False):
            org_imp = float(row.get("_sqp_organic_impression_share", 0))
            if org_imp > 10:
                score = min(score + 5, 100)
                bd["sqp_bonus"] = (org_imp, 5)
            elif org_imp > 5:
                score = min(score + 3, 100)
                bd["sqp_bonus"] = (org_imp, 3)
        scores.append(score)
        breakdowns.append(bd)

    ranking["_rank_score"] = scores
    ranking["_rank_status"] = ranking["_rank_score"].apply(assign_status)
    ranking["_breakdown"] = breakdowns

    # Bait & switch check
    bs_results = []
    for _, row in ranking.iterrows():
        ready, instruction = check_bait_switch(row)
        bs_results.append({"_bs_ready": ready, "_bs_instruction": instruction})
    bs_df = pd.DataFrame(bs_results, index=ranking.index)
    ranking = pd.concat([ranking, bs_df], axis=1)

    # Budget and day parting
    budget_results = []
    for _, row in ranking.iterrows():
        cname = str(row.get("Campaign Name", ""))
        budget = camps_budget_map.get(cname, 0)
        spend = float(row.get("Spend", 0))
        util = spend / budget if budget > 0 else 0

        dp_status, dp_rec = daypart_recommendation(util)
        ba_status, ba_rec_budget, ba_detail = budget_adequacy(spend, budget)

        budget_results.append({
            "_budget":       budget,
            "_util":         util,
            "_dp_status":    dp_status,
            "_dp_rec":       dp_rec,
            "_ba_status":    ba_status,
            "_ba_rec_budget": ba_rec_budget,
            "_ba_detail":    ba_detail,
        })
    budget_df = pd.DataFrame(budget_results, index=ranking.index)
    ranking = pd.concat([ranking, budget_df], axis=1)

    return ranking


# -- Excel Helpers -----------------------------------------------------------

def apply_header(ws, row_num, headers, bg, fc="FFFFFF", height=20):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=i, value=h)
        c.font = hfont(color=fc)
        c.fill = hex_fill(bg)
        c.alignment = cal()
        c.border = thin_border()

def sheet_title(ws, title, subtitle, bg, height=28):
    ws.row_dimensions[1].height = height
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    max_col = max(ws.max_column or 1, 12)
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = hex_fill(bg)
        ws.cell(row=1, column=col).border = thin_border()
    ws.row_dimensions[2].height = 16
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    s.fill = hex_fill("FFFFFF")

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# -- Sheet Builders ----------------------------------------------------------

def build_dashboard(wb, ranking, brand, target_acos):
    ws = wb.create_sheet("Ranking Dashboard")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Amazon PPC Rank Tracker Dashboard"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hex_fill(C["header_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = (f"Brand: {brand}   |   Target ACoS: {target_acos:.0%}   |   "
                f"Generated: {datetime.today().strftime('%b %d, %Y')}")
    c2.font = Font(name=FONT, italic=True, size=9, color="888888")
    c2.fill = hex_fill("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # KPIs
    total_kws = len(ranking)
    total_spend = ranking["Spend"].sum()
    total_sales = ranking["Sales"].sum()
    total_orders = ranking["Orders"].sum()
    avg_score = ranking["_rank_score"].mean() if total_kws > 0 else 0
    bs_ready = int(ranking["_bs_ready"].sum())

    kpis = [
        ("Ranking Keywords", f"{total_kws}",             C["header_dark"]),
        ("Avg Rank Score",   f"{avg_score:.0f}/100",     C["header_green"] if avg_score >= 50 else C["header_red"]),
        ("Total Spend",      f"${total_spend:,.2f}",     C["header_dark"]),
        ("Total Sales",      f"${total_sales:,.2f}",     C["header_green"]),
        ("Total Orders",     f"{int(total_orders):,}",   C["header_dark"]),
        ("B&S Ready",        f"{bs_ready}",              C["header_purple"] if bs_ready > 0 else C["header_gray"]),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    for i, (label, value, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hex_fill("F5F5F5")
        lc.alignment = cal()
        lc.border = thin_border()
        vc = ws.cell(row=5, column=i, value=value)
        vc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = cal()
        vc.border = thin_border()

    # Status distribution
    ws.row_dimensions[7].height = 18
    apply_header(ws, 7,
                 ["Status", "Count", "% of Total", "Avg Score",
                  "Avg Spend", "Avg Orders", "Guidance"],
                 C["header_dark"])

    status_order = ["RANKING READY", "PROGRESSING", "STRUGGLING", "NOT RANKING"]
    guidance = {
        "RANKING READY": "Maintain spend. Test bait & switch for organic hold.",
        "PROGRESSING":   "Continue pushing. Increase budget if throttled.",
        "STRUGGLING":    "Review listing, pricing, keyword relevancy.",
        "NOT RANKING":   "Consider different approach or keyword change.",
    }
    for r_idx, status in enumerate(status_order):
        rn = 8 + r_idx
        ws.row_dimensions[rn].height = 18
        subset = ranking[ranking["_rank_status"] == status]
        count = len(subset)
        pct = count / total_kws if total_kws > 0 else 0
        avg_s = subset["_rank_score"].mean() if count > 0 else 0
        avg_sp = subset["Spend"].mean() if count > 0 else 0
        avg_ord = subset["Orders"].mean() if count > 0 else 0
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])

        vals = [status, count, pct, f"{avg_s:.0f}",
                avg_sp, avg_ord, guidance[status]]
        fmts = [None, None, "0.0%", None, '"$"#,##0.00', "0.0", None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = bfont(bold=(c_idx == 1))
            if c_idx == 1:
                cell.font = Font(name=FONT, bold=True, size=10,
                                 color=status_color(status))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val not in ("", None):
                cell.number_format = fmt

    ws.column_dimensions["G"].width = 52
    ws.freeze_panes = "A3"


def build_keyword_scores(wb, ranking):
    ws = wb.create_sheet("Keyword Ranking Scores")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Keyword Ranking Scores",
                f"{len(ranking)} keywords scored across 5 ranking dimensions",
                C["header_dark"])

    headers = ["Campaign", "Ad Group", "Keyword", "Match",
               "Score", "Status", "CVR", "CVR Pts", "CTR", "CTR Pts",
               "ToS IS", "ToS Pts", "Daily Orders", "Consistency Pts",
               "Budget Util", "Util Pts", "Impressions", "Clicks",
               "Spend", "Sales", "Orders", "Ranking Reason"]
    apply_header(ws, 3, headers, C["header_dark"])

    sorted_df = ranking.sort_values("_rank_score", ascending=False)

    for r_idx, (_, row) in enumerate(sorted_df.iterrows()):
        er = 4 + r_idx
        status = row.get("_rank_status", "")
        fill_color = {
            "RANKING READY": C["green_light"],
            "PROGRESSING":   C["amber_light"],
            "STRUGGLING":    C["red_light"],
            "NOT RANKING":   C["gray_light"],
        }.get(status, "FFFFFF")
        fill = hex_fill(fill_color) if r_idx % 2 == 0 else hex_fill("FFFFFF")
        ws.row_dimensions[er].height = 15

        bd = row.get("_breakdown", {})
        cvr_val, cvr_pts = bd.get("cvr", (0, 0))
        ctr_val, ctr_pts = bd.get("ctr", (0, 0))
        tos_val, tos_pts = bd.get("tos_is", (0, 0))
        sc_val, sc_pts = bd.get("sales_consistency", (0, 0))
        bu_val, bu_pts = bd.get("budget_util", (0, 0))

        vals = [
            row.get("Campaign Name", ""), row.get("Ad Group Name", ""),
            row.get("Keyword Text", ""), row.get("Match Type", ""),
            row.get("_rank_score", 0), status,
            cvr_val, cvr_pts, ctr_val, ctr_pts,
            tos_val, tos_pts, sc_val, sc_pts,
            bu_val, bu_pts,
            row.get("Impressions", 0), row.get("Clicks", 0),
            row.get("Spend", 0), row.get("Sales", 0),
            row.get("Orders", 0), row.get("ranking_reason", ""),
        ]
        fmts_list = [
            None, None, None, None, "0", None,
            "0.0%", "0", "0.00%", "0",
            "0.0%", "0", "0.0", "0",
            "0.0%", "0",
            "#,##0", "#,##0",
            '"$"#,##0.00', '"$"#,##0.00', "#,##0", None,
        ]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts_list), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            if c_idx == 6:  # Status column
                cell.font = Font(name=FONT, bold=True, size=9,
                                 color=status_color(status))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal() if c_idx <= 4 or c_idx == 22 else cal()
            if fmt and val != "":
                cell.number_format = fmt

    widths = {
        "A": 28, "B": 22, "C": 32, "D": 10, "E": 8, "F": 16,
        "G": 8, "H": 9, "I": 8, "J": 9, "K": 8, "L": 9,
        "M": 12, "N": 14, "O": 12, "P": 9, "Q": 12, "R": 10,
        "S": 12, "T": 12, "U": 10, "V": 36,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_bait_switch(wb, ranking):
    ws = wb.create_sheet("Bait Switch Candidates")
    ws.sheet_view.showGridLines = False

    candidates = ranking[ranking["_bs_ready"]].copy()
    not_ready = ranking[(ranking["_rank_score"] >= 50) & (~ranking["_bs_ready"])].copy()

    sheet_title(ws, "Bait & Switch Candidates",
                f"{len(candidates)} keywords ready | {len(not_ready)} approaching readiness",
                C["header_purple"])

    headers = ["Campaign", "Keyword", "Score", "Status", "Orders",
               "ToS IS", "Spend", "Sales", "CVR", "Instruction"]
    apply_header(ws, 3, headers, C["header_purple"])

    all_rows = pd.concat([candidates, not_ready]).sort_values(
        "_rank_score", ascending=False)

    for r_idx, (_, row) in enumerate(all_rows.iterrows()):
        er = 4 + r_idx
        is_ready = row.get("_bs_ready", False)
        fill = hex_fill(C["purple_light"]) if is_ready else hex_fill("FFFFFF")
        ws.row_dimensions[er].height = 15

        vals = [
            row.get("Campaign Name", ""), row.get("Keyword Text", ""),
            row.get("_rank_score", 0), row.get("_rank_status", ""),
            row.get("Orders", 0), row.get("_tos_is", 0),
            row.get("Spend", 0), row.get("Sales", 0),
            row.get("_cvr", 0), row.get("_bs_instruction", ""),
        ]
        fmts_list = [None, None, "0", None, "#,##0", "0.0%",
                     '"$"#,##0.00', '"$"#,##0.00', "0.0%", None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts_list), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            if is_ready and c_idx == 10:
                cell.font = Font(name=FONT, bold=True, size=9, color=C["positive"])
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal() if c_idx in (1, 2, 10) else cal()
            if fmt and val != "":
                cell.number_format = fmt

    widths = {
        "A": 30, "B": 32, "C": 8, "D": 16, "E": 10,
        "F": 10, "G": 12, "H": 12, "I": 8, "J": 60,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"
    return len(candidates)


def build_struggling(wb, ranking):
    ws = wb.create_sheet("Struggling Keywords")
    ws.sheet_view.showGridLines = False

    struggling = ranking[ranking["_rank_status"].isin(
        ["STRUGGLING", "NOT RANKING"])].copy()

    sheet_title(ws, "Struggling Keywords",
                f"{len(struggling)} keywords with weak ranking signals",
                C["header_red"])

    headers = ["Campaign", "Keyword", "Score", "Status",
               "CVR", "CTR", "ToS IS", "Impressions", "Clicks",
               "Spend", "Orders", "Issue", "Recommendation"]
    apply_header(ws, 3, headers, C["header_red"])

    struggling = struggling.sort_values("_rank_score", ascending=True)

    for r_idx, (_, row) in enumerate(struggling.iterrows()):
        er = 4 + r_idx
        fill = hex_fill(C["red_light"]) if r_idx % 2 == 0 else hex_fill("FFFFFF")
        ws.row_dimensions[er].height = 15

        # Diagnose primary issue
        cvr = float(row.get("_cvr", 0))
        ctr = float(row.get("_ctr", 0))
        imps = float(row.get("Impressions", 0))
        clicks = float(row.get("Clicks", 0))

        if imps < 100:
            issue = "Low impressions"
            rec = "Increase bid or check keyword relevancy"
        elif ctr < 0.001:
            issue = "Very low CTR"
            rec = "Review ad copy, main image, and title relevancy"
        elif cvr < 0.05:
            issue = "Low conversion rate"
            rec = "Review listing: price, reviews, images, A+ content"
        else:
            issue = "Low overall traction"
            rec = "Consider different keyword strategy or product fit"

        vals = [
            row.get("Campaign Name", ""), row.get("Keyword Text", ""),
            row.get("_rank_score", 0), row.get("_rank_status", ""),
            cvr, ctr, row.get("_tos_is", 0),
            row.get("Impressions", 0), row.get("Clicks", 0),
            row.get("Spend", 0), row.get("Orders", 0),
            issue, rec,
        ]
        fmts_list = [None, None, "0", None, "0.0%", "0.00%", "0.0%",
                     "#,##0", "#,##0", '"$"#,##0.00', "#,##0", None, None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts_list), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            if c_idx == 4:
                cell.font = Font(name=FONT, bold=True, size=9,
                                 color=C["negative"])
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal() if c_idx in (1, 2, 12, 13) else cal()
            if fmt and val != "":
                cell.number_format = fmt

    widths = {
        "A": 28, "B": 32, "C": 8, "D": 16, "E": 8, "F": 8, "G": 8,
        "H": 12, "I": 10, "J": 12, "K": 10, "L": 24, "M": 44,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"
    return len(struggling)


def build_dayparting(wb, ranking):
    ws = wb.create_sheet("Day Parting Recs")
    ws.sheet_view.showGridLines = False

    needs_dp = ranking[ranking["_dp_status"] == "NEEDS DAY PARTING"].copy()
    monitor = ranking[ranking["_dp_status"] == "MONITOR"].copy()
    all_dp = pd.concat([needs_dp, monitor])

    sheet_title(ws, "Day Parting Recommendations",
                f"{len(needs_dp)} campaigns need day parting | {len(monitor)} to monitor",
                C["header_amber"])

    headers = ["Campaign", "Keyword", "Budget", "Spend",
               "Utilisation", "Status", "Peak Hours Recommendation"]
    apply_header(ws, 3, headers, C["header_amber"])

    for r_idx, (_, row) in enumerate(all_dp.iterrows()):
        er = 4 + r_idx
        is_urgent = row.get("_dp_status", "") == "NEEDS DAY PARTING"
        fill = hex_fill(C["amber_light"]) if is_urgent else hex_fill("FFFFFF")
        ws.row_dimensions[er].height = 15

        vals = [
            row.get("Campaign Name", ""), row.get("Keyword Text", ""),
            row.get("_budget", 0), row.get("Spend", 0),
            row.get("_util", 0), row.get("_dp_status", ""),
            row.get("_dp_rec", ""),
        ]
        fmts_list = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                     "0.0%", None, None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts_list), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            if is_urgent and c_idx == 6:
                cell.font = Font(name=FONT, bold=True, size=9, color=C["negative"])
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal() if c_idx in (1, 2, 7) else cal()
            if fmt and val != "":
                cell.number_format = fmt

    widths = {
        "A": 30, "B": 32, "C": 12, "D": 12,
        "E": 12, "F": 20, "G": 60,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_budget_adequacy(wb, ranking):
    ws = wb.create_sheet("Budget Adequacy")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Budget Adequacy Analysis",
                "Campaign budget health and recommended adjustments",
                C["header_blue"])

    headers = ["Campaign", "Keyword", "Current Budget", "Spend",
               "Utilisation", "Status", "Recommended Budget", "Detail"]
    apply_header(ws, 3, headers, C["header_blue"])

    sorted_df = ranking.sort_values("_util", ascending=False)

    for r_idx, (_, row) in enumerate(sorted_df.iterrows()):
        er = 4 + r_idx
        ba_status = row.get("_ba_status", "")
        fill_map = {
            "THROTTLED":     C["red_light"],
            "UNDERSPENDING": C["amber_light"],
            "ADEQUATE":      C["green_light"],
        }
        fill = hex_fill(fill_map.get(ba_status, "FFFFFF")) if r_idx % 2 == 0 \
            else hex_fill("FFFFFF")
        ws.row_dimensions[er].height = 15

        vals = [
            row.get("Campaign Name", ""), row.get("Keyword Text", ""),
            row.get("_budget", 0), row.get("Spend", 0),
            row.get("_util", 0), ba_status,
            row.get("_ba_rec_budget", 0), row.get("_ba_detail", ""),
        ]
        fmts_list = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                     "0.0%", None, '"$"#,##0.00', None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts_list), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            if c_idx == 6:
                color = {"THROTTLED": C["negative"],
                         "UNDERSPENDING": C["neutral"],
                         "ADEQUATE": C["positive"]}.get(ba_status, "1A1A1A")
                cell.font = Font(name=FONT, bold=True, size=9, color=color)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal() if c_idx in (1, 2, 8) else cal()
            if fmt and val != "":
                cell.number_format = fmt

    widths = {
        "A": 30, "B": 32, "C": 14, "D": 12,
        "E": 12, "F": 16, "G": 18, "H": 52,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_keyword_groups(wb, ranking):
    ws = wb.create_sheet("Keyword Groups")
    ws.sheet_view.showGridLines = False

    keywords = ranking["Keyword Text"].dropna().unique().tolist()
    groups = cluster_keywords(keywords)

    sheet_title(ws, "Keyword Groups",
                f"{len(groups)} clusters from {len(keywords)} keywords",
                C["header_dark"])

    headers = ["Group Root", "Keywords in Group", "Avg Score",
               "Total Spend", "Total Sales", "Total Orders", "Keywords"]
    apply_header(ws, 3, headers, C["header_dark"])

    r_idx = 0
    for root, kw_list in sorted(groups.items(),
                                 key=lambda x: len(x[1]), reverse=True):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[er].height = 15

        subset = ranking[ranking["Keyword Text"].isin(kw_list)]
        avg_score = subset["_rank_score"].mean() if len(subset) > 0 else 0
        total_spend = subset["Spend"].sum() if len(subset) > 0 else 0
        total_sales = subset["Sales"].sum() if len(subset) > 0 else 0
        total_orders = subset["Orders"].sum() if len(subset) > 0 else 0
        kw_str = " | ".join(kw_list[:10])
        if len(kw_list) > 10:
            kw_str += f" ... +{len(kw_list) - 10} more"

        vals = [root, len(kw_list), avg_score,
                total_spend, total_sales, total_orders, kw_str]
        fmts_list = [None, "#,##0", "0", '"$"#,##0.00',
                     '"$"#,##0.00', "#,##0", None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts_list), 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal() if c_idx in (1, 7) else cal()
            if fmt and val != "":
                cell.number_format = fmt
        r_idx += 1

    widths = {
        "A": 24, "B": 18, "C": 12, "D": 14,
        "E": 14, "F": 14, "G": 60,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_action_plan(wb, ranking, brand):
    ws = wb.create_sheet("Ranking Action Plan")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Ranking Action Plan",
                f"Prioritised weekly actions for {brand}",
                C["header_green"])

    headers = ["Priority", "Action", "Keyword", "Campaign",
               "Current Score", "Current Status", "Detail", "Timeline"]
    apply_header(ws, 3, headers, C["header_green"])

    actions = []

    # 1. Bait & switch candidates
    bs_ready = ranking[ranking["_bs_ready"]]
    for _, row in bs_ready.iterrows():
        actions.append({
            "priority": "HIGH",
            "action": "TEST BAIT & SWITCH",
            "keyword": row.get("Keyword Text", ""),
            "campaign": row.get("Campaign Name", ""),
            "score": row.get("_rank_score", 0),
            "status": row.get("_rank_status", ""),
            "detail": "Pause ranking campaign 48-72h. Monitor organic rank.",
            "timeline": "This Week",
        })

    # 2. Throttled campaigns
    throttled = ranking[ranking["_ba_status"] == "THROTTLED"]
    for _, row in throttled.iterrows():
        actions.append({
            "priority": "HIGH",
            "action": "INCREASE BUDGET",
            "keyword": row.get("Keyword Text", ""),
            "campaign": row.get("Campaign Name", ""),
            "score": row.get("_rank_score", 0),
            "status": row.get("_rank_status", ""),
            "detail": f"Budget throttled. Increase to ${row.get('_ba_rec_budget', 0):.2f}/day",
            "timeline": "Immediate",
        })

    # 3. Progressing keywords to push
    progressing = ranking[ranking["_rank_status"] == "PROGRESSING"]
    for _, row in progressing.nlargest(10, "_rank_score").iterrows():
        actions.append({
            "priority": "MEDIUM",
            "action": "CONTINUE PUSH",
            "keyword": row.get("Keyword Text", ""),
            "campaign": row.get("Campaign Name", ""),
            "score": row.get("_rank_score", 0),
            "status": row.get("_rank_status", ""),
            "detail": f"Score {row.get('_rank_score', 0):.0f} — getting traction. "
                      f"Maintain or slightly increase bids.",
            "timeline": "This Week",
        })

    # 4. Struggling keywords to review
    struggling = ranking[ranking["_rank_status"] == "STRUGGLING"]
    for _, row in struggling.nlargest(10, "Spend").iterrows():
        actions.append({
            "priority": "LOW",
            "action": "REVIEW STRATEGY",
            "keyword": row.get("Keyword Text", ""),
            "campaign": row.get("Campaign Name", ""),
            "score": row.get("_rank_score", 0),
            "status": row.get("_rank_status", ""),
            "detail": "Weak signals. Review listing quality, pricing, and relevancy.",
            "timeline": "Next Week",
        })

    # 5. Not ranking — consider dropping
    not_ranking = ranking[ranking["_rank_status"] == "NOT RANKING"]
    for _, row in not_ranking.nlargest(5, "Spend").iterrows():
        actions.append({
            "priority": "LOW",
            "action": "CONSIDER DROPPING",
            "keyword": row.get("Keyword Text", ""),
            "campaign": row.get("Campaign Name", ""),
            "score": row.get("_rank_score", 0),
            "status": row.get("_rank_status", ""),
            "detail": f"Score {row.get('_rank_score', 0):.0f}. No traction. "
                      f"Reallocate budget to performing keywords.",
            "timeline": "Next Week",
        })

    priority_order = {"Immediate": 0, "This Week": 1, "Next Week": 2}
    actions.sort(key=lambda x: (
        {"HIGH": 0, "MEDIUM": 1, "LOW": 2}.get(x["priority"], 3),
        priority_order.get(x["timeline"], 3)
    ))

    for r_idx, act in enumerate(actions):
        er = 4 + r_idx
        priority = act["priority"]
        fill_map = {
            "HIGH":   C["red_light"],
            "MEDIUM": C["amber_light"],
            "LOW":    C["light_gray"],
        }
        fill = hex_fill(fill_map.get(priority, "FFFFFF"))
        ws.row_dimensions[er].height = 15

        vals = [act["priority"], act["action"], act["keyword"],
                act["campaign"], act["score"], act["status"],
                act["detail"], act["timeline"]]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            if c_idx == 1:
                color = {"HIGH": C["negative"], "MEDIUM": C["neutral"],
                         "LOW": C["header_gray"]}.get(priority, "1A1A1A")
                cell.font = Font(name=FONT, bold=True, size=9, color=color)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal() if c_idx in (2, 3, 4, 7) else cal()

    widths = {
        "A": 12, "B": 22, "C": 32, "D": 28,
        "E": 14, "F": 16, "G": 56, "H": 14,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"

    return actions


def build_raw_sheet(wb, ranking):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "All Ranking Keywords - Full Data",
                "Every ranking keyword with all metrics and scores",
                C["header_dark"])

    all_cols = ["Portfolio Name", "Campaign Name", "Ad Group Name",
                "Keyword Text", "Match Type", "State",
                "Bid", "Impressions", "Clicks", "Spend", "Sales",
                "Orders", "_acos", "_ctr", "_cvr", "_tos_is",
                "_rank_score", "_rank_status", "is_ranking",
                "ranking_reason", "_bs_ready", "_bs_instruction",
                "_budget", "_util", "_ba_status"]
    all_hdrs = ["Portfolio", "Campaign", "Ad Group",
                "Keyword", "Match", "State",
                "Bid", "Impressions", "Clicks", "Spend", "Sales",
                "Orders", "ACoS", "CTR", "CVR", "ToS IS",
                "Rank Score", "Rank Status", "Is Ranking",
                "Ranking Reason", "B&S Ready", "B&S Instruction",
                "Budget", "Utilisation", "Budget Status"]
    all_fmts = {
        "Bid":    '"$"#,##0.00', "Spend":  '"$"#,##0.00',
        "Sales":  '"$"#,##0.00', "_acos":  "0.0%",
        "_ctr":   "0.00%",       "_cvr":   "0.0%",
        "_tos_is": "0.0%",       "_budget": '"$"#,##0.00',
        "_util":  "0.0%",
    }
    status_bg = {
        "RANKING READY": C["green_light"],
        "PROGRESSING":   C["amber_light"],
        "STRUGGLING":    C["red_light"],
        "NOT RANKING":   C["gray_light"],
    }

    apply_header(ws, 3, all_hdrs, C["header_dark"])
    sorted_df = ranking.sort_values("_rank_score", ascending=False)

    for r_idx, (_, row) in enumerate(sorted_df.iterrows()):
        er = 4 + r_idx
        status = row.get("_rank_status", "")
        fill = hex_fill(status_bg.get(status, "FFFFFF"))
        ws.row_dimensions[er].height = 14
        for c_idx, col in enumerate(all_cols, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in all_fmts and val != "":
                cell.number_format = all_fmts[col]

    for i, w in enumerate(
        [20, 28, 22, 32, 10, 10, 12, 12, 10, 12, 12,
         10, 10, 10, 10, 10, 10, 16, 10, 34, 10, 50,
         12, 10, 14], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# -- Main -------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Amazon PPC Rank Tracker")
    parser.add_argument("--bulk",        required=True,
                        help="Path to Sponsored Products bulk file")
    parser.add_argument("--sqp",         default=None,
                        help="Path to Search Query Performance file (optional)")
    parser.add_argument("--output",      required=True,
                        help="Output Excel file path")
    parser.add_argument("--target-acos", type=float, default=0.60,
                        help="Target ACoS for ranking campaigns (default 0.60)")
    parser.add_argument("--brand",       required=True,
                        help="Brand name for filtering and labelling")
    parser.add_argument("--asin",        default=None,
                        help="Product ASIN for focused analysis (optional)")
    args = parser.parse_args()

    print(f"Loading bulk file: {args.bulk}")
    raw_df = load_bulk(args.bulk)
    camps_df = extract_campaigns(raw_df)
    kws = extract_keywords(raw_df)
    print(f"  {len(kws)} keyword rows loaded")
    print(f"  {len(camps_df)} campaign rows loaded")

    # Load SQP if provided
    sqp_df = None
    if args.sqp:
        print(f"\nLoading SQP file: {args.sqp}")
        sqp_df = load_sqp(args.sqp)
        print(f"  {len(sqp_df)} SQP rows loaded")

    print("\nRunning ranking analysis...")
    ranking = analyse(kws, camps_df, args.brand, sqp_df)

    # Print summary
    counts = ranking["_rank_status"].value_counts()
    print("\nRanking Status Distribution:")
    for status, n in counts.items():
        print(f"  {status}: {n}")

    bs_count = int(ranking["_bs_ready"].sum())
    print(f"\n  Bait & Switch ready: {bs_count}")
    print(f"  Avg ranking score: {ranking['_rank_score'].mean():.1f}/100")

    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb, ranking, args.brand, args.target_acos)
    build_keyword_scores(wb, ranking)
    bs_n = build_bait_switch(wb, ranking)
    struggle_n = build_struggling(wb, ranking)
    build_dayparting(wb, ranking)
    build_budget_adequacy(wb, ranking)
    build_keyword_groups(wb, ranking)
    actions = build_action_plan(wb, ranking, args.brand)
    build_raw_sheet(wb, ranking)

    # -- Write findings JSON -------------------------------------------------
    top_kws = ranking.nlargest(5, "_rank_score")
    bottom_kws = ranking.nsmallest(5, "_rank_score")
    findings = {
        "tool":              "rank_tracker",
        "brand":             args.brand,
        "asin":              args.asin,
        "target_acos":       args.target_acos,
        "total_ranking_kws": int(len(ranking)),
        "avg_rank_score":    round(float(ranking["_rank_score"].mean()), 1),
        "status_distribution": {
            status: int(count)
            for status, count in ranking["_rank_status"].value_counts().items()
        },
        "bait_switch_ready": bs_count,
        "struggling_count":  int(struggle_n),
        "throttled_campaigns": int(
            (ranking["_ba_status"] == "THROTTLED").sum()),
        "actions": [
            {"priority": a["priority"], "type": a["action"],
             "subject": a["keyword"], "campaign": a["campaign"],
             "detail": a["detail"], "timeline": a["timeline"]}
            for a in actions[:20]
        ],
        "top_keywords": [
            {"keyword": r.get("Keyword Text", ""),
             "score": int(r.get("_rank_score", 0)),
             "status": r.get("_rank_status", ""),
             "orders": int(r.get("Orders", 0)),
             "spend": round(float(r.get("Spend", 0)), 2)}
            for _, r in top_kws.iterrows()
        ],
        "bottom_keywords": [
            {"keyword": r.get("Keyword Text", ""),
             "score": int(r.get("_rank_score", 0)),
             "status": r.get("_rank_status", ""),
             "orders": int(r.get("Orders", 0)),
             "spend": round(float(r.get("Spend", 0)), 2)}
            for _, r in bottom_kws.iterrows()
        ],
    }
    findings_path = args.output.replace(".xlsx", "_findings.json")
    with open(findings_path, "w") as _f:
        json.dump(findings, _f, indent=2, default=str)
    print(f"   Findings: {findings_path}")

    wb.save(args.output)
    print(f"\nSaved: {args.output}")
    print(f"   Ranking keywords: {len(ranking)}")
    print(f"   Bait & Switch ready: {bs_n}")
    print(f"   Struggling: {struggle_n}")
    print(f"   Action items: {len(actions)}")


if __name__ == "__main__":
    main()
