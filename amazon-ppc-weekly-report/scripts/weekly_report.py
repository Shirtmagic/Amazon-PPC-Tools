#!/usr/bin/env python3
"""
Amazon PPC Weekly Report Generator
Combines Bulk Operations file + Search Term Report into a
single executive dashboard Excel file.
"""

import argparse
import glob
import json
import os
import sys
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import warnings
warnings.filterwarnings("ignore")

C = {
    "navy":        "1F3864",
    "green":       "1E6B3C",
    "red":         "8B0000",
    "amber":       "7B4F00",
    "blue":        "1A4A7A",
    "green_light": "F0FFF4",
    "red_light":   "FFF0F0",
    "amber_light": "FFFBF0",
    "gray_light":  "F5F5F5",
    "mid_gray":    "D9D9D9",
    "positive":    "1E6B3C",
    "negative":    "8B0000",
    "neutral":     "7B4F00",
    "white":       "FFFFFF",
}
FONT = "Arial"

def hx(h):  return PatternFill("solid", start_color=h, end_color=h)
def tb():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)
def hf(sz=10, bold=True, col="FFFFFF"):  return Font(name=FONT, bold=bold, size=sz, color=col)
def bf(sz=10, bold=False, col="1A1A1A"): return Font(name=FONT, bold=bold, size=sz, color=col)
def ca(): return Alignment(horizontal="center", vertical="center")
def la(): return Alignment(horizontal="left",   vertical="center")
def ra(): return Alignment(horizontal="right",  vertical="center")

def header_row(ws, r, cols, bg, fc="FFFFFF", h=20):
    ws.row_dimensions[r].height = h
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=r, column=i, value=c)
        cell.font = hf(col=fc)
        cell.fill = hx(bg)
        cell.alignment = ca()
        cell.border = tb()

def title_block(ws, title, subtitle, bg, h=30):
    ws.row_dimensions[1].height = h
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hx(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, 14):
        ws.cell(row=1, column=col).fill = hx(bg)
        ws.cell(row=1, column=col).border = tb()
    ws.row_dimensions[2].height = 16
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    s.fill = hx("F8F8F8")


# ── Loaders ──────────────────────────────────────────────────────────────────

def load_bulk(path):
    df = pd.read_excel(path, sheet_name="Sponsored Products Campaigns", engine="openpyxl")

    camps = df[df["Entity"] == "Campaign"].copy()
    camps["campaign"] = camps["Campaign Name"].fillna(
        camps.get("Campaign Name (Informational only)", ""))
    camps["portfolio"] = camps.get("Portfolio Name (Informational only)",
                                   pd.Series(dtype=str)).fillna("")
    for col in ["Daily Budget", "Spend", "Sales", "Orders", "Clicks",
                "Impressions", "ACOS", "CPC"]:
        if col in camps.columns:
            camps[col] = pd.to_numeric(camps[col], errors="coerce").fillna(0)
    camps["_acos"] = np.where(camps["Sales"] > 0, camps["Spend"] / camps["Sales"], np.nan)

    kws = df[df["Entity"] == "Keyword"].copy()
    for col in ["Bid", "Spend", "Sales", "Orders", "Clicks", "ACOS"]:
        if col in kws.columns:
            kws[col] = pd.to_numeric(kws[col], errors="coerce").fillna(0)
    kws["campaign"] = kws["Campaign Name (Informational only)"].fillna(
        kws.get("Campaign Name", ""))
    kws["portfolio"] = kws["Portfolio Name (Informational only)"].fillna("")
    kws["_acos"] = np.where(kws["Sales"] > 0, kws["Spend"] / kws["Sales"], np.nan)

    return camps, kws


def load_search_terms(path):
    df = pd.read_excel(path, engine="openpyxl")
    col_map = {}
    aliases = {
        "portfolio":   ["Portfolio name"],
        "campaign":    ["Campaign Name"],
        "ad_group":    ["Ad Group Name"],
        "match_type":  ["Match Type"],
        "search_term": ["Customer Search Term"],
        "impressions": ["Impressions"],
        "clicks":      ["Clicks"],
        "spend":       ["Spend"],
        "sales":       ["7 Day Total Sales ","7 Day Total Sales"],
        "orders":      ["7 Day Total Orders (#)"],
        "acos":        ["Total Advertising Cost of Sales (ACOS) ",
                        "Total Advertising Cost of Sales (ACOS)"],
    }
    for clean, options in aliases.items():
        for opt in options:
            matches = [c for c in df.columns if c.strip() == opt.strip()]
            if matches:
                col_map[matches[0]] = clean
                break
    df = df.rename(columns=col_map)
    for col in ["impressions", "clicks", "spend", "sales", "orders"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    df["_acos"] = np.where(df.get("sales", pd.Series([0]*len(df))) > 0,
                           df.get("spend", 0) / df.get("sales", 1), np.nan)
    return df


# ── KPI helpers ──────────────────────────────────────────────────────────────

def portfolio_kpis(camps, targets):
    grp = camps.groupby("portfolio").agg(
        campaigns=("campaign", "count"),
        budget=("Daily Budget", "sum"),
        spend=("Spend", "sum"),
        sales=("Sales", "sum"),
        orders=("Orders", "sum"),
        clicks=("Clicks", "sum"),
        impressions=("Impressions", "sum"),
    ).reset_index()
    grp["acos"]  = np.where(grp["sales"] > 0, grp["spend"] / grp["sales"], np.nan)
    grp["roas"]  = np.where(grp["spend"] > 0, grp["sales"] / grp["spend"], np.nan)
    grp["cpc"]   = np.where(grp["clicks"] > 0, grp["spend"] / grp["clicks"], np.nan)
    grp["cvr"]   = np.where(grp["clicks"] > 0, grp["orders"] / grp["clicks"], np.nan)
    grp["target_acos"] = grp["portfolio"].map(targets).fillna(
        targets.get("default", 0.25))
    grp["vs_target"] = grp["acos"] - grp["target_acos"]
    grp["status"] = grp.apply(lambda r:
        "✅ On Target" if not pd.isna(r["acos"]) and r["acos"] <= r["target_acos"] * 1.1
        else ("⚠ Slightly Over" if not pd.isna(r["acos"]) and r["acos"] <= r["target_acos"] * 1.5
        else "🔴 Over Target"), axis=1)
    return grp.sort_values("spend", ascending=False)


def top_bottom_campaigns(camps, n=5):
    active = camps[(camps["Orders"] > 0) & (camps["Spend"] > 20)].copy()
    active["_acos"] = active["Spend"] / active["Sales"].replace(0, np.nan)
    top    = active.nsmallest(n, "_acos")
    bottom = active.nlargest(n, "_acos")
    return top, bottom


def top_search_terms(st, n=10):
    if st is None:
        return pd.DataFrame()
    grp = st.groupby("search_term").agg(
        spend=("spend", "sum"),
        sales=("sales", "sum"),
        orders=("orders", "sum"),
        clicks=("clicks", "sum"),
    ).reset_index()
    grp["acos"] = np.where(grp["sales"] > 0, grp["spend"] / grp["sales"], np.nan)
    return grp[grp["orders"] > 0].nlargest(n, "orders")


def wasted_search_terms(st, min_spend=10, n=10):
    if st is None:
        return pd.DataFrame()
    grp = st.groupby("search_term").agg(
        spend=("spend", "sum"),
        orders=("orders", "sum"),
        clicks=("clicks", "sum"),
    ).reset_index()
    return grp[(grp["orders"] == 0) & (grp["spend"] >= min_spend)]\
        .nlargest(n, "spend")


# ── Sheet Builders ────────────────────────────────────────────────────────────

def build_executive_summary(wb, camps, kws, st, targets, brand, date_range):
    ws = wb.create_sheet("📋 Executive Summary")
    ws.sheet_view.showGridLines = False

    # ── Title ──
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = f"📊  Weekly PPC Report — {brand}"
    c.font = Font(name=FONT, bold=True, size=16, color="FFFFFF")
    c.fill = hx(C["navy"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:L2")
    c2 = ws["A2"]
    c2.value = (f"Report period: {date_range}   |   "
                f"Generated: {datetime.today().strftime('%B %d, %Y')}   |   "
                f"Target ACoS: {targets.get('default', 0.25):.0%}")
    c2.font = Font(name=FONT, italic=True, size=10, color="555555")
    c2.fill = hx("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Account-level KPIs ──
    total_spend   = camps["Spend"].sum()
    total_sales   = camps["Sales"].sum()
    total_orders  = int(camps["Orders"].sum())
    total_clicks  = int(camps["Clicks"].sum())
    total_impr    = int(camps["Impressions"].sum())
    overall_acos  = total_spend / total_sales if total_sales > 0 else 0
    overall_roas  = total_sales / total_spend if total_spend > 0 else 0
    overall_cvr   = total_orders / total_clicks if total_clicks > 0 else 0
    avg_cpc       = total_spend / total_clicks if total_clicks > 0 else 0
    target        = targets.get("default", 0.25)

    kpi_data = [
        ("Total Spend",    f"${total_spend:,.2f}",   C["navy"]),
        ("Total Sales",    f"${total_sales:,.2f}",   C["green"]),
        ("Overall ACoS",   f"{overall_acos:.1%}",
         C["negative"] if overall_acos > target * 1.1 else C["green"]),
        ("Overall ROAS",   f"{overall_roas:.1f}x",   C["navy"]),
        ("Total Orders",   f"{total_orders:,}",      C["navy"]),
        ("Total Clicks",   f"{total_clicks:,}",      C["navy"]),
        ("Avg CPC",        f"${avg_cpc:.2f}",        C["navy"]),
        ("Conv. Rate",     f"{overall_cvr:.1%}",     C["navy"]),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 26
    for i, (label, val, color) in enumerate(kpi_data, 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hx("F0F0F0")
        lc.alignment = ca()
        lc.border = tb()
        vc = ws.cell(row=5, column=i, value=val)
        vc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        vc.fill = hx(color)
        vc.alignment = ca()
        vc.border = tb()

    # ── Portfolio Performance Table ──
    ws.row_dimensions[7].height = 18
    port_hdrs = ["Portfolio", "Campaigns", "Spend", "Sales", "ACoS",
                 "Target ACoS", "vs Target", "ROAS", "Orders", "Status"]
    header_row(ws, 7, port_hdrs, C["navy"])

    port = portfolio_kpis(camps, targets)
    for r_idx, (_, row) in enumerate(port.iterrows()):
        rn = 8 + r_idx
        ws.row_dimensions[rn].height = 18
        status = row.get("status", "")
        bg = (C["green_light"] if "✅" in status
              else C["amber_light"] if "⚠" in status
              else C["red_light"])
        fill = hx(bg) if r_idx % 2 == 0 else hx("FFFFFF")

        acos_val   = row["acos"]   if not pd.isna(row.get("acos", np.nan))   else ""
        target_val = row["target_acos"]
        vs_val     = row["vs_target"] if not pd.isna(row.get("vs_target", np.nan)) else ""
        roas_val   = row["roas"]   if not pd.isna(row.get("roas", np.nan))   else ""

        vals = [row["portfolio"], int(row["campaigns"]),
                row["spend"], row["sales"],
                acos_val, target_val, vs_val,
                roas_val, int(row["orders"]), status]
        fmts = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                '0.0%', '0.0%', '+0.0%;-0.0%;0.0%', '0.0"x"', None, None]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = bfont = Font(name=FONT, size=10,
                                     bold=(c_idx == 1 or c_idx == 10))
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()
            if fmt and val not in ("", None):
                cell.number_format = fmt

        # Colour vs-target cell
        if vs_val != "":
            vc = ws.cell(row=rn, column=7)
            vc.font = Font(name=FONT, size=10, bold=True,
                           color=C["negative"] if float(vs_val) > 0 else C["positive"])

    # ── Top & Bottom Campaigns ──
    top_c, bot_c = top_bottom_campaigns(camps)
    offset = 8 + len(port) + 2

    ws.row_dimensions[offset].height = 18
    ws.merge_cells(f"A{offset}:E{offset}")
    tc = ws.cell(row=offset, column=1, value="🏆  Top 5 Campaigns — Lowest ACoS")
    tc.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    tc.fill = hx(C["green"])
    tc.alignment = la()
    for col in range(2, 6):
        ws.cell(row=offset, column=col).fill = hx(C["green"])

    mini_hdrs = ["Campaign", "Portfolio", "Spend", "Sales", "ACoS"]
    header_row(ws, offset + 1, mini_hdrs, C["green"])
    for r_idx, (_, row) in enumerate(top_c.iterrows()):
        rn = offset + 2 + r_idx
        ws.row_dimensions[rn].height = 16
        fill = hx("FFFFFF") if r_idx % 2 == 0 else hx(C["green_light"])
        for c_idx, (col, fmt) in enumerate(zip(
            ["campaign", "portfolio", "Spend", "Sales", "_acos"],
            [None, None, '"$"#,##0.00', '"$"#,##0.00', '0.0%']
        ), 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = ""
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=9)
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()
            if fmt and val not in ("", None): cell.number_format = fmt

    offset2 = offset
    ws.merge_cells(f"G{offset2}:K{offset2}")
    bc = ws.cell(row=offset2, column=7, value="⚠️  Bottom 5 Campaigns — Highest ACoS")
    bc.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    bc.fill = hx(C["red"])
    bc.alignment = la()
    for col in range(8, 12):
        ws.cell(row=offset2, column=col).fill = hx(C["red"])

    mini_hdrs2 = ["Campaign", "Portfolio", "Spend", "Sales", "ACoS"]
    for i, h in enumerate(mini_hdrs2, 7):
        cell = ws.cell(row=offset + 1, column=i, value=h)
        cell.font = hf()
        cell.fill = hx(C["red"])
        cell.alignment = ca()
        cell.border = tb()

    for r_idx, (_, row) in enumerate(bot_c.iterrows()):
        rn = offset + 2 + r_idx
        fill = hx("FFFFFF") if r_idx % 2 == 0 else hx(C["red_light"])
        for c_idx, (col, fmt) in enumerate(zip(
            ["campaign", "portfolio", "Spend", "Sales", "_acos"],
            [None, None, '"$"#,##0.00', '"$"#,##0.00', '0.0%']
        ), 7):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = ""
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=9)
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()
            if fmt and val not in ("", None): cell.number_format = fmt

    # Column widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 34
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 14
    ws.freeze_panes = "A3"


def build_portfolio_detail(wb, camps, kws, targets):
    ws = wb.create_sheet("📦 Portfolio Detail")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Portfolio-Level Performance Detail",
                "Campaign-by-campaign breakdown within each portfolio", C["navy"])

    hdrs = ["Portfolio", "Campaign", "Type", "State",
            "Daily Budget", "Avg Daily Spend", "Budget Util%",
            "Spend", "Sales", "Orders", "ACoS", "Target", "vs Target",
            "ROAS", "Clicks", "CPC", "CVR"]
    header_row(ws, 3, hdrs, C["navy"])

    camps_s = camps.sort_values(["portfolio", "Spend"], ascending=[True, False])
    target = targets.get("default", 0.25)

    for r_idx, (_, row) in enumerate(camps_s.iterrows()):
        rn   = 4 + r_idx
        ws.row_dimensions[rn].height = 15
        acos = row.get("_acos", np.nan)
        t    = targets.get(str(row.get("portfolio", "")), target)
        vs   = acos - t if not pd.isna(acos) else np.nan

        budget     = float(row.get("Daily Budget", 0))
        daily_sp   = float(row.get("Spend", 0)) / 30
        util       = daily_sp / budget if budget > 0 else np.nan

        ctype = "Auto" if "Auto" in str(row.get("campaign","")) else \
                "Exact" if "Exact" in str(row.get("campaign","")) else \
                "Broad" if "Broad" in str(row.get("campaign","")) else \
                "Phrase" if "Phrase" in str(row.get("campaign","")) else "Mixed"

        bg = (C["green_light"] if not pd.isna(acos) and acos <= t
              else C["red_light"]   if not pd.isna(acos) and acos > t * 1.5
              else C["amber_light"] if not pd.isna(acos)
              else "FFFFFF")
        fill = hx(bg) if r_idx % 2 == 0 else hx("FFFFFF")

        clicks  = float(row.get("Clicks", 0))
        orders  = float(row.get("Orders", 0))
        spend   = float(row.get("Spend", 0))
        sales   = float(row.get("Sales", 0))
        roas    = sales / spend if spend > 0 else np.nan
        cvr     = orders / clicks if clicks > 0 else np.nan
        cpc     = spend / clicks if clicks > 0 else np.nan

        vals = [row.get("portfolio",""), row.get("campaign",""),
                ctype, row.get("State",""),
                budget, daily_sp, util if not pd.isna(util) else "",
                spend, sales, int(orders),
                acos if not pd.isna(acos) else "",
                t,
                vs if not pd.isna(vs) else "",
                roas if not pd.isna(roas) else "",
                int(clicks), cpc if cpc > 0 else "",
                cvr if not pd.isna(cvr) else ""]
        fmts = [None, None, None, None,
                '"$"#,##0.00', '"$"#,##0.00', '0.0%',
                '"$"#,##0.00', '"$"#,##0.00', None,
                '0.0%', '0.0%', '+0.0%;-0.0%;0.0%',
                '0.0"x"', None, '"$"#,##0.00', '0.0%']

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=9)
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()
            if fmt and val not in ("", None): cell.number_format = fmt

        if vs != "" and not pd.isna(vs):
            vc = ws.cell(row=rn, column=13)
            vc.font = Font(name=FONT, size=9, bold=True,
                           color=C["negative"] if float(vs) > 0 else C["positive"])

    for i, w in enumerate([24,34,10,10,16,16,13,
                            14,14,10,10,10,13,10,10,12,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_keyword_analysis(wb, kws, targets):
    ws = wb.create_sheet("🔑 Keyword Analysis")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Keyword Performance Analysis",
                "All active keywords ranked by spend — colour coded vs target ACoS",
                C["navy"])

    hdrs = ["Portfolio", "Campaign", "Keyword", "Match Type",
            "Bid", "Clicks", "Spend", "Sales", "Orders",
            "ACoS", "ROAS", "CVR", "Status"]
    header_row(ws, 3, hdrs, C["navy"])

    target = targets.get("default", 0.25)
    active = kws[kws["Spend"] > 0].copy()
    active["_acos"] = np.where(active["Sales"] > 0,
                                active["Spend"] / active["Sales"], np.nan)
    active = active.sort_values("Spend", ascending=False)

    for r_idx, (_, row) in enumerate(active.iterrows()):
        rn  = 4 + r_idx
        ws.row_dimensions[rn].height = 14
        acos = row.get("_acos", np.nan)
        t    = targets.get(str(row.get("portfolio","")), target)
        status = ("✅ On Target" if not pd.isna(acos) and acos <= t
                  else "⚠ Slightly Over" if not pd.isna(acos) and acos <= t*1.5
                  else "🔴 Over Target" if not pd.isna(acos)
                  else "➖ No Sales")
        bg = (C["green_light"] if "✅" in status
              else C["amber_light"] if "⚠" in status
              else C["red_light"])
        fill = hx(bg) if r_idx % 2 == 0 else hx("FFFFFF")

        clicks = float(row.get("Clicks",0))
        spend  = float(row.get("Spend",0))
        sales  = float(row.get("Sales",0))
        orders = float(row.get("Orders",0))
        roas   = sales/spend if spend>0 else np.nan
        cvr    = orders/clicks if clicks>0 else np.nan

        vals = [row.get("portfolio",""), row.get("campaign",""),
                row.get("Keyword Text",""), row.get("Match Type",""),
                row.get("Bid",""),
                int(clicks), spend, sales, int(orders),
                acos if not pd.isna(acos) else "",
                roas if not pd.isna(roas) else "",
                cvr  if not pd.isna(cvr)  else "",
                status]
        fmts = [None,None,None,None,'"$"#,##0.00',
                None,'"$"#,##0.00','"$"#,##0.00',None,
                '0.0%','0.0"x"','0.0%',None]

        for c_idx,(val,fmt) in enumerate(zip(vals,fmts),1):
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=9)
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()
            if fmt and val not in ("",None): cell.number_format = fmt

    for i,w in enumerate([20,30,34,10,10,10,12,12,10,10,10,10,16],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_search_term_sheet(wb, st):
    if st is None:
        return
    ws = wb.create_sheet("🔍 Search Term Insights")
    ws.sheet_view.showGridLines = False

    top    = top_search_terms(st, n=15)
    wasted = wasted_search_terms(st, n=15)

    title_block(ws,
        "Search Term Insights — Top Converters & Wasted Spend",
        "Aggregated from the Search Term Report for this period",
        C["navy"])

    # Top converters
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:G3")
    tc = ws.cell(row=3, column=1, value="🏆  Top Converting Search Terms (by Orders)")
    tc.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    tc.fill = hx(C["green"])
    tc.alignment = la()
    for col in range(2,8):
        ws.cell(row=3,column=col).fill = hx(C["green"])

    header_row(ws, 4,
               ["Search Term","Orders","Spend","Sales","ACoS","Clicks","Action"],
               C["green"])

    for r_idx,(_, row) in enumerate(top.iterrows()):
        rn   = 5 + r_idx
        fill = hx("FFFFFF") if r_idx%2==0 else hx(C["green_light"])
        ws.row_dimensions[rn].height = 15
        acos = row.get("acos",np.nan)
        vals = [row.get("search_term",""), int(row.get("orders",0)),
                row.get("spend",0), row.get("sales",0),
                acos if not pd.isna(acos) else "",
                int(row.get("clicks",0)), "Consider adding as EXACT keyword"]
        fmts = [None,None,'"$"#,##0.00','"$"#,##0.00','0.0%',None,None]
        for c_idx,(val,fmt) in enumerate(zip(vals,fmts),1):
            cell = ws.cell(row=rn,column=c_idx,value=val)
            cell.font = Font(name=FONT,size=9)
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()
            if fmt and val not in ("",None): cell.number_format = fmt

    # Wasted spend
    offset = 5 + len(top) + 2
    ws.row_dimensions[offset].height = 18
    ws.merge_cells(f"A{offset}:G{offset}")
    wc = ws.cell(row=offset, column=1,
                 value="💸  Wasted Spend — High Spend, Zero Orders")
    wc.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    wc.fill = hx(C["red"])
    wc.alignment = la()
    for col in range(2,8):
        ws.cell(row=offset,column=col).fill = hx(C["red"])

    header_row(ws, offset+1,
               ["Search Term","Spend","Clicks","Orders","Action","",""],
               C["red"])

    for r_idx,(_, row) in enumerate(wasted.iterrows()):
        rn   = offset + 2 + r_idx
        fill = hx("FFFFFF") if r_idx%2==0 else hx(C["red_light"])
        ws.row_dimensions[rn].height = 15
        vals = [row.get("search_term",""),
                row.get("spend",0), int(row.get("clicks",0)),
                0, "Add as NEGATIVE EXACT","",""]
        fmts = [None,'"$"#,##0.00',None,None,None,None,None]
        for c_idx,(val,fmt) in enumerate(zip(vals,fmts),1):
            cell = ws.cell(row=rn,column=c_idx,value=val)
            cell.font = Font(name=FONT,size=9)
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()
            if fmt and val not in ("",None): cell.number_format = fmt

    for i,w in enumerate([44,12,14,14,12,10,28],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def build_action_checklist(wb, camps, kws, st, targets):
    ws = wb.create_sheet("✅ Action Checklist")
    ws.sheet_view.showGridLines = False

    title_block(ws,
        "Weekly Action Checklist",
        "Prioritised actions for this week — check off as you complete each one",
        C["navy"])

    target = targets.get("default", 0.25)

    actions = []

    # Budget-constrained campaigns
    camps["_daily"] = camps["Spend"] / 30
    camps["_util"]  = camps["_daily"] / camps["Daily Budget"].replace(0, np.nan)
    constrained = camps[(camps["_util"] >= 0.8) &
                        (camps["_acos"].fillna(1) <= target * 1.5)]
    for _, row in constrained.iterrows():
        actions.append(("🚨 Budget", "HIGH",
                        f"Increase budget for {row['campaign']} — hitting "
                        f"{row['_util']:.0%} of cap at "
                        f"{row.get('_acos',0):.1%} ACoS"))

    # Over-target high-spend keywords
    heavy_kws = kws[(kws["Spend"] > 100) &
                    (kws["_acos"].fillna(0) > target * 1.5)]
    for _, row in heavy_kws.head(5).iterrows():
        actions.append(("🎯 Bid", "HIGH",
                        f"Lower bid on '{row.get('Keyword Text','')}' "
                        f"({row.get('Match Type','')}) — ${row['Spend']:.0f} "
                        f"spend at {row['_acos']:.1%} ACoS"))

    # Great performers to scale
    winners = kws[(kws["_acos"].fillna(1) <= target * 0.7) &
                  (kws["Spend"] > 20) & (kws["Orders"] > 0)]
    for _, row in winners.head(3).iterrows():
        actions.append(("📈 Scale", "MEDIUM",
                        f"Raise bid on '{row.get('Keyword Text','')}' "
                        f"({row.get('Match Type','')}) — only "
                        f"{row['_acos']:.1%} ACoS on ${row['Spend']:.0f} spend"))

    # Wasted search term spend
    if st is not None:
        wasted = wasted_search_terms(st, min_spend=15, n=5)
        for _, row in wasted.iterrows():
            actions.append(("🔴 Negative", "HIGH",
                            f"Add '{row['search_term']}' as NEGATIVE EXACT — "
                            f"${row['spend']:.2f} spend, 0 orders"))

    # Zero spend campaigns
    dark = camps[(camps["Spend"] == 0) & (camps["State"] == "enabled")]
    for _, row in dark.iterrows():
        actions.append(("🚨 Alert", "HIGH",
                        f"Campaign '{row['campaign']}' is enabled but has 0 spend — "
                        f"check bids and targeting"))

    # Sort: HIGH first
    actions.sort(key=lambda x: 0 if x[1] == "HIGH" else 1)

    ws.row_dimensions[3].height = 18
    header_row(ws, 3, ["☐", "Category", "Priority", "Action Item"], C["navy"])
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 80

    priority_color = {"HIGH": C["negative"], "MEDIUM": C["neutral"], "LOW": C["green"]}

    for r_idx, (cat, priority, action) in enumerate(actions):
        rn   = 4 + r_idx
        ws.row_dimensions[rn].height = 20
        fill = hx("FFFFFF") if r_idx%2==0 else hx(C["gray_light"])
        for c_idx, val in enumerate(["☐", cat, priority, action], 1):
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=10,
                             color=priority_color.get(priority,"1A1A1A") if c_idx==3 else "1A1A1A",
                             bold=(c_idx==3))
            cell.fill = fill
            cell.border = tb()
            cell.alignment = la()

    ws.freeze_panes = "A4"


# ── Cross-Tool Master Action Plan ─────────────────────────────────────────────

def load_all_findings(output_path):
    """
    Auto-detect _findings.json files in the same directory as the output.
    When multiple findings exist for the same tool, use the most recently
    modified one (handles reruns cleanly).
    """
    out_dir = os.path.dirname(output_path)
    # Collect all findings grouped by tool
    by_tool = {}
    for path in glob.glob(os.path.join(out_dir, "*_findings.json")):
        try:
            with open(path) as f:
                data = json.load(f)
            tool = data.get("tool", os.path.basename(path))
            mtime = os.path.getmtime(path)
            if tool not in by_tool or mtime > by_tool[tool][1]:
                by_tool[tool] = (data, mtime, os.path.basename(path))
        except Exception:
            pass
    findings = {}
    for tool, (data, _, fname) in by_tool.items():
        findings[tool] = data
        print(f"  Loaded findings: {fname}")
    return findings


def build_master_action_plan(wb, findings, targets):
    """
    Creates a 🎯 Master Action Plan sheet by aggregating prioritised actions
    from all 4 tool findings JSONs, ranked by dollar impact.
    """
    ws = wb.create_sheet("🎯 Master Action Plan")
    ws.sheet_view.showGridLines = False

    # Title
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "🎯  Master Action Plan — Cross-Tool Priority Actions"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hx("1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, 10):
        ws.cell(row=1, column=col).fill = hx("1F3864")

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = (f"Aggregated from all 5 tools  |  Ranked by dollar impact  |  "
                 f"Target ACoS: {targets.get('default', 0.25):.0%}  |  "
                 f"Generated: {datetime.today().strftime('%B %d, %Y')}")
    sub.font = Font(name=FONT, italic=True, size=9, color="555555")
    sub.fill = hx("F8F8F8")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Stats bar from findings
    tool_labels = {
        "harvester":            "🌱 Harvest",
        "bid_optimizer":        "🎯 Bids",
        "budget_manager":       "💰 Budget",
        "placement_optimizer":  "📍 Placement",
    }
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 26
    col_i = 1
    for tool_key, label in tool_labels.items():
        if tool_key not in findings:
            continue
        fd = findings[tool_key]
        if tool_key == "harvester":
            stat = f"{fd.get('harvest_exact',0)+fd.get('harvest_phrase',0)} terms | ${fd.get('wasted_spend',0):,.0f} wasted"
        elif tool_key == "bid_optimizer":
            stat = f"{fd.get('raise_count',0)} raise | {fd.get('lower_count',0)} lower"
        elif tool_key == "budget_manager":
            stat = f"{fd.get('alert_count',0)} alerts | {fd.get('investigate_count',0)} investigate"
        elif tool_key == "placement_optimizer":
            stat = f"${fd.get('leakage_spend',0):,.0f} leakage | {fd.get('base_bid_reduce_count',0)} phantom"
        else:
            stat = "—"
        lc = ws.cell(row=4, column=col_i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hx("F0F0F0"); lc.alignment = Alignment(horizontal="center", vertical="center"); lc.border = tb()
        vc = ws.cell(row=5, column=col_i, value=stat)
        vc.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
        vc.fill = hx("1A4A7A"); vc.alignment = Alignment(horizontal="center", vertical="center"); vc.border = tb()
        ws.column_dimensions[get_column_letter(col_i)].width = 26
        col_i += 1

    # Collect and rank all actions
    all_actions = []
    for tool_key, fd in findings.items():
        for act in fd.get("actions", []):
            all_actions.append({
                "priority":     act.get("priority", "LOW"),
                "tool":         tool_labels.get(tool_key, tool_key),
                "type":         act.get("type", ""),
                "subject":      act.get("subject", ""),
                "campaign":     act.get("campaign", ""),
                "impact_spend": float(act.get("impact_spend", 0)),
                "detail":       act.get("detail", ""),
            })

    # Sort: HIGH first, then by dollar impact descending
    priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    all_actions.sort(key=lambda x: (priority_order.get(x["priority"], 3),
                                     -x["impact_spend"]))

    # Table header
    hdrs = ["Priority", "Tool", "Action Type", "Keyword / Campaign",
            "Portfolio / Campaign", "$ Impact", "What To Do", "Done?", "#"]
    bg_map = {"HIGH": "8B0000", "MEDIUM": "7B4F00", "LOW": "1E6B3C"}
    type_label = {
        "ADD_NEGATIVE":       "➕ Add Negative Keyword",
        "HARVEST_EXACT":      "🌱 Add Keyword (Exact)",
        "HARVEST_PHRASE":     "🌱 Add Keyword (Phrase)",
        "LOWER_BID":          "📉 Lower Bid",
        "RAISE_BID":          "📈 Raise Bid",
        "SUPPRESS_PLACEMENT": "🚫 Suppress Placement Modifier",
        "BASE_BID_REDUCE":    "⚠ Reduce Base Bid (Phantom Leak)",
        "ZERO_SPEND_ALERT":   "🚨 Zero Spend — Investigate",
        "BUDGET_CONSTRAINED": "💰 Increase Budget",
        "STRUCTURE_FIX":      "🏗 Fix Campaign Structure",
        "STRUCTURE_MONITOR":  "🔍 Monitor Campaign Purity",
    }
    priority_emoji = {"HIGH": "🔴 HIGH", "MEDIUM": "🟡 MEDIUM", "LOW": "🟢 LOW"}

    ws.row_dimensions[7].height = 20
    for i, h in enumerate(hdrs, 1):
        cell = ws.cell(row=7, column=i, value=h)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = hx("1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = tb()

    row_num = 8
    last_priority = None
    for idx, act in enumerate(all_actions, 1):
        pri = act["priority"]
        # Section divider when priority changes
        if pri != last_priority:
            ws.row_dimensions[row_num].height = 18
            ws.merge_cells(f"A{row_num}:I{row_num}")
            div = ws.cell(row=row_num, column=1,
                          value=f"{'🔴' if pri=='HIGH' else '🟡' if pri=='MEDIUM' else '🟢'}  {pri} PRIORITY ACTIONS")
            div.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
            div.fill = hx(bg_map.get(pri, "555555"))
            div.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            row_num += 1
            last_priority = pri

        bg = "FFF8F8" if pri == "HIGH" else "FFFDF0" if pri == "MEDIUM" else "F8FFF8"
        if idx % 2 == 0:
            bg = "FFFFFF"
        ws.row_dimensions[row_num].height = 18
        vals = [
            priority_emoji.get(pri, pri),
            act["tool"],
            type_label.get(act["type"], act["type"]),
            act["subject"][:50] if act["subject"] else "",
            act["campaign"][:50] if act["campaign"] else "",
            act["impact_spend"],
            act["detail"][:120] if act["detail"] else "",
            "☐",
            idx,
        ]
        fmts = [None, None, None, None, None, '"$"#,##0.00', None, None, None]
        for c_i, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=row_num, column=c_i, value=val)
            cell.font = Font(name=FONT, size=9)
            cell.fill = hx(bg)
            cell.border = tb()
            cell.alignment = Alignment(horizontal="left" if c_i not in (1,2,6,8,9) else "center",
                                       vertical="center", wrap_text=(c_i == 7))
            if fmt and val not in ("", None) and isinstance(val, (int, float)):
                cell.number_format = fmt
        row_num += 1

    if not all_actions:
        ws.row_dimensions[8].height = 30
        ws.merge_cells("A8:I8")
        nc = ws["A8"]
        nc.value = ("ℹ  No cross-tool findings available yet. "
                    "Run all 5 tools first, then rerun the Weekly Report — "
                    "findings JSONs will be auto-detected from the same folder.")
        nc.font = Font(name=FONT, italic=True, size=10, color="888888")
        nc.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    for i, w in enumerate([14, 16, 22, 44, 44, 12, 60, 8, 6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A8"

    print(f"  Master Action Plan: {len(all_actions)} actions from "
          f"{len(findings)} tools")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Amazon PPC Weekly Report")
    parser.add_argument("--bulk",         required=True, help="Bulk operations file")
    parser.add_argument("--search-terms", default=None,  help="Search term report (optional)")
    parser.add_argument("--output",       required=True)
    parser.add_argument("--brand",        default="My Brand")
    parser.add_argument("--target-acos",  type=float, default=0.25)
    parser.add_argument("--date-range",   default="")
    args = parser.parse_args()

    targets = {"default": args.target_acos}
    date_range = args.date_range or "Jan 25 – Feb 24, 2026"

    print(f"Loading bulk file: {args.bulk}")
    camps, kws = load_bulk(args.bulk)
    print(f"  {len(camps)} campaigns | {len(kws)} keywords")

    st = None
    if args.search_terms:
        print(f"Loading search term report: {args.search_terms}")
        st = load_search_terms(args.search_terms)
        print(f"  {len(st)} search term rows")

    # Auto-detect findings from other tools in same output directory
    print(f"\nScanning for cross-tool findings...")
    findings = load_all_findings(args.output)
    if not findings:
        print("  (none found — run all 5 tools first for Master Action Plan)")

    print(f"\nBuilding weekly report for {args.brand}...")
    wb = Workbook()
    wb.remove(wb.active)

    # Master Action Plan goes FIRST so it's the default sheet users see
    build_master_action_plan(wb, findings, targets)
    build_executive_summary(wb, camps, kws, st, targets, args.brand, date_range)
    build_portfolio_detail(wb, camps, kws, targets)
    build_keyword_analysis(wb, kws, targets)
    if st is not None:
        build_search_term_sheet(wb, st)
    build_action_checklist(wb, camps, kws, st, targets)

    wb.save(args.output)
    print(f"\n✅ Saved: {args.output}")
    print(f"   Portfolios: {camps['portfolio'].nunique()}")
    print(f"   Campaigns:  {len(camps)}")
    print(f"   Keywords:   {len(kws[kws['Spend']>0])} active")
    if st is not None:
        print(f"   Search terms: {len(st)}")


if __name__ == '__main__':
    main()
