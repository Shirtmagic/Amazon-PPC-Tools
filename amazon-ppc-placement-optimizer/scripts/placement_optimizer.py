#!/usr/bin/env python3
"""
Amazon PPC Placement Optimizer
Reads a Sponsored Products Bulk Operations file, evaluates placement-level
performance against campaign naming intent (ToS/RoS/PP), and produces a
bulk-upload-ready Excel file with bid modifier recommendations.
"""

import argparse
import json
import sys
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── Colours ──────────────────────────────────────────────────────────────────
C = {
    "header_dark":   "1F3864",
    "header_green":  "1E6B3C",
    "header_red":    "8B0000",
    "header_amber":  "7B4F00",
    "header_blue":   "1A4A7A",
    "header_purple": "4B0082",
    "header_gray":   "555555",
    "green_light":   "F0FFF4",
    "red_light":     "FFF0F0",
    "amber_light":   "FFFBF0",
    "blue_light":    "F0F8FF",
    "light_gray":    "F5F5F5",
    "positive":      "1E6B3C",
    "negative":      "8B0000",
    "neutral":       "7B4F00",
    "white":         "FFFFFF",
}
FONT = "Arial"

# Placement name normalisation
PLACEMENT_LABELS = {
    "Placement Top":              "Top of Search",
    "Placement Rest Of Search":   "Rest of Search",
    "Placement Product Page":     "Product Page",
    "Placement Amazon Business":  "Amazon Business",
}

# Which placement is PRIMARY for each campaign type
PRIMARY_PLACEMENT = {
    "ToS": "Placement Top",
    "RoS": "Placement Rest Of Search",
    "PP":  "Placement Product Page",
}

def hex_fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, bold=True, color="FFFFFF"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def bfont(size=10, bold=False, color="1A1A1A"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def cal():
    return Alignment(horizontal="center", vertical="center")

def lal():
    return Alignment(horizontal="left", vertical="center")


# ── Data Loading ──────────────────────────────────────────────────────────────

def camp_intent(name):
    """Determine campaign placement intent from name."""
    n = str(name).upper()
    if "TOS" in n or "| TOP" in n or "SKC TOS" in n:
        return "ToS"
    if "ROS" in n:
        return "RoS"
    if " PP " in n or n.endswith("| PP") or "| PP |" in n or n.endswith("PP"):
        return "PP"
    return "General"


def load_placements(path):
    df = pd.read_excel(path, sheet_name="Sponsored Products Campaigns",
                       engine="openpyxl")

    # ── Campaigns — get Campaign ID, bidding strategy, daily budget ──
    camps = df[df["Entity"] == "Campaign"][
        ["Campaign ID", "Campaign Name",
         "Campaign Name (Informational only)",
         "Portfolio Name (Informational only)",
         "Daily Budget", "Bidding Strategy", "State"]
    ].copy()
    camps["Campaign Name"] = camps["Campaign Name"].fillna(
        camps["Campaign Name (Informational only)"])
    camps["Portfolio Name"] = camps["Portfolio Name (Informational only)"]

    # ── Bidding Adjustment rows ──
    ba = df[df["Entity"] == "Bidding Adjustment"].copy()

    for col in ["Percentage", "Spend", "Sales", "Orders",
                "Impressions", "Clicks", "ACOS"]:
        if col in ba.columns:
            ba[col] = pd.to_numeric(ba[col], errors="coerce").fillna(0)

    ba["campaign"] = ba["Campaign Name (Informational only)"].fillna(
        ba.get("Campaign Name", ""))
    ba["portfolio"] = ba["Portfolio Name (Informational only)"].fillna("")
    ba["camp_id"]   = ba["Campaign ID"]

    ba["intent"]    = ba["campaign"].apply(camp_intent)
    ba["placement_label"] = ba["Placement"].map(PLACEMENT_LABELS).fillna(
        ba["Placement"])
    ba["is_primary"] = ba.apply(
        lambda r: r["Placement"] == PRIMARY_PLACEMENT.get(r["intent"], ""),
        axis=1
    )
    ba["acos_calc"] = np.where(
        (ba["Spend"] > 0) & (ba["Sales"] > 0),
        ba["Spend"] / ba["Sales"],
        np.nan
    )
    ba["cpc_calc"] = np.where(
        ba["Clicks"] > 0,
        ba["Spend"] / ba["Clicks"],
        np.nan
    )

    # ── Load keyword bids for base-bid calibration ──
    kws = df[df["Entity"] == "Keyword"].copy()
    kws["camp_id"] = kws["Campaign ID"].astype(str)
    kws["Bid"] = pd.to_numeric(kws.get("Bid", 0), errors="coerce").fillna(0)
    camp_avg_bid = kws.groupby("camp_id")["Bid"].mean().to_dict()
    ba["avg_kw_bid"] = ba["camp_id"].astype(str).map(camp_avg_bid).fillna(np.nan)

    return ba, camps


# ── Recommendation Engine ─────────────────────────────────────────────────────

def recommend(row, target_acos, min_clicks, max_increase, max_decrease,
              leakage_spend):
    """
    Returns (action, new_modifier, reason) for a single placement row.
    """
    current_mod = float(row.get("Percentage", 0))
    intent      = row.get("intent", "General")
    is_primary  = row.get("is_primary", False)
    spend       = float(row.get("Spend", 0))
    clicks      = float(row.get("Clicks", 0))
    orders      = float(row.get("Orders", 0))
    acos        = row.get("acos_calc", np.nan)
    placement   = row.get("Placement", "")

    # General campaigns — only suppress severe leakage
    if intent == "General":
        if (not pd.isna(acos) and acos > target_acos * 3
                and spend >= leakage_spend and clicks >= min_clicks):
            return ("SUPPRESS", 0,
                    f"General campaign: {acos:.1%} ACoS on ${spend:.0f} spend — suppress")
        return ("NO CHANGE", current_mod, "General campaign — no intent targeting")

    # Amazon Business — never adjust (B2B traffic, separate economics)
    if placement == "Placement Amazon Business":
        return ("NO CHANGE", current_mod, "Amazon Business placement — skip")

    # ── PRIMARY placement ──────────────────────────────────────────────────
    if is_primary:
        if spend == 0:
            return ("ALERT", current_mod,
                    f"⚠ PRIMARY placement not serving — 0 spend. "
                    f"Check bids are high enough to win {PLACEMENT_LABELS.get(placement,'')}")

        if clicks < min_clicks:
            return ("NO CHANGE", current_mod,
                    f"Only {int(clicks)} clicks on primary placement — need ≥{min_clicks}")

        if pd.isna(acos):
            return ("NO CHANGE", current_mod, "Spend but no sales data yet")

        band_low  = target_acos * 0.85
        band_high = target_acos * 1.15

        if acos <= band_low:
            # Scale up
            ratio   = target_acos / acos
            delta   = min((ratio - 1) * 50, max_increase)
            new_mod = min(round(current_mod + delta), 900)
            return ("RAISE", new_mod,
                    f"Primary {PLACEMENT_LABELS.get(placement,'')} ACoS {acos:.1%} "
                    f"well under target — raise modifier {current_mod:.0f}% → {new_mod:.0f}%")

        elif acos <= band_high:
            return ("NO CHANGE", current_mod,
                    f"Primary {PLACEMENT_LABELS.get(placement,'')} ACoS {acos:.1%} "
                    f"within target band — healthy")

        else:
            ratio   = target_acos / acos
            delta   = min((1 - ratio) * 100, max_decrease)
            new_mod = max(round(current_mod - delta), 0)
            return ("LOWER", new_mod,
                    f"Primary {PLACEMENT_LABELS.get(placement,'')} ACoS {acos:.1%} "
                    f"over target — lower modifier {current_mod:.0f}% → {new_mod:.0f}%")

    # ── NON-PRIMARY (leakage) placement ────────────────────────────────────
    if spend < leakage_spend:
        return ("NO CHANGE", current_mod,
                f"Low leakage spend ${spend:.2f} on non-primary — ignore")

    if pd.isna(acos):
        # Spend but no conversions at all
        if spend >= leakage_spend and clicks >= min_clicks:
            if current_mod == 0:
                return ("BASE_BID_REDUCE", 0,
                        f"⚠ PHANTOM LEAKAGE: modifier 0% but ${spend:.2f} on "
                        f"{PLACEMENT_LABELS.get(placement,'')} with 0 orders — reduce BASE BID")
            return ("SUPPRESS", 0,
                    f"LEAKAGE: ${spend:.2f} on non-primary {PLACEMENT_LABELS.get(placement,'')} "
                    f"with 0 orders — set to 0%")
        return ("NO CHANGE", current_mod, f"Minor leakage ${spend:.2f} — monitor")

    if acos > target_acos * 2:
        if current_mod == 0:
            # Modifier is already 0% but spend is still leaking — the BASE BID is too high
            return ("BASE_BID_REDUCE", 0,
                    f"⚠ PHANTOM LEAKAGE: modifier already 0% but ${spend:.2f} still going to "
                    f"{PLACEMENT_LABELS.get(placement,'')} at {acos:.1%} ACoS — "
                    f"reduce BASE BID via Bid Optimizer to stop this bleed")
        return ("SUPPRESS", 0,
                f"LEAKAGE: ${spend:.2f} on non-primary {PLACEMENT_LABELS.get(placement,'')} "
                f"at {acos:.1%} ACoS ({acos/target_acos:.1f}× target) — suppress to 0%")
    elif acos <= target_acos:
        return ("NO CHANGE", current_mod,
                f"Profitable leakage: ${spend:.2f} on {PLACEMENT_LABELS.get(placement,'')} "
                f"at {acos:.1%} ACoS — keep monitoring")
    else:
        ratio   = target_acos / acos
        delta   = min((1 - ratio) * 100, max_decrease)
        new_mod = max(round(current_mod - delta), 0)
        return ("LOWER", new_mod,
                f"Leakage ${spend:.2f} on non-primary {PLACEMENT_LABELS.get(placement,'')} "
                f"at {acos:.1%} — reduce modifier {current_mod:.0f}% → {new_mod:.0f}%")


def analyse(df, target_acos, min_clicks, max_increase, max_decrease,
            leakage_spend):
    actions, new_mods, reasons = [], [], []
    for _, row in df.iterrows():
        a, m, r = recommend(row, target_acos, min_clicks, max_increase,
                            max_decrease, leakage_spend)
        actions.append(a)
        new_mods.append(m)
        reasons.append(r)
    df = df.copy()
    df["action"]        = actions
    df["new_modifier"]  = new_mods
    df["reason"]        = reasons
    df["modifier_delta"] = df["new_modifier"] - df["Percentage"]
    return df


# ── Excel Helpers ─────────────────────────────────────────────────────────────

def apply_header(ws, row_num, headers, bg, fc="FFFFFF", height=20):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=i, value=h)
        c.font = hfont(color=fc)
        c.fill = hex_fill(bg)
        c.alignment = cal()
        c.border = thin_border()

def sheet_title(ws, title, subtitle, bg, height=28):
    ws.row_dimensions[1].height = height
    max_col = max(ws.max_column or 1, 12)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hex_fill(bg)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 16
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    s.fill = hex_fill("FFFFFF")

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# Standard columns for placement data sheets
PL_COLS = ["portfolio", "campaign", "intent", "is_primary",
           "placement_label", "Percentage", "Spend", "Sales",
           "Orders", "Clicks", "acos_calc",
           "action", "new_modifier", "modifier_delta", "reason"]
PL_HDRS = ["Portfolio", "Campaign", "Intent", "Primary?",
           "Placement", "Current Modifier %", "Spend", "Sales",
           "Orders", "Clicks", "ACoS",
           "Action", "New Modifier %", "Delta", "Reason"]
PL_FMTS = {
    "Spend":         '"$"#,##0.00',
    "Sales":         '"$"#,##0.00',
    "acos_calc":     '0.0%',
    "Percentage":    '0"%"',
    "new_modifier":  '0"%"',
    "modifier_delta":'+0;-0;0',
}

ACTION_BG = {
    "RAISE":     C["green_light"],
    "LOWER":     C["amber_light"],
    "SUPPRESS":  C["red_light"],
    "ALERT":     "FFE4E4",
    "NO CHANGE": "FFFFFF",
}
ACTION_COLOR = {
    "RAISE":    C["positive"],
    "LOWER":    C["neutral"],
    "SUPPRESS": C["negative"],
    "ALERT":    C["negative"],
}


def write_rows(ws, data, start_row, alt_fill_key="light_gray"):
    for r_idx, (_, row) in enumerate(data.iterrows()):
        er   = start_row + r_idx
        act  = row.get("action", "NO CHANGE")
        fill = hex_fill(ACTION_BG.get(act, "FFFFFF")) \
               if r_idx % 2 == 0 else hex_fill(C[alt_fill_key])
        ws.row_dimensions[er].height = 15

        for c_idx, col in enumerate(PL_COLS, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            if col == "is_primary":
                val = "✅ Primary" if val else "↘ Leakage"
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in PL_FMTS and val not in ("", None,
                                               "✅ Primary", "↘ Leakage"):
                cell.number_format = PL_FMTS[col]

        # Colour action + delta cells
        act_cell = ws.cell(row=er, column=12)
        act_cell.font = Font(name=FONT, bold=True, size=9,
                             color=ACTION_COLOR.get(act, "1A1A1A"))
        delta_cell = ws.cell(row=er, column=14)
        delta_val = row.get("modifier_delta", 0)
        if isinstance(delta_val, (int, float)) and not np.isnan(delta_val):
            delta_cell.font = Font(
                name=FONT, bold=True, size=9,
                color=C["positive"] if delta_val > 0 else
                      C["negative"] if delta_val < 0 else "888888")


# ── Sheet Builders ────────────────────────────────────────────────────────────

WIDTHS = {
    "A": 22, "B": 36, "C": 10, "D": 12, "E": 20,
    "F": 17, "G": 12, "H": 12, "I": 10, "J": 10,
    "K": 10, "L": 14, "M": 17, "N": 10, "O": 52,
}


def build_summary(wb, df, target_acos):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "📐  Amazon PPC Placement Optimizer Report"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hex_fill(C["header_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = (f"Target ACoS: {target_acos:.0%}   |   "
                f"Generated: {datetime.today().strftime('%b %d, %Y')}   |   "
                f"Campaign types: ToS = Top of Search | RoS = Rest of Search | PP = Product Page")
    c2.font = Font(name=FONT, italic=True, size=9, color="888888")
    c2.fill = hex_fill("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Action KPIs
    action_counts = df["action"].value_counts()
    leakage_df    = df[(~df["is_primary"]) & (df["action"] == "SUPPRESS")]
    leakage_spend = leakage_df["Spend"].sum()
    total_spend   = df["Spend"].sum()
    total_sales   = df["Sales"].sum()
    overall_acos  = total_spend / total_sales if total_sales > 0 else 0

    kpis = [
        ("Total Placement Spend", f"${total_spend:,.2f}",   C["header_dark"]),
        ("Overall ACoS",          f"{overall_acos:.1%}",
         C["negative"] if overall_acos > target_acos else C["header_green"]),
        ("Modifiers to Raise",    str(action_counts.get("RAISE", 0)),    C["header_green"]),
        ("Modifiers to Lower",    str(action_counts.get("LOWER", 0)),    C["neutral"]),
        ("Modifiers to Suppress", str(action_counts.get("SUPPRESS", 0)), C["negative"]),
        ("Leakage Spend Stopped", f"${leakage_spend:,.2f}",              C["negative"]),
        ("Alerts",                str(action_counts.get("ALERT", 0)),    C["header_amber"]),
        ("No Change",             str(action_counts.get("NO CHANGE", 0)), C["header_gray"]),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    for i, (label, value, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hex_fill("F5F5F5")
        lc.alignment = cal()
        lc.border = thin_border()
        vc = ws.cell(row=5, column=i, value=value)
        vc.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = cal()
        vc.border = thin_border()

    # Intent breakdown table
    ws.row_dimensions[7].height = 18
    apply_header(ws, 7,
                 ["Campaign Type", "Campaigns", "Primary Spend",
                  "Primary ACoS", "Leakage Spend", "Leakage ACoS",
                  "Raise", "Lower", "Suppress", "Alert"],
                 C["header_dark"])

    intents = ["ToS", "RoS", "PP", "General"]
    for r_idx, intent in enumerate(intents):
        rn      = 8 + r_idx
        subset  = df[df["intent"] == intent]
        primary = subset[subset["is_primary"]]
        leakage = subset[~subset["is_primary"] &
                         (subset["Placement"] != "Placement Amazon Business")]

        p_spend = primary["Spend"].sum()
        p_sales = primary["Sales"].sum()
        p_acos  = p_spend / p_sales if p_sales > 0 else np.nan
        l_spend = leakage["Spend"].sum()
        l_sales = leakage["Sales"].sum()
        l_acos  = l_spend / l_sales if l_sales > 0 else np.nan

        def ac(df_sub, a):
            return len(df_sub[df_sub["action"] == a])

        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[rn].height = 18
        vals = [
            {"ToS": "🔝 Top of Search", "RoS": "↔ Rest of Search",
             "PP": "📄 Product Page", "General": "⚙ General/Auto"}.get(intent, intent),
            subset["campaign"].nunique(),
            p_spend, p_acos if not np.isnan(p_acos) else "",
            l_spend, l_acos if not np.isnan(l_acos) else "",
            ac(subset, "RAISE"), ac(subset, "LOWER"),
            ac(subset, "SUPPRESS"), ac(subset, "ALERT"),
        ]
        fmts = [None, None, '"$"#,##0.00', '0.0%',
                '"$"#,##0.00', '0.0%', None, None, None, None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = bfont(bold=(c_idx == 1))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val not in ("", None):
                cell.number_format = fmt

    # Top leakage campaigns
    ws.row_dimensions[14].height = 18
    apply_header(ws, 14,
                 ["🚨 Top Placement Leakage — Campaigns Spending on Wrong Placements",
                  "", "Intent", "Leaking Into", "Spend", "ACoS", "Action"],
                 C["header_red"])

    leakage_all = df[
        (~df["is_primary"]) &
        (df["Spend"] >= 10) &
        (df["Placement"] != "Placement Amazon Business") &
        (df["intent"] != "General")
    ].sort_values("Spend", ascending=False).head(10)

    for r_idx, (_, row) in enumerate(leakage_all.iterrows()):
        rn   = 15 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["red_light"])
        ws.row_dimensions[rn].height = 16
        acos_val = row["acos_calc"] if not pd.isna(row.get("acos_calc", np.nan)) else ""
        vals = [row["campaign"], "", row["intent"],
                row["placement_label"], row["Spend"],
                acos_val, row["action"]]
        fmts = [None, None, None, None, '"$"#,##0.00', '0.0%', None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=c_idx, value=val)
            cell.font = bfont(
                bold=True if c_idx == 7 else False, size=9,
                color=ACTION_COLOR.get(row["action"], "1A1A1A") if c_idx == 7 else "1A1A1A"
            )
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val not in ("", None):
                cell.number_format = fmt

    ws.freeze_panes = "A3"


def build_primary_sheet(wb, df):
    data = df[df["is_primary"] & (df["intent"] != "General")].copy()
    data = data.sort_values(["intent", "acos_calc"])

    ws = wb.create_sheet("🎯 Primary Placement")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,
        "Primary Placement Performance — Each Campaign's Intended Placement",
        f"{len(data)} primary placements | Sorted by intent type then ACoS",
        C["header_blue"])
    apply_header(ws, 3, PL_HDRS, C["header_blue"])
    write_rows(ws, data, 4, "blue_light")
    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A4"


def build_leakage_sheet(wb, df):
    data = df[
        (~df["is_primary"]) &
        (df["intent"] != "General") &
        (df["Spend"] >= 5) &
        (df["Placement"] != "Placement Amazon Business")
    ].copy()
    data = data.sort_values("Spend", ascending=False)

    ws = wb.create_sheet("🚨 Placement Leakage")
    ws.sheet_view.showGridLines = False
    total_leakage = data["Spend"].sum()
    sheet_title(ws,
        f"Placement Leakage — ${total_leakage:,.2f} Spend on Non-Primary Placements",
        "These campaigns are spending on placements outside their intended focus",
        C["header_red"])
    apply_header(ws, 3, PL_HDRS, C["header_red"])
    write_rows(ws, data, 4, "red_light")
    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A4"


def build_changes_sheet(wb, df):
    data = df[df["action"].isin(["RAISE", "LOWER", "SUPPRESS", "ALERT"])].copy()
    data = data.sort_values(
        ["action", "Spend"],
        key=lambda x: x.map(
            {"SUPPRESS": 0, "ALERT": 1, "RAISE": 2, "LOWER": 3}
        ) if x.name == "action" else x,
        ascending=[True, False]
    )

    ws = wb.create_sheet("🔧 Modifier Changes")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,
        f"All Recommended Modifier Changes — {len(data)} adjustments",
        "Sorted by priority: Suppress → Alert → Raise → Lower",
        C["header_amber"])
    apply_header(ws, 3, PL_HDRS, C["header_amber"])
    write_rows(ws, data, 4, "amber_light")
    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A4"


def build_bulk_upload(wb, df, camps):
    """
    Produces Bidding Adjustment rows in Amazon bulk format.
    Only includes rows where action != NO CHANGE.
    """
    changes = df[df["action"].isin(["RAISE", "LOWER", "SUPPRESS"])].copy()

    ws = wb.create_sheet("📤 Amazon Bulk Upload")
    ws.sheet_view.showGridLines = False
    sheet_title(ws,
        f"Amazon Bulk Upload — {len(changes)} placement modifier updates",
        "⚠ Upload via Seller Central → Campaign Manager → Bulk Operations → Upload",
        C["header_dark"])

    headers = [
        "Product", "Entity", "Operation",
        "Campaign ID", "Campaign Name",
        "Bidding Strategy", "Placement",
        "Current Modifier %", "New Modifier %", "Delta",
        "Action", "Reason"
    ]
    apply_header(ws, 3, headers, C["header_dark"])

    # Build campaign ID lookup
    camp_id_map = {}
    if "Campaign ID" in camps.columns and "Campaign Name" in camps.columns:
        for _, row in camps.iterrows():
            cname = str(row.get("Campaign Name") or "")
            cid   = row.get("Campaign ID", "")
            if cname and cid:
                camp_id_map[cname] = cid

    # Also try to get bidding strategy per campaign
    strat_map = {}
    if "Bidding Strategy" in camps.columns:
        for _, row in camps.iterrows():
            cname = str(row.get("Campaign Name") or "")
            strat = row.get("Bidding Strategy", "")
            if cname and strat:
                strat_map[cname] = strat

    action_colors = {
        "RAISE":    C["positive"],
        "LOWER":    C["neutral"],
        "SUPPRESS": C["negative"],
    }

    for r_idx, (_, row) in enumerate(changes.iterrows()):
        er     = 4 + r_idx
        action = row.get("action", "")
        light  = {
            "RAISE":    C["green_light"],
            "LOWER":    C["amber_light"],
            "SUPPRESS": C["red_light"],
        }.get(action, "FFFFFF")
        fill = hex_fill(light) if r_idx % 2 == 0 else hex_fill("FFFFFF")

        cname  = row.get("campaign", "")
        cid    = camp_id_map.get(cname, "⬅ FILL IN")
        strat  = strat_map.get(cname, "Dynamic bids - down only")
        cur    = float(row.get("Percentage", 0))
        new    = float(row.get("new_modifier", cur))
        delta  = new - cur

        vals = {
            "Product":           "Sponsored Products",
            "Entity":            "Bidding Adjustment",
            "Operation":         "update",
            "Campaign ID":       cid,
            "Campaign Name":     cname,
            "Bidding Strategy":  strat,
            "Placement":         row.get("Placement", ""),
            "Current Modifier %": cur,
            "New Modifier %":    new,
            "Delta":             delta,
            "Action":            action,
            "Reason":            row.get("reason", ""),
        }
        fmts = {
            "Current Modifier %": '0"%"',
            "New Modifier %":     '0"%"',
            "Delta":              '+0;-0;0',
        }
        ws.row_dimensions[er].height = 15
        for c_idx, h in enumerate(headers, 1):
            val = vals.get(h, "")
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if h in fmts:
                cell.number_format = fmts[h]

        act_cell = ws.cell(row=er, column=11)
        act_cell.font = Font(name=FONT, bold=True, size=9,
                             color=action_colors.get(action, "1A1A1A"))

    for i, w in enumerate([18,18,12,18,36,26,26,18,16,10,14,52], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_raw_sheet(wb, df):
    ws = wb.create_sheet("🗂 Raw Data")
    ws.sheet_view.showGridLines = False
    sheet_title(ws, "All Placement Rows — Full Analysis",
                "Every bidding adjustment row with intent, performance, and recommendation",
                C["header_dark"])
    apply_header(ws, 3, PL_HDRS, C["header_dark"])
    df_sorted = df.sort_values(
        ["intent", "campaign", "is_primary", "Spend"],
        ascending=[True, True, False, False]
    )
    write_rows(ws, df_sorted, 4)
    set_widths(ws, WIDTHS)
    ws.freeze_panes = "A4"


# ── Placement Structure Analysis ──────────────────────────────────────────────

def placement_purity(df):
    """Per placement-intent campaign: what % of spend went to the intended placement."""
    intent_camps = df[df["intent"] != "General"].copy()
    if intent_camps.empty:
        return pd.DataFrame()

    rows = []
    for camp_id, grp in intent_camps.groupby("camp_id"):
        intent      = grp["intent"].iloc[0]
        camp_name   = grp["campaign"].iloc[0]
        portfolio   = grp["portfolio"].iloc[0]
        total_spend = grp["Spend"].sum()
        primary_row = grp[grp["is_primary"]]
        non_primary = grp[~grp["is_primary"]]
        primary_spend  = primary_row["Spend"].sum()
        primary_clicks = primary_row["Clicks"].sum()
        primary_mod    = primary_row["Percentage"].max() if len(primary_row) else 0
        primary_cpc    = (primary_row["cpc_calc"].mean()
                          if len(primary_row) and not primary_row["cpc_calc"].isna().all()
                          else np.nan)
        leak_spend     = non_primary[non_primary["Spend"] > 0]["Spend"].sum()
        avg_kw_bid     = grp["avg_kw_bid"].mean()
        purity         = (primary_spend / total_spend * 100) if total_spend > 0 else 0
        non_primary_mods       = non_primary["Percentage"].tolist()
        all_suppressed         = all(m == 0 for m in non_primary_mods) if non_primary_mods else True
        unsuppressed_placements = [
            PLACEMENT_LABELS.get(r["Placement"], r["Placement"])
            for _, r in non_primary.iterrows() if r["Percentage"] > 0
        ]
        # Ideal base bid: primary_CPC / (1 + modifier%) × 0.85 safety factor
        if not pd.isna(primary_cpc) and primary_mod > 0 and primary_cpc > 0:
            ideal_base = round(primary_cpc / (1 + primary_mod / 100) * 0.85, 2)
            ideal_base = max(ideal_base, 0.10)
        elif not pd.isna(avg_kw_bid) and primary_mod > 0:
            ideal_base = round(float(avg_kw_bid) * 0.60, 2)
        else:
            ideal_base = np.nan
        # Health assessment
        if purity >= 85 and all_suppressed:
            health = "✅ Healthy"
            rec    = "Campaign properly isolated — monitor purity weekly"
        elif not all_suppressed:
            health = "🔴 Fix Modifiers"
            rec    = (f"Non-primary modifiers not 0%: {', '.join(unsuppressed_placements)}. "
                      f"Suppress immediately to stop intent drift")
        elif purity < 70 and leak_spend > 0:
            health = "🔴 Base Bid Too High"
            bid_str = f"${float(avg_kw_bid):.2f}" if not pd.isna(avg_kw_bid) else "unknown"
            tgt_str = f"${float(ideal_base):.2f}" if not pd.isna(ideal_base) else "unknown"
            rec    = (f"Purity {purity:.0f}% — base bid {bid_str} wins non-primary even at 0% modifier. "
                      f"Target base bid: {tgt_str}")
        elif purity < 85:
            health = "⚠ Monitor"
            tgt_str = f"${float(ideal_base):.2f}" if not pd.isna(ideal_base) else "lower"
            rec    = f"Purity {purity:.0f}% — lower base bid toward {tgt_str} to improve isolation"
        else:
            health = "⚠ Insufficient Data"
            rec    = "Not enough spend data to evaluate fully"

        rows.append({
            "campaign":             camp_name,
            "portfolio":            portfolio,
            "intent":               intent,
            "total_spend":          round(total_spend, 2),
            "primary_spend":        round(primary_spend, 2),
            "leak_spend":           round(leak_spend, 2),
            "purity_score":         round(purity, 1),
            "primary_modifier_pct": primary_mod,
            "all_non_primary_at_0": all_suppressed,
            "unsuppressed":         ", ".join(unsuppressed_placements) if unsuppressed_placements else "—",
            "avg_kw_bid":           round(float(avg_kw_bid), 2) if not pd.isna(avg_kw_bid) else None,
            "ideal_base_bid":       float(ideal_base) if not pd.isna(ideal_base) else None,
            "health":               health,
            "recommendation":       rec,
        })

    return pd.DataFrame(rows).sort_values("purity_score")


def build_structure_audit(wb, df):
    """Sheet showing placement-intent campaign health, purity scores, base bid calibration."""
    ws = wb.create_sheet("🏗 Structure Audit")
    ws.sheet_view.showGridLines = False

    purity_df = placement_purity(df)

    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "🏗  Placement Campaign Structure Audit"
    c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill("1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, 14):
        ws.cell(row=1, column=col).fill = hex_fill("1F3864")

    ws.row_dimensions[2].height = 15
    ws.merge_cells("A2:M2")
    sub = ws["A2"]
    sub.value = ("Goal: ≥85% of spend at primary placement. Non-primary modifiers = 0%. "
                 "Base bid calibrated so 0% modifier truly stops spend at non-primary placements. "
                 "Ideal base bid = primary_CPC ÷ (1 + modifier%) × 0.85")
    sub.font = Font(name=FONT, italic=True, size=9, color="555555")
    sub.fill = hex_fill("F8F8F8")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    if purity_df.empty:
        ws.cell(row=4, column=1, value="No placement-intent campaigns found (need TOS/ROS/PP in campaign names).").font = Font(name=FONT, italic=True, size=10, color="888888")
        return purity_df

    # Score summary
    healthy  = (purity_df["health"] == "✅ Healthy").sum()
    monitor  = purity_df["health"].str.contains("Monitor|Insufficient", na=False).sum()
    critical = purity_df["health"].str.contains("🔴", na=False).sum()
    total_leak = purity_df["leak_spend"].sum()
    avg_purity = purity_df["purity_score"].mean()

    score_items = [
        ("✅ Healthy",      str(healthy),         "1E6B3C"),
        ("⚠ Monitor",       str(monitor),          "7B4F00"),
        ("🔴 Fix Now",      str(critical),         "8B0000"),
        ("Avg Purity",      f"{avg_purity:.0f}%",  "1A4A7A"),
        ("Total Leakage $", f"${total_leak:,.2f}",  "4A4A4A"),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 26
    for i, (label, val, color) in enumerate(score_items, 1):
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hex_fill("F0F0F0")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = thin_border()
        vc = ws.cell(row=5, column=i, value=val)
        vc.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = 18

    hdrs = ["Campaign", "Portfolio", "Intent", "Health",
            "Purity %", "Primary $", "Leakage $",
            "Primary Mod %", "Non-Primary = 0%",
            "Avg KW Bid", "Ideal Base Bid", "Bid Reduction?", "Action Required"]
    apply_header(ws, 7, hdrs, "1F3864")

    health_bg = {
        "✅ Healthy":          "F0FFF4",
        "⚠ Monitor":           "FFFBF0",
        "🔴 Fix Modifiers":   "FFF0F0",
        "🔴 Base Bid Too High":"FFF0F0",
        "⚠ Insufficient Data": "F5F5F5",
    }

    for r_idx, (_, row) in enumerate(purity_df.iterrows()):
        rn = 8 + r_idx
        ws.row_dimensions[rn].height = 22
        bg = health_bg.get(row["health"], "FFFFFF")
        avg_bid = row["avg_kw_bid"]
        ideal   = row["ideal_base_bid"]
        bid_red = ("✅ Yes — reduce to ${:.2f}".format(ideal)
                   if (ideal is not None and avg_bid is not None and ideal < avg_bid)
                   else "No change needed")
        vals = [
            row["campaign"], row["portfolio"], row["intent"], row["health"],
            row["purity_score"] / 100,
            row["primary_spend"], row["leak_spend"],
            row["primary_modifier_pct"] / 100,
            "✅ Yes" if row["all_non_primary_at_0"] else "❌ No",
            avg_bid, ideal,
            bid_red,
            row["recommendation"],
        ]
        fmts = [None, None, None, None, "0%", '"$"#,##0.00', '"$"#,##0.00',
                "0%", None, '"$"#,##0.00', '"$"#,##0.00', None, None]
        for c_i, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=rn, column=c_i, value=val)
            cell.font = Font(name=FONT, size=9,
                             bold=(c_i == 4),
                             color=("8B0000" if "🔴" in str(row["health"]) and c_i == 4
                                    else "1E6B3C" if "✅" in str(row["health"]) and c_i == 4
                                    else "1A1A1A"))
            cell.fill = hex_fill(bg)
            cell.border = thin_border()
            cell.alignment = Alignment(
                horizontal="center" if c_i in (3,4,5,6,7,8,9,10,11) else "left",
                vertical="center", wrap_text=(c_i == 13))
            if fmt and val not in (None, "—") and isinstance(val, (int, float)):
                cell.number_format = fmt

    for i, w in enumerate([46, 22, 10, 22, 10, 12, 12, 14, 18, 12, 14, 24, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A8"
    return purity_df


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Amazon PPC Placement Optimizer")
    parser.add_argument("--input",         required=True)
    parser.add_argument("--output",        required=True)
    parser.add_argument("--target-acos",   type=float, default=0.25)
    parser.add_argument("--min-clicks",    type=int,   default=10)
    parser.add_argument("--max-increase",  type=float, default=50)
    parser.add_argument("--max-decrease",  type=float, default=100)
    parser.add_argument("--leakage-spend", type=float, default=20)
    args = parser.parse_args()

    print(f"Loading: {args.input}")
    df, camps = load_placements(args.input)
    print(f"  {len(df)} placement rows | {df['campaign'].nunique()} campaigns")

    intent_counts = df[df["is_primary"]]["intent"].value_counts()
    print("  Intent breakdown:")
    for intent, n in intent_counts.items():
        print(f"    {intent}: {n} campaigns")

    print("\nAnalysing placements...")
    df = analyse(df, args.target_acos, args.min_clicks,
                 args.max_increase, args.max_decrease, args.leakage_spend)

    counts = df["action"].value_counts()
    print("\nResults:")
    for action, n in counts.items():
        print(f"  {action}: {n}")

    leakage_spend = df[
        (~df["is_primary"]) & (df["action"] == "SUPPRESS")
    ]["Spend"].sum()
    print(f"\n  Leakage spend suppressed: ${leakage_spend:,.2f}")

    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb, df, args.target_acos)
    purity_df = build_structure_audit(wb, df)  # NEW: placement purity analysis
    build_primary_sheet(wb, df)
    build_leakage_sheet(wb, df)
    build_changes_sheet(wb, df)
    build_bulk_upload(wb, df, camps)
    build_raw_sheet(wb, df)

    # ── Write cross-tool findings JSON ────────────────────────────────────
    base_bid_rows = df[df["action"] == "BASE_BID_REDUCE"]
    suppress_rows = df[df["action"] == "SUPPRESS"].nlargest(5, "Spend")
    # Purity-based structure issues
    structure_actions = []
    if not purity_df.empty:
        for _, pr in purity_df.iterrows():
            if "🔴" in str(pr["health"]):
                structure_actions.append({
                    "priority": "HIGH",
                    "type": "STRUCTURE_FIX",
                    "subject": pr["campaign"],
                    "campaign": pr["campaign"],
                    "impact_spend": round(float(pr["leak_spend"]), 2),
                    "detail": (f"{pr['health']} — Purity {pr['purity_score']:.0f}% "
                               f"(${pr['leak_spend']:.2f} leaking). {pr['recommendation']}")
                })
            elif "Monitor" in str(pr["health"]) and pr["leak_spend"] >= 10:
                structure_actions.append({
                    "priority": "MEDIUM",
                    "type": "STRUCTURE_MONITOR",
                    "subject": pr["campaign"],
                    "campaign": pr["campaign"],
                    "impact_spend": round(float(pr["leak_spend"]), 2),
                    "detail": (f"{pr['health']} — Purity {pr['purity_score']:.0f}% "
                               f"(${pr['leak_spend']:.2f} leaking). {pr['recommendation']}")
                })

    low_purity = purity_df[purity_df["purity_score"] < 70] if not purity_df.empty else pd.DataFrame()

    findings = {
        "tool":                 "placement_optimizer",
        "target_acos":          args.target_acos,
        "raise_count":          int((df["action"] == "RAISE").sum()),
        "lower_count":          int((df["action"] == "LOWER").sum()),
        "suppress_count":       int((df["action"] == "SUPPRESS").sum()),
        "base_bid_reduce_count":int((df["action"] == "BASE_BID_REDUCE").sum()),
        "leakage_spend":        round(float(df[~df["is_primary"] & df["action"].isin(["SUPPRESS","BASE_BID_REDUCE"])]["Spend"].sum()), 2),
        "low_purity_campaigns": int(len(low_purity)),
        "avg_purity_score":     round(float(purity_df["purity_score"].mean()), 1) if not purity_df.empty else 0,
        "base_bid_reduce_campaigns": [
            {"campaign": r.get("campaign",""), "placement": r.get("Placement",""),
             "spend": round(float(r.get("Spend",0)),2), "acos": round(float(r.get("acos_calc",0)),4)}
            for _, r in base_bid_rows.iterrows()
        ],
        "actions": (
            structure_actions
            +
            [{"priority": "HIGH",   "type": "BASE_BID_REDUCE",
              "subject":  r.get("campaign",""), "campaign": r.get("campaign",""),
              "impact_spend": round(float(r.get("Spend",0)),2),
              "detail": r.get("reason","")}
             for _, r in base_bid_rows.iterrows() if r.get("Spend",0) >= 10]
            +
            [{"priority": "HIGH",   "type": "SUPPRESS_PLACEMENT",
              "subject":  r.get("campaign",""), "campaign": r.get("campaign",""),
              "impact_spend": round(float(r.get("Spend",0)),2),
              "detail": r.get("reason","")}
             for _, r in suppress_rows.iterrows() if r.get("Spend",0) >= 20]
        ),
    }
    findings_path = args.output.replace(".xlsx", "_findings.json")
    with open(findings_path, "w") as _f:
        json.dump(findings, _f, indent=2, default=str)
    print(f"   Findings: {findings_path}")

    wb.save(args.output)
    changes = df[df["action"].isin(["RAISE", "LOWER", "SUPPRESS", "BASE_BID_REDUCE"])]
    print(f"\n✅ Saved: {args.output}")
    print(f"   🎯 Raise:    {(df['action']=='RAISE').sum()} modifiers")
    print(f"   🔧 Lower:    {(df['action']=='LOWER').sum()} modifiers")
    print(f"   🚫 Suppress: {(df['action']=='SUPPRESS').sum()} modifiers")
    print(f"   🔧 Base Bid Reduce: {(df['action']=='BASE_BID_REDUCE').sum()} campaigns (modifier already 0%)")
    print(f"   ⚠  Alerts:   {(df['action']=='ALERT').sum()}")
    print(f"   Bulk upload tab: {len(changes)} rows ready")


if __name__ == "__main__":
    main()
