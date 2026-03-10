#!/usr/bin/env python3
"""
Amazon PPC Single Keyword Campaign (SKC) Builder
Generates complete SKC campaign structures from harvested keywords or a manual
list, outputting a formatted Excel file ready for Amazon Bulk Upload.
"""

import argparse
import json
import sys
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")


# -- Colours -----------------------------------------------------------------
C = {
    "header_dark":   "1F3864",
    "header_green":  "1E6B3C",
    "header_red":    "8B0000",
    "header_amber":  "7B4F00",
    "header_pause":  "4A4A4A",
    "header_gray":   "555555",
    "green_light":   "F0FFF4",
    "red_light":     "FFF0F0",
    "amber_light":   "FFFBF0",
    "gray_light":    "F8F8F8",
    "light_gray":    "F5F5F5",
    "positive":      "1E6B3C",
    "negative":      "8B0000",
    "neutral":       "7B4F00",
    "white":         "FFFFFF",
}
FONT = "Arial"


def hex_fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, bold=True, color="FFFFFF"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def bfont(size=10, bold=False, color="1A1A1A"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def cal():
    return Alignment(horizontal="center", vertical="center")

def lal():
    return Alignment(horizontal="left", vertical="center")


# -- Goal Configuration ------------------------------------------------------

GOAL_CONFIG = {
    "ranking": {
        "label":            "Ranking",
        "starting_bid":     0.50,
        "daily_budget":     25.00,
        "tos_modifier":     100,
        "pp_modifier":      0,
        "bidding_strategy":  "Fixed bids",
        "description":      "Aggressive Top-of-Search push for organic rank gains",
    },
    "profit": {
        "label":            "Profit",
        "starting_bid":     0.50,   # overridden by revenue-based
        "daily_budget":     15.00,
        "tos_modifier":     25,
        "pp_modifier":      0,
        "bidding_strategy":  "Dynamic bids - down only",
        "description":      "Conservative spend — let Amazon lower bids when unlikely to convert",
    },
    "reviews": {
        "label":            "Reviews",
        "starting_bid":     0.75,
        "daily_budget":     20.00,
        "tos_modifier":     50,
        "pp_modifier":      0,
        "bidding_strategy":  "Fixed bids",
        "description":      "Moderate spend to drive orders for review velocity",
    },
    "marketshare": {
        "label":            "Market Share",
        "starting_bid":     0.50,   # overridden by revenue-based
        "daily_budget":     30.00,
        "tos_modifier":     100,
        "pp_modifier":      50,
        "bidding_strategy":  "Fixed bids",
        "description":      "Broad placement coverage to dominate category presence",
    },
    "research": {
        "label":            "Research",
        "starting_bid":     0.30,
        "daily_budget":     10.00,
        "tos_modifier":     0,
        "pp_modifier":      0,
        "bidding_strategy":  "Dynamic bids - down only",
        "description":      "Low-cost data gathering — discover what converts",
    },
}

PLACEMENT_LABELS = {
    "tos": "ToS",
    "ros": "RoS",
    "pp":  "PP",
    "all": "All",
}


# -- Data Loading -------------------------------------------------------------

def load_harvest(path):
    """Read a harvest output Excel file and extract keyword data."""
    xls = pd.ExcelFile(path, engine="openpyxl")
    target_sheets = ["Harvest \u2192 Exact", "Harvest", "Harvest \u2192 Broad"]
    sheet = None
    for name in target_sheets:
        if name in xls.sheet_names:
            sheet = name
            break
    if sheet is None:
        sheet = xls.sheet_names[0]

    df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")

    # Normalise column names for common harvest output patterns
    col_map = {}
    for col in df.columns:
        cl = col.strip().lower()
        if cl in ("keyword text", "keyword", "search term", "query"):
            col_map[col] = "keyword_text"
        elif cl in ("avg cpc", "cpc", "average cpc"):
            col_map[col] = "avg_cpc"
        elif cl in ("orders", "total orders"):
            col_map[col] = "orders"
        elif cl in ("clicks", "total clicks"):
            col_map[col] = "clicks"
        elif cl in ("sales", "total sales", "revenue"):
            col_map[col] = "sales"
        elif cl in ("impressions", "total impressions"):
            col_map[col] = "impressions"
        elif cl in ("spend", "total spend", "cost"):
            col_map[col] = "spend"
        elif cl in ("acos", "acos %"):
            col_map[col] = "acos"

    df = df.rename(columns=col_map)

    if "keyword_text" not in df.columns:
        # Try first column as keyword text
        df = df.rename(columns={df.columns[0]: "keyword_text"})

    # Ensure numeric columns
    for col in ["avg_cpc", "orders", "clicks", "sales", "impressions", "spend"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Drop rows without keyword text
    df = df.dropna(subset=["keyword_text"])
    df["keyword_text"] = df["keyword_text"].astype(str).str.strip()
    df = df[df["keyword_text"].str.len() > 0]

    if df.empty:
        sys.exit("ERROR: No keywords found in harvest file.")

    return df


def keywords_from_list(kw_string):
    """Split a comma-separated keyword string into a DataFrame."""
    kws = [k.strip() for k in kw_string.split(",") if k.strip()]
    if not kws:
        sys.exit("ERROR: No keywords provided in --keywords list.")
    return pd.DataFrame({"keyword_text": kws})


# -- Bid Calculation ----------------------------------------------------------

def revenue_bid(row, target_acos, aov=None):
    """Revenue-based bid: (Orders/Clicks) * AOV * Target ACoS, clamped."""
    clicks = row.get("clicks", 0)
    orders = row.get("orders", 0)
    sales  = row.get("sales", 0)

    if clicks <= 0 or orders <= 0:
        return None

    if aov and aov > 0:
        order_value = aov
    elif sales > 0 and orders > 0:
        order_value = sales / orders
    else:
        return None

    rpc = (orders / clicks) * order_value
    bid = rpc * target_acos
    return round(max(0.20, min(5.00, bid)), 2)


# -- Campaign Structure Builder -----------------------------------------------

def build_campaigns(df, asin, portfolio, goal, target_acos, starting_bid,
                    bid_strategy, placement, daily_budget):
    """Generate the full SKC structure from keyword DataFrame."""
    cfg = GOAL_CONFIG[goal]
    plc = PLACEMENT_LABELS.get(placement, "ToS")
    start_date = datetime.today().strftime("%Y%m%d")
    campaigns = []

    for _, row in df.iterrows():
        kw = str(row["keyword_text"]).strip()
        if not kw:
            continue

        # Determine bid
        bid = starting_bid
        bid_source = "manual"
        if bid_strategy == "revenue":
            rev_bid = revenue_bid(row, target_acos)
            if rev_bid is not None:
                bid = rev_bid
                bid_source = "revenue"
            else:
                bid = cfg["starting_bid"]
                bid_source = "goal-default"

        campaign_name = f"{portfolio} - SP - KW - Exact - {plc} - {kw}"
        ad_group_name = kw
        bidding_strat = cfg["bidding_strategy"]

        # Placement modifiers based on target placement
        if placement == "tos":
            tos_mod = cfg["tos_modifier"]
            pp_mod  = cfg["pp_modifier"]
        elif placement == "pp":
            tos_mod = 0
            pp_mod  = cfg["tos_modifier"]   # flip the modifier to PP
        elif placement == "ros":
            tos_mod = 0
            pp_mod  = 0
        else:  # all
            tos_mod = cfg["tos_modifier"]
            pp_mod  = cfg["pp_modifier"]

        base = {
            "keyword_text":      kw,
            "campaign_name":     campaign_name,
            "ad_group_name":     ad_group_name,
            "asin":              asin,
            "portfolio":         portfolio,
            "bid":               bid,
            "bid_source":        bid_source,
            "daily_budget":      daily_budget,
            "bidding_strategy":  bidding_strat,
            "tos_modifier":      tos_mod,
            "pp_modifier":       pp_mod,
            "start_date":        start_date,
            "placement_label":   plc,
            "goal":              cfg["label"],
        }

        # Attach harvest metrics if present
        for metric in ["avg_cpc", "orders", "clicks", "sales",
                       "impressions", "spend", "acos"]:
            base[f"src_{metric}"] = row.get(metric, "")

        campaigns.append(base)

    return campaigns


# -- Bulk Upload Row Generation -----------------------------------------------

BULK_HEADERS = [
    "Product", "Entity", "Operation", "Campaign Id", "Ad Group Id",
    "Portfolio Id", "Campaign Name", "Ad Group Name", "Portfolio Name",
    "Start Date", "End Date", "Targeting Type", "State", "Daily Budget",
    "Keyword Text", "Match Type", "Bid", "Keyword Id",
    "Product Targeting Id", "Bidding Strategy",
    "Placement Product Page", "Placement Top", "ASIN", "SKU",
]


def generate_bulk_rows(campaigns):
    """Expand each campaign into the 5-6 entity rows Amazon expects."""
    rows = []
    for c in campaigns:
        # 1) Campaign row
        rows.append({
            "Product":           "Sponsored Products",
            "Entity":            "Campaign",
            "Operation":         "create",
            "Campaign Name":     c["campaign_name"],
            "Portfolio Name":    c["portfolio"],
            "Start Date":        c["start_date"],
            "Targeting Type":    "Manual",
            "State":             "enabled",
            "Daily Budget":      c["daily_budget"],
            "Bidding Strategy":  c["bidding_strategy"],
        })
        # 2) Ad Group row
        rows.append({
            "Product":           "Sponsored Products",
            "Entity":            "Ad Group",
            "Operation":         "create",
            "Campaign Name":     c["campaign_name"],
            "Ad Group Name":     c["ad_group_name"],
            "State":             "enabled",
            "Bid":               c["bid"],
        })
        # 3) Product Ad row
        rows.append({
            "Product":           "Sponsored Products",
            "Entity":            "Product Ad",
            "Operation":         "create",
            "Campaign Name":     c["campaign_name"],
            "Ad Group Name":     c["ad_group_name"],
            "State":             "enabled",
            "ASIN":              c["asin"],
        })
        # 4) Keyword row
        rows.append({
            "Product":           "Sponsored Products",
            "Entity":            "Keyword",
            "Operation":         "create",
            "Campaign Name":     c["campaign_name"],
            "Ad Group Name":     c["ad_group_name"],
            "State":             "enabled",
            "Keyword Text":      c["keyword_text"],
            "Match Type":        "Exact",
            "Bid":               c["bid"],
        })
        # 5) Bidding Adjustment — Placement Top
        if c["tos_modifier"] > 0:
            rows.append({
                "Product":           "Sponsored Products",
                "Entity":            "Bidding Adjustment",
                "Operation":         "create",
                "Campaign Name":     c["campaign_name"],
                "Placement Top":     c["tos_modifier"],
            })
        # 6) Bidding Adjustment — Placement Product Page
        if c["pp_modifier"] > 0:
            rows.append({
                "Product":           "Sponsored Products",
                "Entity":            "Bidding Adjustment",
                "Operation":         "create",
                "Campaign Name":     c["campaign_name"],
                "Placement Product Page": c["pp_modifier"],
            })

    return rows


# -- Excel Helpers ------------------------------------------------------------

def apply_header(ws, row_num, headers, bg, fc="FFFFFF", height=20):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=i, value=h)
        c.font = hfont(color=fc)
        c.fill = hex_fill(bg)
        c.alignment = cal()
        c.border = thin_border()


def sheet_title(ws, title, subtitle, bg, height=28):
    ws.row_dimensions[1].height = height
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    max_col = max(ws.max_column or 1, 12)
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = hex_fill(bg)
        ws.cell(row=1, column=col).border = thin_border()
    ws.row_dimensions[2].height = 16
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    s.fill = hex_fill("FFFFFF")


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# -- Sheet Builders -----------------------------------------------------------

def build_summary(wb, campaigns, goal, target_acos, bid_strategy, placement,
                  daily_budget):
    """Build the summary overview sheet."""
    ws = wb.create_sheet("\U0001f4ca Summary")
    ws.sheet_view.showGridLines = False
    cfg = GOAL_CONFIG[goal]
    n_camps = len(campaigns)
    total_budget = n_camps * daily_budget

    # Title row
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Amazon PPC SKC Builder Report"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hex_fill(C["header_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = (f"Goal: {cfg['label']}   |   Target ACoS: {target_acos:.0%}   |   "
                f"Generated: {datetime.today().strftime('%b %d, %Y')}")
    c2.font = Font(name=FONT, italic=True, size=9, color="888888")
    c2.fill = hex_fill("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # KPI cards
    kpis = [
        ("Campaigns",       f"{n_camps}",                C["header_dark"]),
        ("Total Daily $",   f"${total_budget:,.2f}",     C["header_green"]),
        ("Monthly Est.",    f"${total_budget * 30:,.0f}", C["header_amber"]),
        ("Bid Strategy",    bid_strategy.title(),         C["header_dark"]),
        ("Placement",       PLACEMENT_LABELS.get(placement, placement).upper(),
                                                          C["header_dark"]),
        ("Bidding Mode",    cfg["bidding_strategy"],      C["header_dark"]),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    for i, (label, value, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hex_fill("F5F5F5")
        lc.alignment = cal()
        lc.border = thin_border()
        vc = ws.cell(row=5, column=i, value=value)
        vc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = cal()
        vc.border = thin_border()

    # Goal description
    ws.row_dimensions[7].height = 18
    desc_cell = ws.cell(row=7, column=1,
                        value=f"Goal Strategy: {cfg['description']}")
    desc_cell.font = Font(name=FONT, italic=True, size=10, color="555555")

    # Campaign table
    table_hdrs = ["#", "Campaign Name", "Keyword", "Bid", "Budget",
                  "ToS Mod", "PP Mod", "Bid Source"]
    apply_header(ws, 9, table_hdrs, C["header_dark"])

    for r_idx, camp in enumerate(campaigns):
        er = 10 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[er].height = 16
        vals = [
            r_idx + 1,
            camp["campaign_name"],
            camp["keyword_text"],
            camp["bid"],
            camp["daily_budget"],
            f"{camp['tos_modifier']}%",
            f"{camp['pp_modifier']}%",
            camp["bid_source"],
        ]
        fmts_map = {3: '"$"#,##0.00', 4: '"$"#,##0.00'}
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if c_idx in fmts_map and isinstance(val, (int, float)):
                cell.number_format = fmts_map[c_idx]

    set_widths(ws, {
        "A": 6, "B": 52, "C": 32, "D": 12, "E": 12,
        "F": 10, "G": 10, "H": 14,
    })
    ws.freeze_panes = "A10"


def build_structure(wb, campaigns):
    """Visual campaign hierarchy sheet."""
    ws = wb.create_sheet("\U0001f3d7 Campaign Structure")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Campaign Structure — SKC Hierarchy",
                f"{len(campaigns)} single-keyword campaigns",
                C["header_dark"])

    hdrs = ["Campaign", "Ad Group", "Keyword (Exact)", "ASIN",
            "Bid", "Budget", "Bidding Strategy", "ToS %", "PP %"]
    apply_header(ws, 3, hdrs, C["header_dark"])

    for r_idx, camp in enumerate(campaigns):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[er].height = 16
        vals = [
            camp["campaign_name"],
            camp["ad_group_name"],
            camp["keyword_text"],
            camp["asin"],
            camp["bid"],
            camp["daily_budget"],
            camp["bidding_strategy"],
            f"{camp['tos_modifier']}%",
            f"{camp['pp_modifier']}%",
        ]
        fmts_map = {4: '"$"#,##0.00', 5: '"$"#,##0.00'}
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if c_idx in fmts_map and isinstance(val, (int, float)):
                cell.number_format = fmts_map[c_idx]

    set_widths(ws, {
        "A": 52, "B": 32, "C": 32, "D": 16, "E": 12,
        "F": 12, "G": 26, "H": 10, "I": 10,
    })
    ws.freeze_panes = "A4"


def build_bulk_upload(wb, bulk_rows):
    """Amazon-ready bulk upload sheet."""
    ws = wb.create_sheet("\U0001f4e4 Amazon Bulk Upload")
    ws.sheet_view.showGridLines = False

    entity_counts = {}
    for r in bulk_rows:
        e = r.get("Entity", "")
        entity_counts[e] = entity_counts.get(e, 0) + 1
    camp_count = entity_counts.get("Campaign", 0)

    sheet_title(ws,
        f"Amazon Bulk Upload \u2014 {camp_count} campaigns, {len(bulk_rows)} rows",
        "Upload via Seller Central \u2192 Campaign Manager \u2192 Bulk Operations",
        C["header_dark"])

    apply_header(ws, 3, BULK_HEADERS, C["header_dark"])

    entity_colors = {
        "Campaign":           C["green_light"],
        "Ad Group":           "F0F5FF",
        "Product Ad":         "FFF8F0",
        "Keyword":            "F5F0FF",
        "Bidding Adjustment": C["amber_light"],
    }
    fmts = {
        "Daily Budget": '"$"#,##0.00',
        "Bid":          '"$"#,##0.00',
        "Placement Top":          '0"%"',
        "Placement Product Page": '0"%"',
    }

    for r_idx, row_data in enumerate(bulk_rows):
        er = 4 + r_idx
        entity = row_data.get("Entity", "")
        bg = entity_colors.get(entity, "FFFFFF")
        fill = hex_fill(bg)
        ws.row_dimensions[er].height = 15

        for c_idx, header in enumerate(BULK_HEADERS, 1):
            val = row_data.get(header, "")
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9, bold=(header == "Entity"))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if header in fmts and val != "" and val is not None:
                cell.number_format = fmts[header]

    # Footer note
    last_row = 4 + len(bulk_rows) + 1
    ws.row_dimensions[last_row].height = 20
    note = ws.cell(row=last_row, column=1,
                   value="NOTE: Campaign ID and Ad Group ID are left blank "
                         "\u2014 Amazon assigns these on upload. "
                         "Portfolio must already exist in Seller Central.")
    note.font = Font(name=FONT, bold=True, size=9, color=C["neutral"])
    note.fill = hex_fill(C["amber_light"])
    ws.merge_cells(start_row=last_row, start_column=1,
                   end_row=last_row, end_column=10)

    set_widths(ws, {
        "A": 18, "B": 20, "C": 12, "D": 14, "E": 14, "F": 14,
        "G": 52, "H": 32, "I": 20, "J": 12, "K": 10, "L": 16,
        "M": 10, "N": 14, "O": 32, "P": 12, "Q": 12, "R": 14,
        "S": 18, "T": 26, "U": 22, "V": 16, "W": 16, "X": 10,
    })
    ws.freeze_panes = "A4"


def build_inch_up_schedule(wb, campaigns, bid_strategy):
    """Week-by-week bid increase plan for inch-up strategy."""
    ws = wb.create_sheet("\U0001f4cb Inch-Up Schedule")
    ws.sheet_view.showGridLines = False

    if bid_strategy != "inch-up":
        sheet_title(ws, "Inch-Up Schedule \u2014 Not Applicable",
                    "Bid strategy is revenue-based; inch-up schedule not generated",
                    C["header_gray"])
        ws.cell(row=4, column=1,
                value="Revenue-based bids were calculated from harvest data. "
                      "No inch-up schedule needed.").font = bfont()
        set_widths(ws, {"A": 60})
        return

    sheet_title(ws, "Inch-Up Schedule \u2014 8-Week Bid Ramp Plan",
                "Raise bids $0.05\u2013$0.10 every 3\u20137 days until "
                "impressions flow consistently",
                C["header_dark"])

    # Headers
    week_hdrs = ["Keyword", "Starting Bid"]
    increments = [0.05, 0.05, 0.10, 0.10, 0.10, 0.15, 0.15, 0.20]
    for w in range(1, 9):
        week_hdrs.append(f"Week {w}")
    week_hdrs.append("Week 8 Max")
    apply_header(ws, 3, week_hdrs, C["header_dark"])

    today = datetime.today()
    for r_idx, camp in enumerate(campaigns):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[er].height = 16

        base = camp["bid"]
        running = base
        vals = [camp["keyword_text"], base]
        for inc in increments:
            running = round(min(running + inc, 5.00), 2)
            vals.append(running)
        vals.append(running)  # Week 8 Max repeat

        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = cal() if c_idx > 1 else lal()
            if c_idx >= 2:
                cell.number_format = '"$"#,##0.00'

    # Week date reference row
    date_row = 4 + len(campaigns) + 1
    ws.row_dimensions[date_row].height = 14
    ws.cell(row=date_row, column=1, value="Target Date").font = bfont(
        size=8, bold=True, color="888888")
    ws.cell(row=date_row, column=2, value="Now").font = bfont(
        size=8, color="888888")
    for w in range(1, 9):
        d = today + timedelta(days=w * 7)
        cell = ws.cell(row=date_row, column=2 + w, value=d.strftime("%b %d"))
        cell.font = bfont(size=8, color="888888")
        cell.alignment = cal()

    # Tip row
    tip_row = date_row + 2
    tip = ws.cell(row=tip_row, column=1,
                  value="TIP: Only raise bids if the keyword is getting "
                        "fewer than 100 impressions/day. If impressions are "
                        "healthy, hold the current bid.")
    tip.font = Font(name=FONT, italic=True, size=9, color=C["neutral"])
    ws.merge_cells(start_row=tip_row, start_column=1,
                   end_row=tip_row, end_column=8)

    widths = {"A": 36, "B": 14}
    for i in range(3, 12):
        widths[get_column_letter(i)] = 12
    set_widths(ws, widths)
    ws.freeze_panes = "B4"


def build_source_data(wb, df):
    """Original keywords with any harvest metrics."""
    ws = wb.create_sheet("\U0001f5c2 Source Data")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Source Data \u2014 Original Keywords",
                f"{len(df)} keywords loaded",
                C["header_dark"])

    # Build headers from available columns
    base_cols = ["keyword_text"]
    metric_cols = ["impressions", "clicks", "spend", "sales", "orders",
                   "avg_cpc", "acos"]
    display_hdrs = ["Keyword"]
    display_cols = list(base_cols)
    for m in metric_cols:
        if m in df.columns:
            display_cols.append(m)
            display_hdrs.append(m.replace("_", " ").title())

    apply_header(ws, 3, display_hdrs, C["header_dark"])

    fmt_map = {
        "avg_cpc": '"$"#,##0.00',
        "spend":   '"$"#,##0.00',
        "sales":   '"$"#,##0.00',
        "acos":    '0.0%',
    }

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[er].height = 15
        for c_idx, col in enumerate(display_cols, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in fmt_map and val != "":
                cell.number_format = fmt_map[col]

    col_widths = {"A": 36}
    for i in range(2, len(display_cols) + 1):
        col_widths[get_column_letter(i)] = 14
    set_widths(ws, col_widths)
    ws.freeze_panes = "A4"


# -- Main --------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Amazon PPC Single Keyword Campaign (SKC) Builder")
    parser.add_argument("--input",         default=None,
                        help="Harvest output Excel file")
    parser.add_argument("--keywords",      default=None,
                        help="Comma-separated keyword list (alternative to --input)")
    parser.add_argument("--output",        required=True,
                        help="Output Excel file path")
    parser.add_argument("--asin",          required=True,
                        help="Product ASIN to advertise")
    parser.add_argument("--portfolio",     required=True,
                        help="Portfolio / brand name for naming convention")
    parser.add_argument("--goal",          default="ranking",
                        choices=["ranking", "profit", "reviews",
                                 "marketshare", "research"],
                        help="Campaign goal (default: ranking)")
    parser.add_argument("--target-acos",   type=float, default=0.60,
                        help="Target ACoS as decimal (default: 0.60)")
    parser.add_argument("--starting-bid",  type=float, default=0.50,
                        help="Starting bid in dollars (default: 0.50)")
    parser.add_argument("--bid-strategy",  default="inch-up",
                        choices=["inch-up", "revenue"],
                        help="Bid strategy (default: inch-up)")
    parser.add_argument("--placement",     default="tos",
                        choices=["tos", "ros", "pp", "all"],
                        help="Target placement (default: tos)")
    parser.add_argument("--daily-budget",  type=float, default=25.00,
                        help="Daily budget per campaign (default: 25.00)")
    args = parser.parse_args()

    # Validate inputs
    if not args.input and not args.keywords:
        sys.exit("ERROR: Provide either --input (harvest file) or --keywords.")

    # Load keywords
    if args.input:
        print(f"Loading harvest file: {args.input}")
        df = load_harvest(args.input)
    else:
        print(f"Using keyword list: {args.keywords}")
        df = keywords_from_list(args.keywords)
    print(f"  {len(df)} keywords loaded")

    # Build campaign structures
    print(f"\nBuilding SKC campaigns...")
    print(f"  Goal:         {GOAL_CONFIG[args.goal]['label']}")
    print(f"  Bid strategy: {args.bid_strategy}")
    print(f"  Placement:    {PLACEMENT_LABELS.get(args.placement, args.placement)}")
    print(f"  Target ACoS:  {args.target_acos:.0%}")

    campaigns = build_campaigns(
        df, args.asin, args.portfolio, args.goal, args.target_acos,
        args.starting_bid, args.bid_strategy, args.placement, args.daily_budget)

    print(f"  {len(campaigns)} campaigns generated")

    # Generate bulk upload rows
    bulk_rows = generate_bulk_rows(campaigns)
    print(f"  {len(bulk_rows)} bulk upload rows")

    # Bid summary
    bids = [c["bid"] for c in campaigns]
    if bids:
        print(f"\n  Bid range: ${min(bids):.2f} - ${max(bids):.2f}")
        print(f"  Avg bid:   ${sum(bids)/len(bids):.2f}")
    total_budget = len(campaigns) * args.daily_budget
    print(f"  Total daily budget: ${total_budget:,.2f}")
    print(f"  Est. monthly spend: ${total_budget * 30:,.0f}")

    # Build Excel workbook
    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb, campaigns, args.goal, args.target_acos,
                  args.bid_strategy, args.placement, args.daily_budget)
    build_structure(wb, campaigns)
    build_bulk_upload(wb, bulk_rows)
    build_inch_up_schedule(wb, campaigns, args.bid_strategy)
    build_source_data(wb, df)

    wb.save(args.output)
    print(f"\nSaved: {args.output}")

    # Write findings JSON
    rev_count = sum(1 for c in campaigns if c["bid_source"] == "revenue")
    manual_count = sum(1 for c in campaigns if c["bid_source"] == "manual")
    default_count = sum(1 for c in campaigns if c["bid_source"] == "goal-default")

    findings = {
        "tool":              "skc_builder",
        "goal":              args.goal,
        "target_acos":       args.target_acos,
        "bid_strategy":      args.bid_strategy,
        "placement":         args.placement,
        "campaign_count":    len(campaigns),
        "bulk_rows":         len(bulk_rows),
        "total_daily_budget": round(total_budget, 2),
        "est_monthly_spend": round(total_budget * 30, 2),
        "bid_min":           round(min(bids), 2) if bids else 0,
        "bid_max":           round(max(bids), 2) if bids else 0,
        "bid_avg":           round(sum(bids) / len(bids), 2) if bids else 0,
        "bids_revenue":      rev_count,
        "bids_manual":       manual_count,
        "bids_goal_default": default_count,
        "asin":              args.asin,
        "portfolio":         args.portfolio,
        "bidding_strategy":  GOAL_CONFIG[args.goal]["bidding_strategy"],
        "tos_modifier":      GOAL_CONFIG[args.goal]["tos_modifier"],
        "campaigns": [
            {
                "keyword":    c["keyword_text"],
                "bid":        c["bid"],
                "bid_source": c["bid_source"],
                "campaign":   c["campaign_name"],
                "budget":     c["daily_budget"],
                "tos_mod":    c["tos_modifier"],
                "pp_mod":     c["pp_modifier"],
            }
            for c in campaigns
        ],
    }

    findings_path = args.output.replace(".xlsx", "_findings.json")
    with open(findings_path, "w") as f:
        json.dump(findings, f, indent=2, default=str)
    print(f"   Findings: {findings_path}")

    print(f"\n   Campaigns:    {len(campaigns)}")
    print(f"   Bulk rows:    {len(bulk_rows)}")
    print(f"   Bid strategy: {args.bid_strategy}")
    if rev_count:
        print(f"   Revenue bids: {rev_count}")
    if default_count:
        print(f"   Default bids: {default_count} (no harvest data for revenue calc)")
    print(f"\n   Upload the 'Amazon Bulk Upload' tab to Seller Central "
          f"\u2192 Bulk Operations.")


if __name__ == "__main__":
    main()
