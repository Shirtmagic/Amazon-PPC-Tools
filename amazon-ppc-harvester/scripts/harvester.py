#!/usr/bin/env python3
"""
Amazon PPC Search Term Harvester
Processes a Sponsored Products Search Term Report and outputs
a formatted Excel action file with harvest and negative recommendations.
"""

import argparse
import json
import sys
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
warnings.filterwarnings("ignore")


# ── Colours ────────────────────────────────────────────────────────────────
C = {
    "header_dark":  "1F3864",   # dark navy
    "header_mid":   "2F5496",   # mid navy
    "header_green": "1E6B3C",   # dark green
    "header_red":   "8B0000",   # dark red
    "header_amber": "7B4F00",   # dark amber
    "header_blue":  "1A4A7A",   # steel blue (review)
    "gold":         "FFD700",
    "gold_light":   "FFF8DC",
    "silver_light": "F0F4FF",
    "red_light":    "FFF0F0",
    "amber_light":  "FFFBF0",
    "blue_light":   "F0F8FF",
    "green_light":  "F0FFF4",
    "white":        "FFFFFF",
    "light_gray":   "F5F5F5",
    "mid_gray":     "D9D9D9",
    "text_white":   "FFFFFF",
    "text_dark":    "1A1A1A",
    "positive":     "1E6B3C",
    "negative":     "8B0000",
    "neutral":      "7B4F00",
}

FONT = "Arial"

def hex_fill(hex_code):
    return PatternFill("solid", start_color=hex_code, end_color=hex_code)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=10, bold=True, color="FFFFFF"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def body_font(size=10, bold=False, color="1A1A1A"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


# ── Data Loading & Cleaning ─────────────────────────────────────────────────

COLUMN_MAP = {
    "Start Date":                               "start_date",
    "End Date":                                 "end_date",
    "Portfolio name":                           "portfolio",
    "Currency":                                 "currency",
    "Campaign Name":                            "campaign",
    "Ad Group Name":                            "ad_group",
    "Retailer":                                 "retailer",
    "Country":                                  "country",
    "Targeting":                                "targeting",
    "Match Type":                               "match_type",
    "Customer Search Term":                     "search_term",
    "Impressions":                              "impressions",
    "Clicks":                                   "clicks",
    "Click-Thru Rate (CTR)":                    "ctr",
    "Click-through Rate":                       "ctr",
    "Cost Per Click (CPC)":                     "cpc",
    "CPC":                                      "cpc",
    "Spend":                                    "spend",
    "7 Day Total Sales ":                       "sales",
    "Total Advertising Cost of Sales (ACOS) ":  "acos",
    "Total Return on Advertising Spend (ROAS)": "roas",
    "7 Day Total Orders (#)":                   "orders",
    "7 Day Total Units (#)":                    "units",
    "7 Day Conversion Rate":                    "cvr",
    "7 Day Advertised SKU Units (#)":           "adv_sku_units",
    "7 Day Other SKU Units (#)":                "other_sku_units",
    "7 Day Advertised SKU Sales ":              "adv_sku_sales",
    "7 Day Other SKU Sales ":                   "other_sku_sales",
}

# Fuzzy match for column names (strip whitespace)
def normalise_cols(df):
    col_map = {}
    for orig, clean in COLUMN_MAP.items():
        for c in df.columns:
            if c.strip() == orig.strip():
                col_map[c] = clean
                break
    return df.rename(columns=col_map)


def load_existing_keywords(bulk_path):
    """
    Load all existing Keyword and Negative Keyword rows from the bulk file.
    Returns two dicts:
      existing_kws:  { normalised_term -> set of (campaign, match_type) }
      existing_negs: { normalised_term -> set of (campaign, neg_match_type) }
    Also returns the raw keywords DataFrame for reference.
    """
    df = pd.read_excel(bulk_path, sheet_name="Sponsored Products Campaigns",
                       engine="openpyxl")

    kw_entities  = ["Keyword"]
    neg_entities = ["Negative Keyword", "Negative keyword",
                    "Campaign Negative Keyword"]

    kws  = df[df["Entity"].isin(kw_entities)].copy()
    negs = df[df["Entity"].isin(neg_entities)].copy()

    # Use informational campaign name if Campaign Name is blank
    for frame in [kws, negs]:
        camp_col = "Campaign Name (Informational only)"
        if camp_col in frame.columns:
            frame["_camp"] = frame["Campaign Name"].fillna(
                frame[camp_col].fillna(""))
        else:
            frame["_camp"] = frame.get("Campaign Name", "")

    def build_lookup(frame):
        lookup = {}
        for _, row in frame.iterrows():
            kw = str(row.get("Keyword Text", "") or "").strip().lower()
            if not kw:
                continue
            mt = str(row.get("Match Type", "") or "").strip().lower()
            camp = str(row.get("_camp", "") or "").strip()
            state = str(row.get("State", "enabled") or "enabled").strip().lower()
            if kw not in lookup:
                lookup[kw] = []
            lookup[kw].append({"match_type": mt, "campaign": camp,
                                "state": state})
        return lookup

    existing_kws  = build_lookup(kws)
    existing_negs = build_lookup(negs)
    return existing_kws, existing_negs, kws


def dedup_status(search_term, target_match_type, existing_kws, existing_negs):
    """
    Returns a (status, detail) tuple for a harvest candidate.
    status values:
      'NEW'           – not found anywhere, safe to add
      'DUPLICATE'     – already exists as this exact match type (skip)
      'BROADER'       – already targeted as broad/phrase (note but still add exact)
      'NEG_CONFLICT'  – term or superstring exists as a negative (flag)
    """
    term = str(search_term).strip().lower()
    tmt  = str(target_match_type).strip().lower()

    matches = existing_kws.get(term, [])
    neg_matches = existing_negs.get(term, [])

    # Check for exact duplicate match type
    for m in matches:
        if m["match_type"] == tmt:
            return ("DUPLICATE",
                    f"Already {tmt} in: {m['campaign']} [{m['state']}]")

    # Check if already targeted more broadly
    broader_found = []
    for m in matches:
        if tmt == "exact" and m["match_type"] in ("phrase", "broad"):
            broader_found.append(f"{m['match_type']} in {m['campaign']}")
        elif tmt == "phrase" and m["match_type"] == "broad":
            broader_found.append(f"broad in {m['campaign']}")
    if broader_found:
        return ("BROADER", "Already targeted as: " + "; ".join(broader_found))

    # Check for negative conflict
    if neg_matches:
        camps = "; ".join(m["campaign"] for m in neg_matches[:2])
        return ("NEG_CONFLICT",
                f"⚠ Exists as negative in: {camps}")

    return ("NEW", "✅ Not found — safe to add")


def load_report(path):
    df = pd.read_excel(path, engine="openpyxl")
    df = normalise_cols(df)
    required = ["campaign", "ad_group", "match_type", "search_term",
                "clicks", "spend", "sales", "orders"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        sys.exit(f"ERROR: Missing columns after mapping: {missing}\n"
                 f"Available: {list(df.columns)}")
    def _col(df, name):
        return df[name] if name in df.columns else pd.Series(0, index=df.index)
    df["spend"]       = pd.to_numeric(_col(df, "spend"),       errors="coerce").fillna(0)
    df["sales"]       = pd.to_numeric(_col(df, "sales"),       errors="coerce").fillna(0)
    df["orders"]      = pd.to_numeric(_col(df, "orders"),      errors="coerce").fillna(0)
    df["clicks"]      = pd.to_numeric(_col(df, "clicks"),      errors="coerce").fillna(0)
    df["impressions"] = pd.to_numeric(_col(df, "impressions"), errors="coerce").fillna(0)
    df["cpc"]         = pd.to_numeric(_col(df, "cpc"),         errors="coerce").fillna(0)
    df["units"]       = pd.to_numeric(_col(df, "units"),       errors="coerce").fillna(0)
    return df


def aggregate(df):
    """Aggregate daily rows → one row per campaign/adgroup/search_term/match_type."""
    grp = ["portfolio", "campaign", "ad_group", "match_type", "search_term"]
    grp = [c for c in grp if c in df.columns]
    agg = df.groupby(grp, as_index=False).agg(
        impressions=("impressions", "sum"),
        clicks=("clicks", "sum"),
        spend=("spend", "sum"),
        sales=("sales", "sum"),
        orders=("orders", "sum"),
        units=("units",  "sum"),
        avg_cpc=("cpc", "mean"),
    )
    agg["acos"] = np.where(agg["sales"] > 0, agg["spend"] / agg["sales"], np.nan)
    agg["roas"] = np.where(agg["spend"] > 0, agg["sales"] / agg["spend"], np.nan)
    agg["cvr"]  = np.where(agg["clicks"] > 0, agg["orders"] / agg["clicks"], np.nan)
    return agg


# ── Classification ──────────────────────────────────────────────────────────

def classify(df, target_acos, min_clicks, min_orders, neg_spend,
             existing_kws=None, existing_negs=None):
    """Tag every row with a harvest category."""
    is_auto   = df["match_type"] == "-"
    is_broad  = df["match_type"].str.upper() == "BROAD"
    is_phrase = df["match_type"].str.upper() == "PHRASE"
    is_exact  = df["match_type"].str.upper() == "EXACT"

    has_orders = df["orders"] >= min_orders
    has_clicks = df["clicks"] >= min_clicks
    low_acos   = df["acos"] <= target_acos
    mid_acos   = df["acos"] <= target_acos * 1.5
    high_acos  = df["acos"] > target_acos * 2
    zero_orders = df["orders"] == 0
    big_spend   = df["spend"] >= neg_spend
    med_spend   = df["spend"] >= neg_spend * 0.5
    many_clicks = df["clicks"] >= 5
    some_clicks = df["clicks"] >= 3

    word_count = df["search_term"].astype(str).str.split().str.len()
    long_tail  = word_count >= 3

    gold   = has_orders & has_clicks & low_acos & (is_auto | is_broad | is_phrase) & ~is_exact
    silver = has_orders & (df["clicks"] >= 2) & mid_acos & is_auto & ~gold
    neg_e  = zero_orders & big_spend & many_clicks
    neg_p  = zero_orders & med_spend & some_clicks & long_tail & ~neg_e
    review = has_orders & high_acos & (df["spend"] >= 5)

    df = df.copy()
    df["category"] = "No Action"
    df.loc[review, "category"] = "Review"
    df.loc[neg_p,  "category"] = "Negative → Phrase"
    df.loc[neg_e,  "category"] = "Negative → Exact"
    df.loc[silver, "category"] = "Harvest → Phrase"
    df.loc[gold,   "category"] = "Harvest → Exact"

    # ── Expert bid formula: ACoS-ratio × CPC, Auto terms get slight premium ──
    def _expert_bid(row, tgt=target_acos):
        if not str(row.get("category","")).startswith("Harvest"):
            return np.nan
        cpc  = float(row["avg_cpc"]) if float(row["avg_cpc"]) > 0 else 0.50
        acos = row["acos"]
        is_auto = str(row.get("match_type","")).strip() == "-"
        if pd.isna(acos) or acos <= 0:
            bid = cpc * 0.75          # no conversion proof — start conservative
        elif acos <= tgt * 0.5:
            bid = cpc * min(tgt / acos, 1.5)  # very efficient — bid up
        elif acos <= tgt:
            bid = cpc * (tgt / acos)  # on-target — proportional scale
        else:
            bid = cpc * (tgt / acos) * 0.90   # over-target — discounted start
        if is_auto:
            bid *= 1.10  # Auto terms proven by Amazon's own matching — premium
        return round(float(np.clip(bid, 0.10, 5.00)), 2)

    df["suggested_bid"] = df.apply(_expert_bid, axis=1)
    # Source type: Auto campaign terms are highest-value harvest candidates
    df["source_type"]    = np.where(df["match_type"].str.strip() == "-",
                                    "🤖 Auto", "✋ Manual")
    # Priority score: revenue × statistical confidence
    df["priority_score"] = np.where(
        df["category"].str.startswith("Harvest"),
        (df["sales"] * np.log1p(df["clicks"])).round(1),
        np.nan
    )

    # ── Dedup check ────────────────────────────────────────────────────────
    if existing_kws is not None and existing_negs is not None:
        dedup_status_list = []
        dedup_detail_list = []
        for _, row in df.iterrows():
            cat = row.get("category", "No Action")
            if cat.startswith("Harvest"):
                tmt = "exact" if "Exact" in cat else "phrase"
                status, detail = dedup_status(
                    row["search_term"], tmt, existing_kws, existing_negs)
            else:
                status, detail = ("N/A", "")
            dedup_status_list.append(status)
            dedup_detail_list.append(detail)
        df["dedup_status"] = dedup_status_list
        df["dedup_detail"] = dedup_detail_list

        # Demote DUPLICATE rows so they don't appear in harvest sheets
        dup_mask = df["dedup_status"] == "DUPLICATE"
        df.loc[dup_mask, "category"] = "Duplicate — Already Targeted"
    else:
        df["dedup_status"] = "N/A (no bulk file)"
        df["dedup_detail"] = ""

    return df


# ── Excel Writing Helpers ───────────────────────────────────────────────────

def apply_header_row(ws, row_num, headers, bg_color, font_color="FFFFFF", row_height=20):
    ws.row_dimensions[row_num].height = row_height
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=i, value=h)
        cell.font = header_font(color=font_color)
        cell.fill = hex_fill(bg_color)
        cell.alignment = center()
        cell.border = thin_border()

def write_data_rows(ws, df, cols, start_row, fmt_map=None, alt_fill=None):
    fmt_map = fmt_map or {}
    for r_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = start_row + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(alt_fill or "F8F9FA")
        for c_idx, col in enumerate(cols, 1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font()
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()
            if col in fmt_map:
                cell.number_format = fmt_map[col]

def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def add_sheet_title(ws, title, subtitle, bg_color, row_height=28):
    ws.row_dimensions[1].height = row_height
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    cell.fill = hex_fill(bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    max_col = ws.max_column or 10
    for c in range(2, max_col + 1):
        cell2 = ws.cell(row=1, column=c)
        cell2.fill = hex_fill(bg_color)
        cell2.border = thin_border()

    ws.row_dimensions[2].height = 18
    cell3 = ws.cell(row=2, column=1, value=subtitle)
    cell3.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    cell3.fill = hex_fill("FFFFFF")


# ── Sheet Builders ──────────────────────────────────────────────────────────

def build_summary(wb, df, target_acos, date_range, portfolios):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    total_spend  = df["spend"].sum()
    total_sales  = df["sales"].sum()
    total_orders = df["orders"].sum()
    total_clicks = df["clicks"].sum()
    overall_acos = total_spend / total_sales if total_sales > 0 else 0
    overall_roas = total_sales / total_spend if total_spend > 0 else 0
    overall_cvr  = total_orders / total_clicks if total_clicks > 0 else 0

    gold_count   = (df["category"] == "Harvest → Exact").sum()
    silver_count = (df["category"] == "Harvest → Phrase").sum()
    neg_count    = df["category"].str.startswith("Negative").sum()
    review_count = (df["category"] == "Review").sum()
    neg_spend    = df[df["category"].str.startswith("Negative")]["spend"].sum()

    # Title bar
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "🔍  Amazon PPC Search Term Harvest Report"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hex_fill(C["header_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = f"Date Range: {date_range}   |   Portfolios: {', '.join(portfolios)}   |   Generated: {datetime.today().strftime('%b %d, %Y')}"
    c2.font = Font(name=FONT, italic=True, size=9, color="888888")
    c2.fill = hex_fill("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── KPI Cards row ──
    ws.row_dimensions[4].height = 18
    kpis = [
        ("Total Spend",     f"${total_spend:,.2f}",  C["header_mid"]),
        ("Total Sales",     f"${total_sales:,.2f}",  C["header_green"]),
        ("Overall ACoS",    f"{overall_acos:.1%}",   C["header_red"] if overall_acos > target_acos else C["header_green"]),
        ("Overall ROAS",    f"{overall_roas:.1f}x",  C["header_green"]),
        ("Total Orders",    f"{int(total_orders):,}", C["header_mid"]),
        ("Conv. Rate",      f"{overall_cvr:.1%}",    C["header_amber"]),
        ("Target ACoS",     f"{target_acos:.0%}",    C["header_dark"]),
        ("Total Clicks",    f"{int(total_clicks):,}", C["header_dark"]),
    ]
    for col_idx, (label, value, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
        ws.row_dimensions[4].height = 14
        lc = ws.cell(row=4, column=col_idx, value=label)
        lc.font = Font(name=FONT, size=8, color="888888", bold=False)
        lc.fill = hex_fill("F5F5F5")
        lc.alignment = center()
        lc.border = thin_border()

        ws.row_dimensions[5].height = 22
        vc = ws.cell(row=5, column=col_idx, value=value)
        vc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = center()
        vc.border = thin_border()

    # ── Action Summary ──
    ws.row_dimensions[7].height = 18
    action_headers = ["Action Category", "Count", "Est. Spend Impact", "What It Means"]
    apply_header_row(ws, 7, action_headers, C["header_dark"])

    actions = [
        ("🟡  Harvest → Exact",   gold_count,   "",
         "High-converting terms to add as Exact match keywords"),
        ("🔵  Harvest → Phrase",  silver_count, "",
         "Good Auto terms to add as Phrase match keywords"),
        ("🔴  Negative → Exact",  neg_count,    f"-${neg_spend:,.2f}/period",
         "Terms burning budget with zero orders — add as negatives"),
        ("🟠  Review",            review_count, "",
         "Converting but ACoS > 2× target — needs manual attention"),
    ]
    for r_idx, (cat, count, impact, meaning) in enumerate(actions, 1):
        row_num = 7 + r_idx
        ws.row_dimensions[row_num].height = 18
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        for c_idx, val in enumerate([cat, count, impact, meaning], 1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font = body_font(bold=(c_idx == 1))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()

    # ── Portfolio breakdown ──
    ws.row_dimensions[13].height = 18
    port_headers = ["Portfolio", "Spend", "Sales", "ACoS", "Orders",
                    "Harvest Terms", "Negative Terms"]
    apply_header_row(ws, 13, port_headers, C["header_mid"])

    if "portfolio" in df.columns:
        port_df = df.groupby("portfolio").agg(
            spend=("spend", "sum"),
            sales=("sales", "sum"),
            orders=("orders", "sum"),
            harvest=("category", lambda x: x.str.startswith("Harvest").sum()),
            negatives=("category", lambda x: x.str.startswith("Negative").sum()),
        ).reset_index()
        port_df["acos"] = np.where(port_df["sales"] > 0,
                                   port_df["spend"] / port_df["sales"], np.nan)
        port_df = port_df.sort_values("spend", ascending=False)

        for r_idx, (_, row) in enumerate(port_df.iterrows()):
            row_num = 14 + r_idx
            ws.row_dimensions[row_num].height = 17
            fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
            vals = [
                row["portfolio"],
                row["spend"],
                row["sales"],
                row["acos"] if not pd.isna(row["acos"]) else "",
                int(row["orders"]),
                int(row["harvest"]),
                int(row["negatives"]),
            ]
            fmts = [None, '"$"#,##0.00', '"$"#,##0.00', '0.0%', None, None, None]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=row_num, column=c_idx, value=val)
                cell.font = body_font()
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = left()
                if fmt:
                    cell.number_format = fmt

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 36

    freeze(ws, "A3")


def build_harvest_sheet(wb, df, match_type_label, bg_color, light_color, tab_emoji):
    category = f"Harvest → {match_type_label}"
    data = df[df["category"] == category].copy()
    data = data.sort_values(["portfolio", "campaign", "orders", "spend"],
                            ascending=[True, True, False, False])

    tab_name = f"{tab_emoji} Harvest → {match_type_label}"
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    title = f"Harvest Candidates — Add as {match_type_label.upper()} Keywords"
    subtitle = (f"{len(data)} terms found | "
                f"Sorted by portfolio then orders ↓")
    add_sheet_title(ws, title, subtitle, bg_color)

    headers = [
        "Portfolio", "Campaign", "Ad Group", "Source Match Type",
        "Search Term", "Impressions", "Clicks", "Spend", "Sales",
        "Orders", "ACoS", "CVR", "Avg CPC", "Suggested Bid",
        "Dedup Status", "Dedup Detail", "Action"
    ]
    apply_header_row(ws, 3, headers, bg_color)

    cols = ["portfolio", "campaign", "ad_group", "match_type", "search_term",
            "impressions", "clicks", "spend", "sales", "orders", "acos", "cvr",
            "avg_cpc", "suggested_bid", "dedup_status", "dedup_detail"]

    fmt_map = {
        "spend": '"$"#,##0.00',
        "sales": '"$"#,##0.00',
        "acos":  '0.0%',
        "cvr":   '0.0%',
        "avg_cpc": '"$"#,##0.00',
        "suggested_bid": '"$"#,##0.00',
    }

    for r_idx, (_, row) in enumerate(data.iterrows()):
        excel_row = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(light_color)
        ws.row_dimensions[excel_row].height = 16
        for c_idx, col in enumerate(cols, 1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font()
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()
            if col in fmt_map and val != "":
                cell.number_format = fmt_map[col]

        # Dedup status colouring
        ds = row.get("dedup_status", "NEW")
        ds_color = {"NEW": C["positive"], "DUPLICATE": C["negative"],
                    "BROADER": C["neutral"], "NEG_CONFLICT": C["negative"]
                    }.get(str(ds), C["positive"])
        ds_cell = ws.cell(row=excel_row, column=15, value=ds)
        ds_cell.font = Font(name=FONT, bold=True, size=9, color=ds_color)
        ds_cell.fill = fill
        ds_cell.border = thin_border()
        ds_cell.alignment = center()

        dd_cell = ws.cell(row=excel_row, column=16,
                          value=row.get("dedup_detail", ""))
        dd_cell.font = body_font(size=9)
        dd_cell.fill = fill
        dd_cell.border = thin_border()
        dd_cell.alignment = left()

        # Action column
        ds_val = str(row.get("dedup_status", "NEW"))
        if ds_val == "DUPLICATE":
            action_text = "⛔ SKIP — already targeted"
            action_color = C["negative"]
        elif ds_val == "NEG_CONFLICT":
            action_text = "⚠ CHECK — negative conflict"
            action_color = C["neutral"]
        elif ds_val == "BROADER":
            action_text = f"➕ Add as {match_type_label.upper()} (more specific)"
            action_color = C["neutral"]
        else:
            action_text = f"✅ Add as {match_type_label.upper()} keyword"
            action_color = C["positive"]

        action_cell = ws.cell(row=excel_row, column=17, value=action_text)
        action_cell.font = Font(name=FONT, bold=True, size=9, color=action_color)
        action_cell.fill = fill
        action_cell.border = thin_border()
        action_cell.alignment = left()

    widths = {
        "A": 22, "B": 32, "C": 24, "D": 18, "E": 42,
        "F": 12, "G": 10, "H": 12, "I": 12, "J": 10,
        "K": 10, "L": 10, "M": 12, "N": 14, "O": 14,
        "P": 44, "Q": 36,
    }
    set_col_widths(ws, widths)
    freeze(ws, "A4")
    return len(data)


def build_negatives_sheet(wb, df, neg_spend_threshold):
    data = df[df["category"].str.startswith("Negative")].copy()
    data = data.sort_values("spend", ascending=False)

    ws = wb.create_sheet("🔴 Negatives")
    ws.sheet_view.showGridLines = False

    wasted = data["spend"].sum()
    title = f"Negative Keyword Candidates — ${wasted:,.2f} Wasted Spend"
    subtitle = (f"{len(data)} terms | Adding these negatives recovers "
                f"${wasted:,.2f} per report period")
    add_sheet_title(ws, title, subtitle, C["header_red"])

    headers = [
        "Portfolio", "Campaign", "Ad Group", "Match Type",
        "Search Term", "Impressions", "Clicks", "Spend",
        "Orders", "Why Flagged", "Action"
    ]
    apply_header_row(ws, 3, headers, C["header_red"])

    for r_idx, (_, row) in enumerate(data.iterrows()):
        excel_row = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["red_light"])
        ws.row_dimensions[excel_row].height = 16

        spend_val = row.get("spend", 0)
        clicks_val = row.get("clicks", 0)
        cat = row.get("category", "")
        reason = (f"${spend_val:.2f} spend, {int(clicks_val)} clicks, "
                  f"0 orders")

        vals = [
            row.get("portfolio", ""), row.get("campaign", ""),
            row.get("ad_group", ""), row.get("match_type", ""),
            row.get("search_term", ""), row.get("impressions", 0),
            int(clicks_val), spend_val, 0, reason,
            "Negative Exact" if "Exact" in cat else "Negative Phrase"
        ]
        fmts = [None, None, None, None, None, None, None,
                '"$"#,##0.00', None, None, None]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font()
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()
            if fmt and val not in ("", None):
                cell.number_format = fmt

        # Action cell coloured
        action_cell = ws.cell(row=excel_row, column=11,
                              value=vals[10])
        action_cell.font = Font(name=FONT, bold=True, size=9,
                                color=C["negative"])
        action_cell.fill = fill
        action_cell.border = thin_border()

    widths = {
        "A": 22, "B": 32, "C": 24, "D": 16, "E": 42,
        "F": 12, "G": 10, "H": 12, "I": 10, "J": 38, "K": 20,
    }
    set_col_widths(ws, widths)
    freeze(ws, "A4")


def build_review_sheet(wb, df, target_acos):
    data = df[df["category"] == "Review"].copy()
    data = data.sort_values("acos", ascending=False)

    ws = wb.create_sheet("🟠 Review")
    ws.sheet_view.showGridLines = False

    title = "Manual Review — Converting but High ACoS"
    subtitle = (f"{len(data)} terms with ACoS > {target_acos*2:.0%} | "
                "These convert but are too expensive — consider reducing bids")
    add_sheet_title(ws, title, subtitle, C["header_amber"])

    headers = [
        "Portfolio", "Campaign", "Ad Group", "Match Type",
        "Search Term", "Clicks", "Spend", "Sales", "Orders",
        "ACoS", "Recommended Action"
    ]
    apply_header_row(ws, 3, headers, C["header_amber"])

    for r_idx, (_, row) in enumerate(data.iterrows()):
        excel_row = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["amber_light"])
        ws.row_dimensions[excel_row].height = 16

        acos_val = row.get("acos", None)
        ratio = (acos_val / target_acos) if (acos_val and target_acos > 0) else 0
        if ratio > 4:
            rec = "Consider pausing — ACoS extremely high"
        elif ratio > 3:
            rec = "Reduce bid 30-40%"
        else:
            rec = "Reduce bid 15-25%"

        vals = [
            row.get("portfolio", ""), row.get("campaign", ""),
            row.get("ad_group", ""), row.get("match_type", ""),
            row.get("search_term", ""), int(row.get("clicks", 0)),
            row.get("spend", 0), row.get("sales", 0),
            int(row.get("orders", 0)), acos_val, rec
        ]
        fmts = [None, None, None, None, None, None,
                '"$"#,##0.00', '"$"#,##0.00', None, '0.0%', None]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if pd.isna(val if val is not None else float('nan')):
                val = ""
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font()
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()
            if fmt and val not in ("", None):
                cell.number_format = fmt

    widths = {
        "A": 22, "B": 32, "C": 24, "D": 16, "E": 42,
        "F": 10, "G": 12, "H": 12, "I": 10, "J": 10, "K": 38,
    }
    set_col_widths(ws, widths)
    freeze(ws, "A4")


def build_bulk_upload_sheet(wb, df):
    """
    Produces an Amazon Sponsored Products bulk upload-compatible sheet.
    Rows: Keyword (harvest) + Negative keyword rows.
    User must fill in Campaign Id / Ad Group Id before uploading.
    """
    ws = wb.create_sheet("📤 Amazon Bulk Upload")
    ws.sheet_view.showGridLines = False

    title = "Amazon Bulk Upload — Keyword & Negative Keyword Rows"
    subtitle = ("⚠ Fill in Campaign Id and Ad Group Id before uploading to Seller Central → Campaign Manager → Bulk Operations")
    add_sheet_title(ws, title, subtitle, C["header_dark"])

    # Amazon bulk operations columns (standard SP format)
    headers = [
        "Product", "Entity", "Operation", "Campaign Id", "Ad Group Id",
        "Portfolio Id", "Ad Id", "Keyword Id", "Campaign Name",
        "Ad Group Name", "Start Date", "End Date", "Targeting Type",
        "State", "Daily Budget", "SKU", "ASIN", "Ad Group Default Bid",
        "Bid", "Keyword Text", "Match Type", "Bidding Strategy",
        "Placement", "Percentage", "Product Targeting Expression"
    ]
    apply_header_row(ws, 3, headers, C["header_dark"], row_height=18)

    harvest = df[df["category"].str.startswith("Harvest")].copy()
    negatives = df[df["category"].str.startswith("Negative")].copy()

    excel_row = 4
    for _, row in harvest.iterrows():
        mt = "exact" if "Exact" in row["category"] else "phrase"
        bid = row.get("suggested_bid", 0.75)
        if pd.isna(bid) or bid == 0:
            bid = 0.75
        vals = {
            "Product": "Sponsored Products",
            "Entity": "Keyword",
            "Operation": "create",
            "Campaign Id": "⬅ FILL IN",
            "Ad Group Id": "⬅ FILL IN",
            "Portfolio Id": "",
            "Ad Id": "", "Keyword Id": "",
            "Campaign Name": row.get("campaign", ""),
            "Ad Group Name": row.get("ad_group", ""),
            "Start Date": "", "End Date": "",
            "Targeting Type": "manual",
            "State": "enabled",
            "Daily Budget": "", "SKU": "", "ASIN": "",
            "Ad Group Default Bid": "",
            "Bid": round(float(bid), 2),
            "Keyword Text": row.get("search_term", ""),
            "Match Type": mt,
            "Bidding Strategy": "", "Placement": "",
            "Percentage": "", "Product Targeting Expression": "",
        }
        ws.row_dimensions[excel_row].height = 15
        fill = hex_fill(C["gold_light"])
        for c_idx, h in enumerate(headers, 1):
            val = vals.get(h, "")
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()
            if h == "Bid":
                cell.number_format = '"$"#,##0.00'
        excel_row += 1

    for _, row in negatives.iterrows():
        mt = "negativeExact" if "Exact" in row["category"] else "negativePhrase"
        vals = {
            "Product": "Sponsored Products",
            "Entity": "Negative keyword",
            "Operation": "create",
            "Campaign Id": "⬅ FILL IN",
            "Ad Group Id": "⬅ FILL IN",
            "Portfolio Id": "",
            "Ad Id": "", "Keyword Id": "",
            "Campaign Name": row.get("campaign", ""),
            "Ad Group Name": row.get("ad_group", ""),
            "Start Date": "", "End Date": "",
            "Targeting Type": "manual",
            "State": "enabled",
            "Daily Budget": "", "SKU": "", "ASIN": "",
            "Ad Group Default Bid": "", "Bid": "",
            "Keyword Text": row.get("search_term", ""),
            "Match Type": mt,
            "Bidding Strategy": "", "Placement": "",
            "Percentage": "", "Product Targeting Expression": "",
        }
        ws.row_dimensions[excel_row].height = 15
        fill = hex_fill(C["red_light"])
        for c_idx, h in enumerate(headers, 1):
            val = vals.get(h, "")
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()
        excel_row += 1

    # Column widths
    for i, w in enumerate([16,18,12,16,16,14,8,12,36,28,12,12,14,12,
                            14,10,12,18,12,36,14,16,14,14,26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Warning note
    ws.row_dimensions[excel_row + 1].height = 20
    note = ws.cell(row=excel_row + 1, column=1,
                   value="⚠ ACTION REQUIRED: Fill in Campaign Id and Ad Group Id "
                         "columns before uploading. Get these IDs from Campaign Manager → "
                         "Columns → Campaign Id / Ad Group Id.")
    note.font = Font(name=FONT, bold=True, size=9, color=C["negative"])
    note.fill = hex_fill("FFF0F0")
    ws.merge_cells(
        start_row=excel_row + 1, start_column=1,
        end_row=excel_row + 1, end_column=10
    )

    freeze(ws, "A4")


def build_duplicates_sheet(wb, df):
    """Sheet showing all harvest terms that were suppressed due to dedup."""
    data = df[df["category"] == "Duplicate — Already Targeted"].copy()
    if data.empty:
        return 0

    ws = wb.create_sheet("⛔ Duplicates Suppressed")
    ws.sheet_view.showGridLines = False

    add_sheet_title(ws,
        f"Suppressed Duplicates — {len(data)} terms already in your campaigns",
        "These were harvest candidates but already exist as that match type — no action needed",
        C["header_dark"])

    headers = ["Portfolio", "Campaign", "Ad Group", "Source Match Type",
               "Search Term", "Clicks", "Orders", "ACoS",
               "Dedup Status", "Already Exists In"]
    apply_header_row(ws, 3, headers, C["header_dark"])

    for r_idx, (_, row) in enumerate(data.iterrows()):
        excel_row = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[excel_row].height = 15
        vals = [
            row.get("portfolio",""), row.get("campaign",""),
            row.get("ad_group",""), row.get("match_type",""),
            row.get("search_term",""), int(row.get("clicks",0)),
            int(row.get("orders",0)),
            row.get("acos","") if not pd.isna(row.get("acos", float("nan"))) else "",
            row.get("dedup_status",""), row.get("dedup_detail",""),
        ]
        fmts = [None,None,None,None,None,None,None,'0.0%',None,None]
        for c_idx,(val,fmt) in enumerate(zip(vals,fmts),1):
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left()
            if fmt and val not in ("", None):
                cell.number_format = fmt

    for i,w in enumerate([20,30,22,16,40,10,10,10,14,44],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    freeze(ws, "A4")
    return len(data)


def build_raw_sheet(wb, df):
    ws = wb.create_sheet("🗂 Raw Data")
    ws.sheet_view.showGridLines = False

    add_sheet_title(ws, "Aggregated Raw Data with Classifications",
                    "All search terms — aggregated across date range with category flags",
                    C["header_dark"])

    cols = ["portfolio", "campaign", "ad_group", "match_type", "search_term",
            "impressions", "clicks", "spend", "sales", "orders", "acos",
            "roas", "cvr", "avg_cpc", "suggested_bid", "category"]
    headers = [c.replace("_", " ").title() for c in cols]
    apply_header_row(ws, 3, headers, C["header_dark"])

    cat_colors = {
        "Harvest → Exact":   C["gold_light"],
        "Harvest → Phrase":  C["silver_light"],
        "Negative → Exact":  C["red_light"],
        "Negative → Phrase": C["red_light"],
        "Review":            C["amber_light"],
        "No Action":         "FFFFFF",
    }
    fmt_map = {
        "spend": '"$"#,##0.00', "sales": '"$"#,##0.00',
        "acos": '0.0%', "roas": '0.0x',
        "cvr": '0.0%', "avg_cpc": '"$"#,##0.00',
        "suggested_bid": '"$"#,##0.00',
    }

    for r_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = 4 + r_idx
        cat = row.get("category", "No Action")
        row_fill = hex_fill(cat_colors.get(cat, "FFFFFF"))
        ws.row_dimensions[excel_row].height = 15
        for c_idx, col in enumerate(cols, 1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = body_font(size=9)
            cell.fill = row_fill
            cell.border = thin_border()
            cell.alignment = left()
            if col in fmt_map and val != "":
                cell.number_format = fmt_map[col]

    for i, w in enumerate([20,30,22,14,40,12,10,12,12,10,10,10,
                            10,12,14,20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    freeze(ws, "A4")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Amazon PPC Search Term Harvester")
    parser.add_argument("--input",  required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--bulk-file",           default=None,
                        help="Bulk operations file for keyword dedup check")
    parser.add_argument("--target-acos",        type=float, default=0.25)
    parser.add_argument("--min-clicks",         type=int,   default=3)
    parser.add_argument("--min-orders",         type=int,   default=1)
    parser.add_argument("--neg-spend-threshold",type=float, default=10.0)
    args = parser.parse_args()

    print(f"Loading: {args.input}")
    raw = load_report(args.input)
    print(f"  {len(raw)} rows loaded. Aggregating...")

    df = aggregate(raw)
    print(f"  {len(df)} unique search term × campaign rows after aggregation.")

    existing_kws, existing_negs = None, None
    if args.bulk_file:
        print(f"\nLoading bulk file for dedup: {args.bulk_file}")
        existing_kws, existing_negs, kw_df = load_existing_keywords(args.bulk_file)
        print(f"  {len(existing_kws)} unique keyword terms loaded")
        print(f"  {len(existing_negs)} unique negative terms loaded")

    df = classify(df, args.target_acos, args.min_clicks,
                  args.min_orders, args.neg_spend_threshold,
                  existing_kws, existing_negs)

    counts = df["category"].value_counts()
    print("\nClassification results:")
    for cat, n in counts.items():
        print(f"  {cat}: {n}")

    # Date range
    if "start_date" in raw.columns:
        d_min = pd.to_datetime(raw["start_date"]).min().strftime("%b %d, %Y")
        d_max = pd.to_datetime(raw["start_date"]).max().strftime("%b %d, %Y")
        date_range = f"{d_min} – {d_max}"
        days = (pd.to_datetime(raw["start_date"]).max() -
                pd.to_datetime(raw["start_date"]).min()).days + 1
    else:
        date_range, days = "Unknown", 7

    portfolios = list(df["portfolio"].unique()) if "portfolio" in df.columns else ["All"]

    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    build_summary(wb, df, args.target_acos, date_range, portfolios)
    gold_n   = build_harvest_sheet(wb, df, "Exact",  C["header_mid"],    C["silver_light"], "🟡")
    silver_n = build_harvest_sheet(wb, df, "Phrase", C["header_green"],  C["green_light"],  "🔵")
    build_negatives_sheet(wb, df, args.neg_spend_threshold)
    build_review_sheet(wb, df, args.target_acos)
    build_bulk_upload_sheet(wb, df)
    dup_n = build_duplicates_sheet(wb, df)
    build_raw_sheet(wb, df)

    if days < 7:
        print(f"\n⚠  WARNING: Report only covers {days} days — "
              "results may not be statistically significant.")

    # ── Write cross-tool findings JSON ────────────────────────────────────
    def _safe(v):
        if isinstance(v, (float, np.floating)):
            return round(float(v), 4)
        if isinstance(v, (int, np.integer)):
            return int(v)
        return str(v)

    harvest_rows = df[df["category"] == "Harvest → Exact"].nlargest(5, "sales")
    top_harvest = [
        {k: _safe(v) for k, v in r.items()
         if k in ("search_term","orders","sales","acos","suggested_bid","source_type","priority_score")}
        for _, r in harvest_rows.iterrows()
    ]
    neg_rows = df[df["category"].str.startswith("Negative")].nlargest(5, "spend")
    top_negatives = [
        {k: _safe(v) for k, v in r.items()
         if k in ("search_term","spend","clicks","orders","campaign")}
        for _, r in neg_rows.iterrows()
    ]
    findings = {
        "tool": "harvester",
        "target_acos": args.target_acos,
        "harvest_exact":        int(gold_n),
        "harvest_phrase":       int(silver_n),
        "negatives_count":      int(df["category"].str.startswith("Negative").sum()),
        "wasted_spend":         round(float(df[df["category"].str.startswith("Negative")]["spend"].sum()), 2),
        "review_count":         int((df["category"] == "Review").sum()),
        "duplicates_suppressed":int(dup_n),
        "auto_terms":           int((df["source_type"] == "🤖 Auto").sum()),
        "top_harvest_exact":    top_harvest,
        "top_negatives":        top_negatives,
        "actions": (
            [{"priority": "HIGH",   "type": "ADD_NEGATIVE",
              "subject":  r["search_term"], "campaign": r.get("campaign",""),
              "impact_spend": _safe(r["spend"]),
              "detail": f"Zero orders, ${r['spend']:.2f} wasted spend"}
             for _, r in neg_rows.iterrows() if r["spend"] >= 20]
            +
            [{"priority": "MEDIUM", "type": "HARVEST_EXACT",
              "subject":  r["search_term"], "campaign": r.get("campaign",""),
              "impact_spend": _safe(r.get("sales", 0)),
              "detail": f"{r.get('orders',0):.0f} orders, {r.get('acos',0):.1%} ACoS — add as EXACT at ${r.get('suggested_bid',0):.2f}"}
             for _, r in harvest_rows.iterrows()]
        ),
    }
    findings_path = args.output.replace(".xlsx", "_findings.json")
    with open(findings_path, "w") as _f:
        json.dump(findings, _f, indent=2, default=str)
    print(f"   Findings: {findings_path}")

    wb.save(args.output)
    print(f"\n✅ Output saved: {args.output}")
    print(f"   Harvest → Exact:  {gold_n} terms")
    print(f"   Harvest → Phrase: {silver_n} terms")
    neg_spend = df[df["category"].str.startswith("Negative")]["spend"].sum()
    print(f"   Negatives:        {df['category'].str.startswith('Negative').sum()} terms  (${neg_spend:,.2f} wasted spend)")
    print(f"   Review:           {(df['category']=='Review').sum()} terms")
    if args.bulk_file:
        print(f"   Duplicates (suppressed): {dup_n} terms already in your campaigns")


if __name__ == "__main__":
    main()
