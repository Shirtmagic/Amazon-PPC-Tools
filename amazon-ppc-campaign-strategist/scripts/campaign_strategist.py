#!/usr/bin/env python3
"""
Amazon PPC Campaign Strategist
Reads a Sponsored Products Bulk Operations file, classifies each campaign
by strategic goal, evaluates health, audits wasted spend, and outputs
a formatted Excel strategy workbook with prioritised weekly actions.
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")


# ── Colours ─────────────────────────────────────────────────────────────────
C = {
    "header_dark":   "1F3864",
    "header_green":  "1E6B3C",
    "header_red":    "8B0000",
    "header_amber":  "7B4F00",
    "header_pause":  "4A4A4A",
    "header_gray":   "555555",
    "header_blue":   "1B4F72",
    "header_purple": "4A235A",
    "header_teal":   "0E6655",
    "green_light":   "F0FFF4",
    "red_light":     "FFF0F0",
    "amber_light":   "FFFBF0",
    "gray_light":    "F8F8F8",
    "light_gray":    "F5F5F5",
    "blue_light":    "EBF5FB",
    "purple_light":  "F4ECF7",
    "teal_light":    "E8F8F5",
    "positive":      "1E6B3C",
    "negative":      "8B0000",
    "neutral":       "7B4F00",
    "white":         "FFFFFF",
}
FONT = "Arial"


def hex_fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, bold=True, color="FFFFFF"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def bfont(size=10, bold=False, color="1A1A1A"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def cal():
    return Alignment(horizontal="center", vertical="center")

def lal():
    return Alignment(horizontal="left", vertical="center")


# ── Goal Constants ──────────────────────────────────────────────────────────
GOALS = {
    "PROFIT":        {"label": "💰 Profit",        "color": C["header_green"]},
    "RANKING":       {"label": "🚀 Ranking",       "color": C["header_blue"]},
    "RESEARCH":      {"label": "🔬 Research",      "color": C["header_purple"]},
    "BRAND DEFENSE": {"label": "🏷 Brand Defense", "color": C["header_teal"]},
    "REVIEW":        {"label": "📝 Review",        "color": C["header_amber"]},
    "MARKET SHARE":  {"label": "📈 Market Share",  "color": C["header_dark"]},
    "DEAD":          {"label": "💀 Dead",          "color": C["header_pause"]},
}

HEALTH_LABELS = {
    "HEALTHY":          {"label": "✅ Healthy",          "color": C["positive"]},
    "OVER-PERFORMING":  {"label": "🌟 Over-Performing",  "color": C["header_blue"]},
    "UNDER-PERFORMING": {"label": "⚠ Under-Performing", "color": C["negative"]},
    "MISALIGNED":       {"label": "🔀 Misaligned",       "color": C["header_amber"]},
    "DEAD":             {"label": "💀 Dead",             "color": C["header_pause"]},
}


# ── Data Loading ────────────────────────────────────────────────────────────

def load_bulk(path):
    """Load campaign and keyword rows from the bulk operations file."""
    df = pd.read_excel(path, sheet_name="Sponsored Products Campaigns",
                       engine="openpyxl")

    camps = df[df["Entity"] == "Campaign"].copy()
    kws   = df[df["Entity"] == "Keyword"].copy()

    # Resolve informational columns
    for frame in [camps, kws]:
        for info_col, base_col in [
            ("Campaign Name (Informational only)", "Campaign Name"),
            ("Ad Group Name (Informational only)",  "Ad Group Name"),
            ("Portfolio Name (Informational only)", "Portfolio Name"),
        ]:
            if info_col in frame.columns:
                if base_col not in frame.columns:
                    frame[base_col] = frame[info_col]
                else:
                    frame[base_col] = frame[base_col].fillna(frame[info_col])

    # Coerce numerics
    numeric = ["Impressions", "Clicks", "Spend", "Sales",
               "Orders", "Units", "Bid", "Budget"]
    for col in numeric:
        for frame in [camps, kws]:
            if col in frame.columns:
                frame[col] = pd.to_numeric(frame[col], errors="coerce").fillna(0)

    # Compute ACoS
    for frame in [camps, kws]:
        frame["_acos"] = np.where(
            (frame["Sales"] > 0) & (frame["Spend"] > 0),
            frame["Spend"] / frame["Sales"],
            np.nan
        )

    return camps, kws


# ── Goal Classification ────────────────────────────────────────────────────

def classify_goal(camp_row, camp_kws, target_acos_profit, brand_patterns):
    """Assign a strategic goal to a campaign based on name, targeting, and ACoS."""
    name   = str(camp_row.get("Campaign Name", "")).lower()
    state  = str(camp_row.get("State", "")).lower()
    imps   = float(camp_row.get("Impressions", 0))
    clicks = float(camp_row.get("Clicks", 0))

    # DEAD: no impressions and enabled
    if imps == 0 and state == "enabled":
        return "DEAD"

    # RANKING: name signals
    ranking_signals = ["rank", "launch", "skc", "aggress"]
    if any(sig in name for sig in ranking_signals):
        return "RANKING"

    # BRAND DEFENSE: contains brand keywords
    if brand_patterns and any(bp in name for bp in brand_patterns):
        return "BRAND DEFENSE"

    # Analyse keyword-level match types for this campaign
    if len(camp_kws) > 0:
        match_types = set(camp_kws["Match Type"].str.lower().dropna().unique())
    else:
        match_types = set()

    # Check targeting type at campaign level
    targeting = str(camp_row.get("Targeting Type", "")).lower()

    # AUTO / RESEARCH: auto targeting or broad/phrase only
    if targeting == "auto" and len(camp_kws) == 0:
        return "RESEARCH"
    if match_types and match_types <= {"broad", "phrase"}:
        return "RESEARCH"

    # Exact-match campaigns: classify by ACoS performance
    if "exact" in match_types:
        acos = camp_row["_acos"]
        if not pd.isna(acos):
            if acos <= target_acos_profit * 1.1:
                return "PROFIT"
            elif acos > target_acos_profit * 2.0:
                return "REVIEW"

    # Market-share signals
    if any(sig in name for sig in ["market", "share", "conquest", "compete"]):
        return "MARKET SHARE"

    # Default: research if broad mix, profit if performing, review otherwise
    acos = camp_row["_acos"]
    if pd.isna(acos):
        return "RESEARCH"
    if acos <= target_acos_profit * 1.3:
        return "PROFIT"
    return "REVIEW"


def get_target_acos(goal, targets):
    """Return the target ACoS for the given goal."""
    return targets.get(goal, targets["PROFIT"])


# ── Health Evaluation ───────────────────────────────────────────────────────

def evaluate_health(camp_row, goal, target_acos):
    """Evaluate campaign health relative to goal target ACoS."""
    imps   = float(camp_row.get("Impressions", 0))
    clicks = float(camp_row.get("Clicks", 0))
    acos   = camp_row["_acos"]

    if imps == 0 or clicks == 0:
        return "DEAD"

    if pd.isna(acos):
        return "DEAD"

    if acos < target_acos * 0.6:
        return "OVER-PERFORMING"
    elif acos <= target_acos * 1.2:
        return "HEALTHY"
    elif acos <= target_acos * 1.5:
        return "UNDER-PERFORMING"
    else:
        return "MISALIGNED"


# ── Wasted Spend Audit ──────────────────────────────────────────────────────

def audit_duplicates(kws):
    """Find keywords targeted in multiple enabled campaigns."""
    enabled = kws[kws["State"].str.lower() == "enabled"].copy()
    if enabled.empty:
        return pd.DataFrame()

    enabled["_kw_lower"] = enabled["Keyword Text"].str.lower().str.strip()
    grouped = enabled.groupby("_kw_lower").agg(
        campaign_count=("Campaign Name", "nunique"),
        campaigns=("Campaign Name", lambda x: " | ".join(sorted(x.unique()))),
        total_spend=("Spend", "sum"),
        total_clicks=("Clicks", "sum"),
        total_orders=("Orders", "sum"),
        total_sales=("Sales", "sum"),
        match_types=("Match Type", lambda x: ", ".join(sorted(x.unique()))),
    ).reset_index()

    dupes = grouped[grouped["campaign_count"] > 1].copy()
    dupes = dupes.sort_values("total_spend", ascending=False)
    dupes.rename(columns={"_kw_lower": "keyword"}, inplace=True)

    # Estimate cannibalization: spend in worst-performing duplicate
    dupes["est_cannibalisation"] = dupes["total_spend"] * 0.3
    return dupes


def audit_zero_roi(camps, kws):
    """Find campaigns and keywords with spend but zero orders."""
    zero_camps = camps[
        (camps["Spend"] > 50) &
        (camps["Orders"] == 0) &
        (camps["State"].str.lower() == "enabled")
    ].copy()
    zero_camps = zero_camps.sort_values("Spend", ascending=False)

    zero_kws = kws[
        (kws["Spend"] > 10) &
        (kws["Clicks"] > 10) &
        (kws["Orders"] == 0) &
        (kws["State"].str.lower() == "enabled")
    ].copy()
    zero_kws = zero_kws.sort_values("Spend", ascending=False)

    return zero_camps, zero_kws


def audit_product_targeting(camps):
    """Identify product-targeting campaigns with high ACoS."""
    pt = camps[
        camps["Campaign Name"].str.lower().str.contains(
            "asin|product|pt|comp|competitor", na=False
        )
    ].copy()
    pt = pt[pt["_acos"].notna() & (pt["_acos"] > 0.4)]
    pt = pt.sort_values("_acos", ascending=False)
    return pt


# ── Day Parting Analysis ────────────────────────────────────────────────────

def analyse_day_parting(camps):
    """Flag campaigns with high budget utilisation as day parting candidates."""
    candidates = camps.copy()
    if "Budget" not in candidates.columns:
        candidates["Budget"] = 0

    candidates["Budget"] = pd.to_numeric(candidates["Budget"],
                                         errors="coerce").fillna(0)
    active = candidates[
        (candidates["State"].str.lower() == "enabled") &
        (candidates["Budget"] > 0) &
        (candidates["Spend"] > 0)
    ].copy()

    active["budget_util"] = active["Spend"] / (active["Budget"] * 30)
    active["budget_util"] = active["budget_util"].clip(upper=2.0)
    day_part = active[active["budget_util"] > 0.9].copy()
    day_part = day_part.sort_values("budget_util", ascending=False)
    return day_part


# ── Brand vs Non-Brand Split ────────────────────────────────────────────────

def split_brand(camps, brand_patterns):
    """Separate campaigns into brand and non-brand groups."""
    if not brand_patterns:
        return camps.copy(), pd.DataFrame(columns=camps.columns)

    mask = camps["Campaign Name"].str.lower().apply(
        lambda n: any(bp in str(n) for bp in brand_patterns)
    )
    brand     = camps[mask].copy()
    non_brand = camps[~mask].copy()
    return brand, non_brand


# ── Weekly Action Plan ──────────────────────────────────────────────────────

def build_action_plan(camp_analysis, dupes, zero_camps, zero_kws, day_part):
    """Generate a prioritised list of weekly actions."""
    actions = []

    # 1. Dead campaigns to pause
    dead = camp_analysis[camp_analysis["health"] == "DEAD"]
    for _, row in dead.iterrows():
        actions.append({
            "priority": "HIGH",
            "action":   "Pause or restructure dead campaign",
            "campaign": row.get("Campaign Name", ""),
            "detail":   f"0 impressions — wasting budget slot",
            "impact":   f"${float(row.get('Budget', 0)):.2f}/day saved",
        })

    # 2. Misaligned campaigns
    mis = camp_analysis[camp_analysis["health"] == "MISALIGNED"]
    for _, row in mis.iterrows():
        actions.append({
            "priority": "HIGH",
            "action":   f"Review misaligned campaign (goal: {row['goal']})",
            "campaign": row.get("Campaign Name", ""),
            "detail":   f"ACoS {row['_acos']:.1%} far from target {row['target_acos']:.0%}",
            "impact":   f"${float(row.get('Spend', 0)):.2f} at risk",
        })

    # 3. Zero-ROI campaigns
    for _, row in zero_camps.head(10).iterrows():
        actions.append({
            "priority": "HIGH",
            "action":   "Pause zero-ROI campaign",
            "campaign": row.get("Campaign Name", ""),
            "detail":   f"${float(row['Spend']):.2f} spent, 0 orders",
            "impact":   f"${float(row['Spend']):.2f} wasted",
        })

    # 4. Duplicate targeting
    for _, row in dupes.head(10).iterrows():
        actions.append({
            "priority": "MEDIUM",
            "action":   f"Consolidate duplicate keyword: {row['keyword']}",
            "campaign": row["campaigns"],
            "detail":   f"In {row['campaign_count']} campaigns, "
                        f"${float(row['total_spend']):.2f} total spend",
            "impact":   f"~${float(row['est_cannibalisation']):.2f} cannibalisation",
        })

    # 5. Zero-ROI keywords
    for _, row in zero_kws.head(10).iterrows():
        actions.append({
            "priority": "MEDIUM",
            "action":   "Negate or pause zero-ROI keyword",
            "campaign": row.get("Campaign Name", ""),
            "detail":   f"'{row.get('Keyword Text', '')}' — "
                        f"${float(row['Spend']):.2f} spent, "
                        f"{int(row['Clicks'])} clicks, 0 orders",
            "impact":   f"${float(row['Spend']):.2f} recoverable",
        })

    # 6. Day parting opportunities
    for _, row in day_part.head(5).iterrows():
        actions.append({
            "priority": "LOW",
            "action":   "Consider day parting schedule",
            "campaign": row.get("Campaign Name", ""),
            "detail":   f"Budget utilisation {row['budget_util']:.0%} — "
                        f"budget exhausting early",
            "impact":   "Improve spend distribution",
        })

    # 7. Over-performing campaigns — scale up
    over = camp_analysis[camp_analysis["health"] == "OVER-PERFORMING"]
    for _, row in over.head(5).iterrows():
        actions.append({
            "priority": "MEDIUM",
            "action":   "Scale budget on over-performing campaign",
            "campaign": row.get("Campaign Name", ""),
            "detail":   f"ACoS {row['_acos']:.1%} well under target "
                        f"{row['target_acos']:.0%} — room to grow",
            "impact":   f"Increase profitable volume",
        })

    return actions


# ── Excel Helpers ───────────────────────────────────────────────────────────

def apply_header(ws, row_num, headers, bg, fc="FFFFFF", height=20):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=i, value=h)
        c.font = hfont(color=fc)
        c.fill = hex_fill(bg)
        c.alignment = cal()
        c.border = thin_border()

def sheet_title(ws, title, subtitle, bg, height=28):
    ws.row_dimensions[1].height = height
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    max_col = max(ws.max_column or 1, 12)
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).fill = hex_fill(bg)
        ws.cell(row=1, column=col).border = thin_border()
    ws.row_dimensions[2].height = 16
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    s.fill = hex_fill("FFFFFF")

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Sheet Builders ──────────────────────────────────────────────────────────

def build_dashboard(wb, camp_analysis, targets, kws):
    """Strategy Dashboard with KPI cards and goal distribution."""
    ws = wb.create_sheet("📊 Strategy Dashboard")
    ws.sheet_view.showGridLines = False

    # Title
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "🎯  Amazon PPC Campaign Strategist Report"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hex_fill(C["header_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = (f"Generated: {datetime.today().strftime('%b %d, %Y')}   |   "
                f"Campaigns analysed: {len(camp_analysis)}   |   "
                f"Keywords: {len(kws)}")
    c2.font = Font(name=FONT, italic=True, size=9, color="888888")
    c2.fill = hex_fill("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Overall KPIs
    total_spend  = camp_analysis["Spend"].sum()
    total_sales  = camp_analysis["Sales"].sum()
    total_orders = camp_analysis["Orders"].sum()
    total_camps  = len(camp_analysis)
    overall_acos = total_spend / total_sales if total_sales > 0 else 0
    healthy_pct  = (len(camp_analysis[camp_analysis["health"] == "HEALTHY"])
                    / total_camps if total_camps > 0 else 0)

    kpis = [
        ("Total Spend",     f"${total_spend:,.2f}",     C["header_dark"]),
        ("Total Sales",     f"${total_sales:,.2f}",     C["header_green"]),
        ("Overall ACoS",    f"{overall_acos:.1%}",
         C["negative"] if overall_acos > targets["PROFIT"] else C["header_green"]),
        ("Total Orders",    f"{int(total_orders):,}",   C["header_dark"]),
        ("Campaigns",       f"{total_camps:,}",         C["header_dark"]),
        ("Health Score",    f"{healthy_pct:.0%}",
         C["header_green"] if healthy_pct > 0.6 else C["header_amber"]),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    for i, (label, value, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hex_fill("F5F5F5")
        lc.alignment = cal()
        lc.border = thin_border()
        vc = ws.cell(row=5, column=i, value=value)
        vc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = cal()
        vc.border = thin_border()

    # Goal distribution
    ws.row_dimensions[7].height = 18
    apply_header(ws, 7,
                 ["Strategic Goal", "Campaigns", "Spend", "Sales",
                  "ACoS", "Target ACoS", "Health Mix"],
                 C["header_dark"])

    goal_groups = camp_analysis.groupby("goal")
    r = 8
    for g_idx, goal in enumerate(GOALS.keys()):
        if goal not in goal_groups.groups:
            continue
        g = goal_groups.get_group(goal)
        g_spend = g["Spend"].sum()
        g_sales = g["Sales"].sum()
        g_acos  = g_spend / g_sales if g_sales > 0 else 0
        target  = get_target_acos(goal, targets)
        health_mix = ", ".join(
            f"{h}: {int(cnt)}"
            for h, cnt in g["health"].value_counts().items()
        )
        fill = hex_fill("FFFFFF") if g_idx % 2 == 0 else hex_fill(C["light_gray"])
        vals = [
            GOALS[goal]["label"], len(g), g_spend, g_sales,
            g_acos, target, health_mix,
        ]
        fmts = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                '0.0%', '0.0%', None]
        ws.row_dimensions[r].height = 18
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = bfont(bold=(c_idx == 1))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val not in ("", None):
                cell.number_format = fmt
        r += 1

    # Health breakdown
    r += 1
    ws.row_dimensions[r].height = 18
    apply_header(ws, r,
                 ["Health Status", "Campaigns", "Spend", "Sales",
                  "Avg ACoS", "", ""],
                 C["header_dark"])
    r += 1
    health_groups = camp_analysis.groupby("health")
    for h_idx, health in enumerate(HEALTH_LABELS.keys()):
        if health not in health_groups.groups:
            continue
        h = health_groups.get_group(health)
        h_spend = h["Spend"].sum()
        h_sales = h["Sales"].sum()
        h_acos  = h_spend / h_sales if h_sales > 0 else 0
        fill = hex_fill("FFFFFF") if h_idx % 2 == 0 else hex_fill(C["light_gray"])
        vals = [HEALTH_LABELS[health]["label"], len(h), h_spend, h_sales, h_acos]
        fmts = [None, None, '"$"#,##0.00', '"$"#,##0.00', '0.0%']
        ws.row_dimensions[r].height = 18
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = bfont(bold=(c_idx == 1))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val not in ("", None):
                cell.number_format = fmt
        r += 1

    ws.column_dimensions["G"].width = 40
    ws.freeze_panes = "A3"


def build_goal_sheet(wb, camp_analysis):
    """Every campaign with goal, target ACoS, health status."""
    ws = wb.create_sheet("🎯 Goal Classification")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Campaign Goal Classification",
                f"{len(camp_analysis)} campaigns classified by strategic intent",
                C["header_dark"])

    headers = ["Campaign Name", "State", "Goal", "Target ACoS",
               "Actual ACoS", "Health", "Impressions", "Clicks",
               "Spend", "Sales", "Orders", "Budget"]
    apply_header(ws, 3, headers, C["header_dark"])

    sorted_df = camp_analysis.sort_values(
        ["goal", "Spend"], ascending=[True, False])

    for r_idx, (_, row) in enumerate(sorted_df.iterrows()):
        er = 4 + r_idx
        goal = row.get("goal", "RESEARCH")
        goal_color = GOALS.get(goal, GOALS["RESEARCH"])["color"]
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[er].height = 15

        vals = [
            row.get("Campaign Name", ""),
            row.get("State", ""),
            GOALS.get(goal, {}).get("label", goal),
            row.get("target_acos", ""),
            row.get("_acos", ""),
            HEALTH_LABELS.get(row.get("health", ""), {}).get("label", row.get("health", "")),
            row.get("Impressions", 0),
            row.get("Clicks", 0),
            row.get("Spend", 0),
            row.get("Sales", 0),
            row.get("Orders", 0),
            row.get("Budget", 0),
        ]
        fmts = [None, None, None, '0.0%', '0.0%', None,
                '#,##0', '#,##0', '"$"#,##0.00', '"$"#,##0.00',
                '#,##0', '"$"#,##0.00']

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val != "":
                cell.number_format = fmt

        # Colour the goal cell
        goal_cell = ws.cell(row=er, column=3)
        goal_cell.font = Font(name=FONT, bold=True, size=9, color=goal_color)

    widths = {"A": 38, "B": 10, "C": 18, "D": 13, "E": 12,
              "F": 22, "G": 13, "H": 10, "I": 13, "J": 13,
              "K": 10, "L": 13}
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_misaligned_sheet(wb, camp_analysis):
    """Campaigns whose performance doesn't match their goal."""
    mis = camp_analysis[camp_analysis["health"].isin(
        ["MISALIGNED", "UNDER-PERFORMING"])].copy()
    mis = mis.sort_values("Spend", ascending=False)

    ws = wb.create_sheet("⚠ Misaligned Campaigns")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Misaligned & Under-Performing Campaigns",
                f"{len(mis)} campaigns need attention",
                C["header_red"])

    headers = ["Campaign Name", "Goal", "Target ACoS", "Actual ACoS",
               "Gap", "Health", "Spend", "Sales", "Orders",
               "Recommendation"]
    apply_header(ws, 3, headers, C["header_red"])

    for r_idx, (_, row) in enumerate(mis.iterrows()):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["red_light"])
        ws.row_dimensions[er].height = 15

        acos = row["_acos"] if not pd.isna(row["_acos"]) else 0
        target = row.get("target_acos", 0)
        gap = acos - target if target > 0 else 0
        health = row.get("health", "")

        if health == "MISALIGNED":
            rec = "Restructure or re-classify — ACoS far from goal target"
        else:
            rec = "Optimise bids/keywords — ACoS exceeds target by >50%"

        vals = [
            row.get("Campaign Name", ""),
            GOALS.get(row.get("goal", ""), {}).get("label", ""),
            target, acos, gap, health,
            row.get("Spend", 0), row.get("Sales", 0),
            row.get("Orders", 0), rec,
        ]
        fmts = [None, None, '0.0%', '0.0%', '+0.0%;-0.0%', None,
                '"$"#,##0.00', '"$"#,##0.00', '#,##0', None]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val != "":
                cell.number_format = fmt

        # Colour the gap cell
        gap_cell = ws.cell(row=er, column=5)
        if isinstance(gap, (int, float)) and gap > 0:
            gap_cell.font = Font(name=FONT, bold=True, size=9,
                                 color=C["negative"])

    widths = {"A": 38, "B": 18, "C": 13, "D": 12, "E": 10,
              "F": 22, "G": 13, "H": 13, "I": 10, "J": 48}
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_wasted_spend_sheet(wb, dupes, zero_camps, zero_kws, pt_camps):
    """Wasted spend audit with duplicate, zero-ROI, and product targeting."""
    ws = wb.create_sheet("💸 Wasted Spend Audit")
    ws.sheet_view.showGridLines = False

    total_waste = (dupes["est_cannibalisation"].sum() if not dupes.empty else 0)
    total_waste += zero_camps["Spend"].sum() if not zero_camps.empty else 0
    total_waste += zero_kws["Spend"].sum() if not zero_kws.empty else 0

    sheet_title(ws, "Wasted Spend Audit",
                f"~${total_waste:,.2f} estimated wasted/at-risk spend identified",
                C["header_red"])

    r = 3
    # ── Section 1: Duplicate Targeting ──
    ws.row_dimensions[r].height = 20
    sec = ws.cell(row=r, column=1,
                  value=f"🔄 Duplicate Targeting ({len(dupes)} keywords)")
    sec.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    sec.fill = hex_fill(C["header_amber"])
    for col in range(2, 8):
        ws.cell(row=r, column=col).fill = hex_fill(C["header_amber"])
    r += 1

    if not dupes.empty:
        dupe_hdrs = ["Keyword", "# Campaigns", "Campaigns",
                     "Total Spend", "Total Orders", "Match Types",
                     "Est. Cannibalisation"]
        apply_header(ws, r, dupe_hdrs, C["header_amber"])
        r += 1
        for d_idx, (_, row) in enumerate(dupes.head(20).iterrows()):
            fill = hex_fill("FFFFFF") if d_idx % 2 == 0 else hex_fill(C["amber_light"])
            ws.row_dimensions[r].height = 15
            vals = [row["keyword"], row["campaign_count"], row["campaigns"],
                    row["total_spend"], row["total_orders"],
                    row["match_types"], row["est_cannibalisation"]]
            fmts = [None, None, None, '"$"#,##0.00', '#,##0', None,
                    '"$"#,##0.00']
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font = bfont(size=9)
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = lal()
                if fmt and val != "":
                    cell.number_format = fmt
            r += 1

    r += 1
    # ── Section 2: Zero-ROI Campaigns ──
    ws.row_dimensions[r].height = 20
    sec = ws.cell(row=r, column=1,
                  value=f"🚫 Zero-ROI Campaigns ({len(zero_camps)} campaigns)")
    sec.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    sec.fill = hex_fill(C["header_red"])
    for col in range(2, 8):
        ws.cell(row=r, column=col).fill = hex_fill(C["header_red"])
    r += 1

    if not zero_camps.empty:
        zc_hdrs = ["Campaign Name", "Spend", "Clicks", "Impressions",
                   "State", "Budget", ""]
        apply_header(ws, r, zc_hdrs, C["header_red"])
        r += 1
        for z_idx, (_, row) in enumerate(zero_camps.head(15).iterrows()):
            fill = hex_fill("FFFFFF") if z_idx % 2 == 0 else hex_fill(C["red_light"])
            ws.row_dimensions[r].height = 15
            vals = [row.get("Campaign Name", ""), row["Spend"],
                    row["Clicks"], row["Impressions"],
                    row.get("State", ""), row.get("Budget", 0), ""]
            fmts = [None, '"$"#,##0.00', '#,##0', '#,##0', None,
                    '"$"#,##0.00', None]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font = bfont(size=9)
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = lal()
                if fmt and val != "":
                    cell.number_format = fmt
            r += 1

    r += 1
    # ── Section 3: Zero-ROI Keywords ──
    ws.row_dimensions[r].height = 20
    sec = ws.cell(row=r, column=1,
                  value=f"🔑 Zero-ROI Keywords ({len(zero_kws)} keywords)")
    sec.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    sec.fill = hex_fill(C["header_red"])
    for col in range(2, 8):
        ws.cell(row=r, column=col).fill = hex_fill(C["header_red"])
    r += 1

    if not zero_kws.empty:
        zk_hdrs = ["Keyword Text", "Campaign", "Match Type",
                   "Spend", "Clicks", "Impressions", ""]
        apply_header(ws, r, zk_hdrs, C["header_red"])
        r += 1
        for z_idx, (_, row) in enumerate(zero_kws.head(20).iterrows()):
            fill = hex_fill("FFFFFF") if z_idx % 2 == 0 else hex_fill(C["red_light"])
            ws.row_dimensions[r].height = 15
            vals = [row.get("Keyword Text", ""),
                    row.get("Campaign Name", ""),
                    row.get("Match Type", ""),
                    row["Spend"], row["Clicks"],
                    row.get("Impressions", 0), ""]
            fmts = [None, None, None, '"$"#,##0.00', '#,##0', '#,##0', None]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font = bfont(size=9)
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = lal()
                if fmt and val != "":
                    cell.number_format = fmt
            r += 1

    r += 1
    # ── Section 4: Product Targeting Overspend ──
    ws.row_dimensions[r].height = 20
    sec = ws.cell(row=r, column=1,
                  value=f"📦 Product Targeting Overspend ({len(pt_camps)} campaigns)")
    sec.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    sec.fill = hex_fill(C["header_amber"])
    for col in range(2, 8):
        ws.cell(row=r, column=col).fill = hex_fill(C["header_amber"])
    r += 1

    if not pt_camps.empty:
        pt_hdrs = ["Campaign Name", "ACoS", "Spend", "Sales",
                   "Orders", "Clicks", ""]
        apply_header(ws, r, pt_hdrs, C["header_amber"])
        r += 1
        for p_idx, (_, row) in enumerate(pt_camps.head(10).iterrows()):
            fill = hex_fill("FFFFFF") if p_idx % 2 == 0 else hex_fill(C["amber_light"])
            ws.row_dimensions[r].height = 15
            vals = [row.get("Campaign Name", ""), row["_acos"],
                    row["Spend"], row["Sales"],
                    row["Orders"], row["Clicks"], ""]
            fmts = [None, '0.0%', '"$"#,##0.00', '"$"#,##0.00',
                    '#,##0', '#,##0', None]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                if isinstance(val, float) and np.isnan(val):
                    val = ""
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font = bfont(size=9)
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = lal()
                if fmt and val != "":
                    cell.number_format = fmt
            r += 1

    widths = {"A": 38, "B": 14, "C": 40, "D": 14, "E": 12,
              "F": 14, "G": 18}
    set_widths(ws, widths)
    ws.freeze_panes = "A3"


def build_day_parting_sheet(wb, day_part):
    """Day parting opportunities."""
    ws = wb.create_sheet("🕐 Day Parting Opportunities")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Day Parting Opportunities",
                f"{len(day_part)} campaigns exhausting budget — "
                f"consider time-based scheduling",
                C["header_blue"])

    headers = ["Campaign Name", "Daily Budget", "Avg Daily Spend",
               "Budget Utilisation", "ACoS", "Spend", "Sales",
               "Orders", "Recommendation"]
    apply_header(ws, 3, headers, C["header_blue"])

    for r_idx, (_, row) in enumerate(day_part.iterrows()):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["blue_light"])
        ws.row_dimensions[er].height = 15

        budget = row.get("Budget", 0)
        util   = row.get("budget_util", 0)
        acos   = row["_acos"] if not pd.isna(row["_acos"]) else ""

        if util > 1.0:
            rec = "Budget capped daily — increase budget or add day parting"
        else:
            rec = "Near budget cap — schedule ads for peak conversion hours"

        vals = [row.get("Campaign Name", ""), budget,
                row["Spend"] / 30, util, acos,
                row["Spend"], row["Sales"], row["Orders"], rec]
        fmts = [None, '"$"#,##0.00', '"$"#,##0.00', '0.0%', '0.0%',
                '"$"#,##0.00', '"$"#,##0.00', '#,##0', None]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val != "":
                cell.number_format = fmt

    widths = {"A": 38, "B": 14, "C": 15, "D": 18, "E": 10,
              "F": 13, "G": 13, "H": 10, "I": 48}
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_brand_sheet(wb, brand, non_brand):
    """Brand vs Non-Brand performance split."""
    ws = wb.create_sheet("🏷 Brand vs Non-Brand")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "Brand vs Non-Brand Performance",
                f"{len(brand)} brand campaigns | "
                f"{len(non_brand)} non-brand campaigns",
                C["header_teal"])

    # Summary cards
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    segments = [
        ("Brand", brand, C["header_teal"]),
        ("Non-Brand", non_brand, C["header_dark"]),
    ]
    col_offset = 0
    for seg_name, seg_df, seg_color in segments:
        s_spend = seg_df["Spend"].sum() if not seg_df.empty else 0
        s_sales = seg_df["Sales"].sum() if not seg_df.empty else 0
        s_orders = seg_df["Orders"].sum() if not seg_df.empty else 0
        s_acos = s_spend / s_sales if s_sales > 0 else 0

        metrics = [
            (f"{seg_name} Spend", f"${s_spend:,.2f}"),
            (f"{seg_name} Sales", f"${s_sales:,.2f}"),
            (f"{seg_name} ACoS",  f"{s_acos:.1%}"),
            (f"{seg_name} Orders", f"{int(s_orders):,}"),
        ]
        for m_idx, (label, value) in enumerate(metrics):
            ci = col_offset + m_idx + 1
            lc = ws.cell(row=4, column=ci, value=label)
            lc.font = Font(name=FONT, size=8, color="888888")
            lc.fill = hex_fill("F5F5F5")
            lc.alignment = cal()
            lc.border = thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = 18
            vc = ws.cell(row=5, column=ci, value=value)
            vc.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
            vc.fill = hex_fill(seg_color)
            vc.alignment = cal()
            vc.border = thin_border()
        col_offset += 4

    # Brand campaigns
    r = 7
    ws.row_dimensions[r].height = 20
    sec = ws.cell(row=r, column=1,
                  value=f"🏷 Brand Campaigns ({len(brand)})")
    sec.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    sec.fill = hex_fill(C["header_teal"])
    for col in range(2, 9):
        ws.cell(row=r, column=col).fill = hex_fill(C["header_teal"])
    r += 1

    camp_hdrs = ["Campaign Name", "ACoS", "Spend", "Sales",
                 "Orders", "Clicks", "Impressions", "Goal"]
    apply_header(ws, r, camp_hdrs, C["header_teal"])
    r += 1

    for c_idx, (_, row) in enumerate(brand.sort_values(
            "Spend", ascending=False).iterrows()):
        fill = hex_fill("FFFFFF") if c_idx % 2 == 0 else hex_fill(C["teal_light"])
        ws.row_dimensions[r].height = 15
        acos = row["_acos"] if not pd.isna(row["_acos"]) else ""
        vals = [row.get("Campaign Name", ""), acos,
                row["Spend"], row["Sales"], row["Orders"],
                row["Clicks"], row["Impressions"],
                GOALS.get(row.get("goal", ""), {}).get("label", "")]
        fmts = [None, '0.0%', '"$"#,##0.00', '"$"#,##0.00',
                '#,##0', '#,##0', '#,##0', None]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val != "":
                cell.number_format = fmt
        r += 1

    r += 1
    ws.row_dimensions[r].height = 20
    sec = ws.cell(row=r, column=1,
                  value=f"🎯 Non-Brand Campaigns ({len(non_brand)})")
    sec.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    sec.fill = hex_fill(C["header_dark"])
    for col in range(2, 9):
        ws.cell(row=r, column=col).fill = hex_fill(C["header_dark"])
    r += 1

    apply_header(ws, r, camp_hdrs, C["header_dark"])
    r += 1

    for c_idx, (_, row) in enumerate(non_brand.sort_values(
            "Spend", ascending=False).iterrows()):
        fill = hex_fill("FFFFFF") if c_idx % 2 == 0 else hex_fill(C["light_gray"])
        ws.row_dimensions[r].height = 15
        acos = row["_acos"] if not pd.isna(row["_acos"]) else ""
        vals = [row.get("Campaign Name", ""), acos,
                row["Spend"], row["Sales"], row["Orders"],
                row["Clicks"], row["Impressions"],
                GOALS.get(row.get("goal", ""), {}).get("label", "")]
        fmts = [None, '0.0%', '"$"#,##0.00', '"$"#,##0.00',
                '#,##0', '#,##0', '#,##0', None]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val != "":
                cell.number_format = fmt
        r += 1

    widths = {"A": 38, "B": 12, "C": 14, "D": 14, "E": 10,
              "F": 10, "G": 14, "H": 18}
    set_widths(ws, widths)
    ws.freeze_panes = "A3"


def build_action_plan_sheet(wb, actions):
    """Weekly action plan with prioritised recommendations."""
    ws = wb.create_sheet("📋 Weekly Action Plan")
    ws.sheet_view.showGridLines = False

    high   = [a for a in actions if a["priority"] == "HIGH"]
    medium = [a for a in actions if a["priority"] == "MEDIUM"]
    low    = [a for a in actions if a["priority"] == "LOW"]

    sheet_title(ws, "Weekly Action Plan",
                f"{len(high)} high | {len(medium)} medium | "
                f"{len(low)} low priority actions",
                C["header_dark"])

    headers = ["Priority", "Action", "Campaign", "Detail", "Impact"]
    apply_header(ws, 3, headers, C["header_dark"])

    priority_colors = {
        "HIGH":   C["negative"],
        "MEDIUM": C["neutral"],
        "LOW":    C["header_gray"],
    }
    priority_bg = {
        "HIGH":   C["red_light"],
        "MEDIUM": C["amber_light"],
        "LOW":    C["light_gray"],
    }

    sorted_actions = high + medium + low
    for r_idx, act in enumerate(sorted_actions):
        er = 4 + r_idx
        pri = act["priority"]
        fill = hex_fill(priority_bg.get(pri, "FFFFFF")) if r_idx % 2 == 0 \
            else hex_fill("FFFFFF")
        ws.row_dimensions[er].height = 16

        vals = [pri, act["action"], act["campaign"],
                act["detail"], act["impact"]]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()

        # Colour the priority cell
        pri_cell = ws.cell(row=er, column=1)
        pri_cell.font = Font(name=FONT, bold=True, size=9,
                             color=priority_colors.get(pri, "1A1A1A"))

    widths = {"A": 12, "B": 40, "C": 38, "D": 50, "E": 28}
    set_widths(ws, widths)
    ws.freeze_panes = "A4"


def build_raw_sheet(wb, camp_analysis):
    """Raw campaign data with all computed fields."""
    ws = wb.create_sheet("🗂 Raw Data")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "All Campaigns — Full Analysis",
                "Every campaign with goal classification and health assessment",
                C["header_dark"])

    cols = ["Campaign Name", "State", "goal", "target_acos", "_acos",
            "health", "Impressions", "Clicks", "Spend", "Sales",
            "Orders", "Budget"]
    hdrs = ["Campaign Name", "State", "Goal", "Target ACoS", "ACoS",
            "Health", "Impressions", "Clicks", "Spend", "Sales",
            "Orders", "Budget"]
    fmts_map = {
        "target_acos": '0.0%', "_acos": '0.0%',
        "Spend": '"$"#,##0.00', "Sales": '"$"#,##0.00',
        "Budget": '"$"#,##0.00',
    }

    apply_header(ws, 3, hdrs, C["header_dark"])

    health_bg = {
        "HEALTHY":          C["green_light"],
        "OVER-PERFORMING":  C["blue_light"],
        "UNDER-PERFORMING": C["red_light"],
        "MISALIGNED":       C["amber_light"],
        "DEAD":             "EEEEEE",
    }

    sorted_df = camp_analysis.sort_values(
        ["health", "Spend"],
        key=lambda x: x.map(
            {"DEAD": 0, "MISALIGNED": 1, "UNDER-PERFORMING": 2,
             "OVER-PERFORMING": 3, "HEALTHY": 4}
        ) if x.name == "health" else x,
        ascending=[True, False]
    )

    for r_idx, (_, row) in enumerate(sorted_df.iterrows()):
        er = 4 + r_idx
        h = row.get("health", "HEALTHY")
        fill = hex_fill(health_bg.get(h, "FFFFFF"))
        ws.row_dimensions[er].height = 14
        for c_idx, col in enumerate(cols, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in fmts_map and val != "":
                cell.number_format = fmts_map[col]

    for i, w in enumerate(
        [38, 10, 18, 13, 10, 22, 13, 10, 13, 13, 10, 13], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Amazon PPC Campaign Strategist")
    parser.add_argument("--input",                required=True)
    parser.add_argument("--output",               required=True)
    parser.add_argument("--target-acos-profit",   type=float, default=0.25)
    parser.add_argument("--target-acos-ranking",  type=float, default=0.60)
    parser.add_argument("--target-acos-research", type=float, default=0.35)
    parser.add_argument("--target-acos-reviews",  type=float, default=0.40)
    parser.add_argument("--target-acos-marketshare", type=float, default=0.30)
    parser.add_argument("--brand-keywords",       type=str, default="")
    parser.add_argument("--days",                 type=int, default=30)
    args = parser.parse_args()

    targets = {
        "PROFIT":        args.target_acos_profit,
        "RANKING":       args.target_acos_ranking,
        "RESEARCH":      args.target_acos_research,
        "REVIEW":        args.target_acos_reviews,
        "BRAND DEFENSE": args.target_acos_profit * 0.8,
        "MARKET SHARE":  args.target_acos_marketshare,
        "DEAD":          0.0,
    }

    brand_patterns = [b.strip().lower() for b in args.brand_keywords.split(",")
                      if b.strip()] if args.brand_keywords else []

    print(f"Loading bulk file: {args.input}")
    camps, kws = load_bulk(args.input)
    print(f"  {len(camps)} campaign rows loaded")
    print(f"  {len(kws)} keyword rows loaded")
    if brand_patterns:
        print(f"  Brand keywords: {brand_patterns}")

    # ── Classify each campaign ──────────────────────────────────────────
    print("\nClassifying campaign goals...")
    goal_list   = []
    target_list = []
    health_list = []

    for _, camp_row in camps.iterrows():
        camp_name = camp_row.get("Campaign Name", "")
        camp_kws  = kws[kws["Campaign Name"] == camp_name]

        goal = classify_goal(camp_row, camp_kws,
                             args.target_acos_profit, brand_patterns)
        target_acos = get_target_acos(goal, targets)
        health = evaluate_health(camp_row, goal, target_acos)

        goal_list.append(goal)
        target_list.append(target_acos)
        health_list.append(health)

    camps["goal"]        = goal_list
    camps["target_acos"] = target_list
    camps["health"]      = health_list

    goal_counts = camps["goal"].value_counts()
    print("\nGoal Distribution:")
    for goal, count in goal_counts.items():
        print(f"  {GOALS.get(goal, {}).get('label', goal)}: {count}")

    health_counts = camps["health"].value_counts()
    print("\nHealth Distribution:")
    for health, count in health_counts.items():
        print(f"  {HEALTH_LABELS.get(health, {}).get('label', health)}: {count}")

    # ── Wasted spend audit ──────────────────────────────────────────────
    print("\nRunning wasted spend audit...")
    dupes = audit_duplicates(kws)
    zero_camps, zero_kws = audit_zero_roi(camps, kws)
    pt_camps = audit_product_targeting(camps)
    print(f"  Duplicate keywords: {len(dupes)}")
    print(f"  Zero-ROI campaigns: {len(zero_camps)}")
    print(f"  Zero-ROI keywords:  {len(zero_kws)}")
    print(f"  Product targeting overspend: {len(pt_camps)}")

    # ── Day parting ─────────────────────────────────────────────────────
    print("\nAnalysing day parting opportunities...")
    day_part = analyse_day_parting(camps)
    print(f"  Day parting candidates: {len(day_part)}")

    # ── Brand split ─────────────────────────────────────────────────────
    brand, non_brand = split_brand(camps, brand_patterns)
    print(f"\nBrand campaigns: {len(brand)}")
    print(f"Non-brand campaigns: {len(non_brand)}")

    # ── Action plan ─────────────────────────────────────────────────────
    actions = build_action_plan(camps, dupes, zero_camps, zero_kws, day_part)
    print(f"\nAction plan: {len(actions)} items")

    # ── Build Excel ─────────────────────────────────────────────────────
    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb, camps, targets, kws)
    build_goal_sheet(wb, camps)
    build_misaligned_sheet(wb, camps)
    build_wasted_spend_sheet(wb, dupes, zero_camps, zero_kws, pt_camps)
    build_day_parting_sheet(wb, day_part)
    build_brand_sheet(wb, brand, non_brand)
    build_action_plan_sheet(wb, actions)
    build_raw_sheet(wb, camps)

    wb.save(args.output)
    print(f"\n✅ Saved: {args.output}")

    # ── Write findings JSON ─────────────────────────────────────────────
    total_waste = (
        (dupes["est_cannibalisation"].sum() if not dupes.empty else 0)
        + (zero_camps["Spend"].sum() if not zero_camps.empty else 0)
        + (zero_kws["Spend"].sum() if not zero_kws.empty else 0)
    )
    findings = {
        "tool":               "campaign_strategist",
        "campaigns_analysed": int(len(camps)),
        "keywords_analysed":  int(len(kws)),
        "goal_distribution":  {g: int(c) for g, c in goal_counts.items()},
        "health_distribution":{h: int(c) for h, c in health_counts.items()},
        "total_spend":        round(float(camps["Spend"].sum()), 2),
        "total_sales":        round(float(camps["Sales"].sum()), 2),
        "overall_acos":       round(float(
            camps["Spend"].sum() / camps["Sales"].sum()
        ) if camps["Sales"].sum() > 0 else 0, 4),
        "wasted_spend_est":   round(float(total_waste), 2),
        "duplicate_keywords": int(len(dupes)),
        "zero_roi_campaigns": int(len(zero_camps)),
        "zero_roi_keywords":  int(len(zero_kws)),
        "day_parting_candidates": int(len(day_part)),
        "brand_campaigns":    int(len(brand)),
        "non_brand_campaigns":int(len(non_brand)),
        "targets":            {k: round(v, 4) for k, v in targets.items()},
        "actions": [
            {"priority": a["priority"], "action": a["action"],
             "campaign": a["campaign"], "detail": a["detail"],
             "impact": a["impact"]}
            for a in actions[:20]
        ],
    }

    findings_path = args.output.replace(".xlsx", "_findings.json")
    with open(findings_path, "w") as _f:
        json.dump(findings, _f, indent=2, default=str)
    print(f"   Findings: {findings_path}")

    # Summary
    print(f"\n   📊 Dashboard: goal & health overview")
    print(f"   🎯 Classified: {len(camps)} campaigns")
    mis_count = len(camps[camps["health"].isin(
        ["MISALIGNED", "UNDER-PERFORMING"])])
    print(f"   ⚠  Misaligned: {mis_count} campaigns")
    print(f"   💸 Wasted: ~${total_waste:,.2f} identified")
    print(f"   🕐 Day parting: {len(day_part)} candidates")
    print(f"   📋 Actions: {len(actions)} items")


if __name__ == "__main__":
    main()
