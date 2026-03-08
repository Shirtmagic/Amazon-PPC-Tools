#!/usr/bin/env python3
"""
Amazon PPC Budget Manager
Reads a Sponsored Products Bulk Operations file and outputs
a formatted Excel action file with campaign budget recommendations.
"""

import argparse
import json
import sys
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── Colours ──────────────────────────────────────────────────────────────────
C = {
    "header_dark":  "1F3864",
    "header_green": "1E6B3C",
    "header_red":   "8B0000",
    "header_amber": "7B4F00",
    "header_alert": "5C0A0A",
    "header_realloc": "1A4A7A",
    "header_gray":  "555555",
    "green_light":  "F0FFF4",
    "red_light":    "FFF0F0",
    "amber_light":  "FFFBF0",
    "blue_light":   "F0F8FF",
    "light_gray":   "F5F5F5",
    "positive":     "1E6B3C",
    "negative":     "8B0000",
    "neutral":      "7B4F00",
    "white":        "FFFFFF",
}
FONT = "Arial"

def hex_fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def thin_border():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, bold=True, color="FFFFFF"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def bfont(size=10, bold=False, color="1A1A1A"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def cal():
    return Alignment(horizontal="center", vertical="center")

def lal():
    return Alignment(horizontal="left", vertical="center")

# ── Data Loading ──────────────────────────────────────────────────────────────

def load_campaigns(path, days):
    df = pd.read_excel(path, sheet_name="Sponsored Products Campaigns",
                       engine="openpyxl")
    camps = df[df["Entity"] == "Campaign"].copy()

    # Resolve names from informational columns
    for info_col, base_col in [
        ("Campaign Name (Informational only)", "Campaign Name"),
        ("Portfolio Name (Informational only)", "Portfolio Name"),
    ]:
        if info_col in camps.columns:
            if base_col not in camps.columns:
                camps[base_col] = camps[info_col]
            else:
                camps[base_col] = camps[base_col].fillna(camps[info_col])

    for col in ["Daily Budget", "Spend", "Sales", "Orders", "Clicks",
                "Impressions", "ACOS", "CPC", "ROAS"]:
        if col in camps.columns:
            camps[col] = pd.to_numeric(camps[col], errors="coerce").fillna(0)

    required = ["Campaign Name", "Daily Budget", "Spend", "State"]
    missing = [c for c in required if c not in camps.columns]
    if missing:
        sys.exit(f"ERROR: Missing columns: {missing}")

    camps["days"] = days
    camps["daily_spend"] = camps["Spend"] / days
    camps["budget_util"] = np.where(
        camps["Daily Budget"] > 0,
        camps["daily_spend"] / camps["Daily Budget"],
        np.nan
    )
    camps["_acos"] = np.where(
        (camps["Spend"] > 0) & (camps["Sales"] > 0),
        camps["Spend"] / camps["Sales"],
        np.nan
    )
    camps["budget_headroom"] = camps["Daily Budget"] - camps["daily_spend"]
    camps["unused_budget_30d"] = camps["budget_headroom"] * days
    return camps


# ── Classification ────────────────────────────────────────────────────────────

def classify(df, target_acos, constrained_thresh, underutil_thresh,
             increase_pct, decrease_pct, min_budget):
    actions, new_budgets, reasons = [], [], []

    for _, row in df.iterrows():
        util   = row.get("budget_util", np.nan)
        acos   = row.get("_acos", np.nan)
        spend  = row.get("Spend", 0)
        state  = str(row.get("State", "enabled")).lower()
        budget = row.get("Daily Budget", 0)
        daily  = row.get("daily_spend", 0)

        # Zero spend — enabled but not serving
        if spend == 0 and state == "enabled":
            actions.append("ALERT")
            new_budgets.append(budget)
            reasons.append("Enabled but zero spend — check bids and targeting")
            continue

        # Not enough data
        if pd.isna(util):
            actions.append("NO CHANGE")
            new_budgets.append(budget)
            reasons.append("No budget or spend data")
            continue

        # Budget-constrained
        if util >= constrained_thresh:
            if not pd.isna(acos) and acos <= target_acos * 1.5:
                new_b = round(budget * (1 + increase_pct), 2)
                actions.append("INCREASE")
                new_budgets.append(new_b)
                reasons.append(
                    f"Hitting {util:.0%} of budget — ACoS {acos:.1%} is efficient, "
                    f"increase budget to capture more volume"
                )
            else:
                new_b = round(max(budget * (1 - decrease_pct), min_budget), 2)
                actions.append("DECREASE")
                new_budgets.append(new_b)
                acos_str = f"{acos:.1%}" if not pd.isna(acos) else "N/A"
                reasons.append(
                    f"Hitting {util:.0%} of budget — ACoS {acos_str} is inefficient, "
                    f"reduce budget to cut losses"
                )
            continue

        # Severely underutilized
        if util <= underutil_thresh and spend > 0:
            suggested = round(max(daily * 1.5, min_budget), 2)
            actions.append("INVESTIGATE")
            new_budgets.append(suggested)
            reasons.append(
                f"Only using {util:.1%} of ${budget:.2f} daily budget "
                f"(${daily:.2f}/day avg) — consider right-sizing to ${suggested:.2f}"
            )
            continue

        actions.append("NO CHANGE")
        new_budgets.append(budget)
        acos_str = f"{acos:.1%}" if not pd.isna(acos) else "N/A"
        reasons.append(f"Utilization {util:.0%} within normal range — ACoS {acos_str}")

    df = df.copy()
    df["action"]     = actions
    df["new_budget"] = new_budgets
    df["reason"]     = reasons
    df["budget_change"] = df["new_budget"] - df["Daily Budget"]
    df["budget_change_pct"] = np.where(
        df["Daily Budget"] > 0,
        df["budget_change"] / df["Daily Budget"],
        np.nan
    )
    return df


# ── Excel Helpers ─────────────────────────────────────────────────────────────

def apply_header(ws, row_num, headers, bg, fc="FFFFFF", height=20):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=i, value=h)
        c.font = hfont(color=fc)
        c.fill = hex_fill(bg)
        c.alignment = cal()
        c.border = thin_border()

def sheet_title(ws, title, subtitle, bg, height=28):
    ws.row_dimensions[1].height = height
    max_col = max(ws.max_column or 1, 12)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=13, color="FFFFFF")
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(2, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hex_fill(bg)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 16
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, italic=True, size=9, color="AAAAAA")
    s.fill = hex_fill("FFFFFF")

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Sheet Builders ────────────────────────────────────────────────────────────

CAMP_COLS = ["Portfolio Name", "Campaign Name", "State",
             "Daily Budget", "daily_spend", "budget_util",
             "Spend", "Sales", "Orders", "_acos",
             "new_budget", "budget_change", "budget_change_pct", "reason"]
CAMP_HDRS = ["Portfolio", "Campaign", "State",
             "Current Daily Budget", "Avg Daily Spend", "Budget Utilization",
             "30d Spend", "30d Sales", "30d Orders", "ACoS",
             "New Budget", "Budget Change $", "Budget Change %", "Reason"]
CAMP_FMTS = {
    "Daily Budget":       '"$"#,##0.00',
    "daily_spend":        '"$"#,##0.00',
    "budget_util":        '0.0%',
    "Spend":              '"$"#,##0.00',
    "Sales":              '"$"#,##0.00',
    "_acos":              '0.0%',
    "new_budget":         '"$"#,##0.00',
    "budget_change":      '+$#,##0.00;-$#,##0.00;$-',
    "budget_change_pct":  '+0.0%;-0.0%;0.0%',
}


def write_action_sheet(wb, df, action, tab_name, bg, light_bg,
                       sort_col="Spend", sort_asc=False):
    data = df[df["action"] == action].copy()
    if not data.empty:
        data = data.sort_values(sort_col, ascending=sort_asc)

    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    count = len(data)
    budget_impact = data["budget_change"].sum() if count else 0
    subtitle = (f"{count} campaigns | "
                f"Total budget impact: ${budget_impact:+,.2f}/day")
    sheet_title(ws, tab_name.split(" ", 1)[-1], subtitle, bg)
    apply_header(ws, 3, CAMP_HDRS, bg)

    for r_idx, (_, row) in enumerate(data.iterrows()):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(light_bg)
        ws.row_dimensions[er].height = 16
        for c_idx, col in enumerate(CAMP_COLS, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in CAMP_FMTS and val not in ("", None):
                cell.number_format = CAMP_FMTS[col]

        # Colour the budget change columns
        chg_cell = ws.cell(row=er, column=12)
        pct_cell = ws.cell(row=er, column=13)
        chg_val = row.get("budget_change", 0)
        color = C["positive"] if chg_val > 0 else C["negative"]
        chg_cell.font = Font(name=FONT, bold=True, size=9, color=color)
        pct_cell.font = Font(name=FONT, bold=True, size=9, color=color)

    widths = {
        "A": 24, "B": 36, "C": 10, "D": 20, "E": 18,
        "F": 18, "G": 14, "H": 14, "I": 12, "J": 10,
        "K": 16, "L": 16, "M": 16, "N": 48,
    }
    set_widths(ws, widths)
    ws.freeze_panes = "A4"
    return count


def build_summary(wb, df, target_acos, days):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "💰  Amazon PPC Budget Manager Report"
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = hex_fill(C["header_dark"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = (f"Lookback: {days} days   |   Target ACoS: {target_acos:.0%}   |   "
                f"Generated: {datetime.today().strftime('%b %d, %Y')}")
    c2.font = Font(name=FONT, italic=True, size=9, color="888888")
    c2.fill = hex_fill("F8F8F8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # KPIs
    total_budget  = df["Daily Budget"].sum()
    total_spend   = df["Spend"].sum()
    total_sales   = df["Sales"].sum()
    total_orders  = int(df["Orders"].sum())
    total_daily   = df["daily_spend"].sum()
    unused_daily  = total_budget - total_daily
    overall_acos  = total_spend / total_sales if total_sales > 0 else 0
    increase_oppt = df[df["action"] == "INCREASE"]["budget_change"].sum()
    wasted_spend  = df[df["action"] == "DECREASE"]["Spend"].sum()

    kpis = [
        ("Total Daily Budget",  f"${total_budget:,.2f}",  C["header_dark"]),
        ("Avg Daily Spend",     f"${total_daily:,.2f}",   C["header_dark"]),
        ("Unused Budget/Day",   f"${unused_daily:,.2f}",  C["neutral"]),
        ("30d Total Spend",     f"${total_spend:,.2f}",   C["header_dark"]),
        ("30d Total Sales",     f"${total_sales:,.2f}",   C["header_green"]),
        ("Overall ACoS",        f"{overall_acos:.1%}",
         C["negative"] if overall_acos > target_acos else C["header_green"]),
        ("Scale Opportunity",   f"+${increase_oppt:,.2f}/day", C["header_green"]),
        ("Wasted Spend (30d)",  f"${wasted_spend:,.2f}",  C["negative"]),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    for i, (label, value, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name=FONT, size=8, color="888888")
        lc.fill = hex_fill("F5F5F5")
        lc.alignment = cal()
        lc.border = thin_border()
        vc = ws.cell(row=5, column=i, value=value)
        vc.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
        vc.fill = hex_fill(color)
        vc.alignment = cal()
        vc.border = thin_border()

    # Action breakdown
    ws.row_dimensions[7].height = 18
    apply_header(ws, 7,
                 ["Action", "Campaigns", "Total Budget Impact/Day",
                  "Total 30d Spend", "What It Means"],
                 C["header_dark"])
    action_meta = {
        "INCREASE":    ("📈 Increase Budget", C["header_green"],
                        "Efficient campaigns being throttled — give them more room"),
        "DECREASE":    ("📉 Decrease Budget", C["header_red"],
                        "Inefficient campaigns hitting cap — cut losses"),
        "INVESTIGATE": ("🔍 Investigate",     C["header_amber"],
                        "Severely underutilized — budgets may be too high"),
        "ALERT":       ("🚨 Alert",           C["header_alert"],
                        "Enabled campaigns with zero spend — not serving at all"),
        "NO CHANGE":   ("✅ No Change",       C["header_gray"],
                        "Within normal operating range"),
    }
    for r_idx, (action, (label, color, meaning)) in enumerate(action_meta.items(), 1):
        row_num = 7 + r_idx
        ws.row_dimensions[row_num].height = 18
        subset = df[df["action"] == action]
        impact = subset["budget_change"].sum() if len(subset) else 0
        spend  = subset["Spend"].sum() if len(subset) else 0
        fill   = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
        for c_idx, val in enumerate(
            [label, len(subset),
             f"${impact:+,.2f}" if impact else "-",
             f"${spend:,.2f}" if spend else "-",
             meaning], 1
        ):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font = bfont(bold=(c_idx == 1))
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()

    # Portfolio health table
    ws.row_dimensions[14].height = 18
    apply_header(ws, 14,
                 ["Portfolio", "Campaigns", "Daily Budget",
                  "Avg Daily Spend", "Utilization", "30d ACoS",
                  "Increase", "Decrease", "Investigate", "Alert"],
                 C["header_dark"])

    if "Portfolio Name" in df.columns:
        port = df.groupby("Portfolio Name").agg(
            campaigns=("Campaign Name", "count"),
            budget=("Daily Budget", "sum"),
            daily_spend=("daily_spend", "sum"),
            spend=("Spend", "sum"),
            sales=("Sales", "sum"),
        ).reset_index()
        port["util"]  = port["daily_spend"] / port["budget"].replace(0, np.nan)
        port["acos"]  = np.where(port["sales"] > 0,
                                 port["spend"] / port["sales"], np.nan)
        port = port.sort_values("spend", ascending=False)

        def ac(p, a):
            return len(df[(df["Portfolio Name"] == p) & (df["action"] == a)])

        for r_idx, (_, row) in enumerate(port.iterrows()):
            rn = 15 + r_idx
            ws.row_dimensions[rn].height = 17
            fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["light_gray"])
            acos_val = row["acos"] if not pd.isna(row["acos"]) else ""
            util_val = row["util"] if not pd.isna(row["util"]) else ""
            vals = [
                row["Portfolio Name"], int(row["campaigns"]),
                row["budget"], row["daily_spend"], util_val, acos_val,
                ac(row["Portfolio Name"], "INCREASE"),
                ac(row["Portfolio Name"], "DECREASE"),
                ac(row["Portfolio Name"], "INVESTIGATE"),
                ac(row["Portfolio Name"], "ALERT"),
            ]
            fmts = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                    '0.0%', '0.0%', None, None, None, None]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=rn, column=c_idx, value=val)
                cell.font = bfont()
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = lal()
                if fmt and val not in ("", None):
                    cell.number_format = fmt

    ws.freeze_panes = "A3"


def build_reallocation_sheet(wb, df):
    """Shows budget that could move from DECREASE campaigns to INCREASE campaigns."""
    ws = wb.create_sheet("♻ Reallocation Map")
    ws.sheet_view.showGridLines = False

    increase_df = df[df["action"] == "INCREASE"].sort_values("_acos")
    decrease_df = df[df["action"] == "DECREASE"].sort_values("Spend", ascending=False)

    freed   = abs(decrease_df["budget_change"].sum())
    needed  = increase_df["budget_change"].sum()
    net     = freed - needed

    sheet_title(ws,
        "Budget Reallocation Map",
        (f"${freed:,.2f}/day available from cuts | "
         f"${needed:,.2f}/day needed for increases | "
         f"Net: ${net:+,.2f}/day"),
        C["header_realloc"])

    # From (decrease) section
    ws.row_dimensions[3].height = 16
    from_headers = ["FROM — Cut These Budgets", "Portfolio",
                    "Current Budget", "Proposed Budget", "Freed/Day", "ACoS"]
    apply_header(ws, 3, from_headers, C["header_red"])

    for r_idx, (_, row) in enumerate(decrease_df.iterrows()):
        er = 4 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["red_light"])
        ws.row_dimensions[er].height = 15
        freed_amt = row["Daily Budget"] - row["new_budget"]
        acos_val  = row["_acos"] if not pd.isna(row.get("_acos", np.nan)) else ""
        vals = [row.get("Campaign Name", ""), row.get("Portfolio Name", ""),
                row["Daily Budget"], row["new_budget"], freed_amt, acos_val]
        fmts = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                '"$"#,##0.00', '0.0%']
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val not in ("", None):
                cell.number_format = fmt

    offset = 4 + len(decrease_df) + 2

    # To (increase) section
    ws.row_dimensions[offset].height = 16
    to_headers = ["TO — Fund These Campaigns", "Portfolio",
                  "Current Budget", "Proposed Budget", "Extra Needed/Day", "ACoS"]
    apply_header(ws, offset, to_headers, C["header_green"])

    for r_idx, (_, row) in enumerate(increase_df.iterrows()):
        er = offset + 1 + r_idx
        fill = hex_fill("FFFFFF") if r_idx % 2 == 0 else hex_fill(C["green_light"])
        ws.row_dimensions[er].height = 15
        extra = row["new_budget"] - row["Daily Budget"]
        acos_val = row["_acos"] if not pd.isna(row.get("_acos", np.nan)) else ""
        vals = [row.get("Campaign Name", ""), row.get("Portfolio Name", ""),
                row["Daily Budget"], row["new_budget"], extra, acos_val]
        fmts = [None, None, '"$"#,##0.00', '"$"#,##0.00',
                '"$"#,##0.00', '0.0%']
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if fmt and val not in ("", None):
                cell.number_format = fmt

    for col, w in zip("ABCDEF", [38, 26, 18, 18, 18, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


def build_bulk_upload(wb, df):
    changes = df[df["action"].isin(["INCREASE", "DECREASE"])].copy()
    ws = wb.create_sheet("📤 Amazon Bulk Upload")
    ws.sheet_view.showGridLines = False

    sheet_title(ws,
        f"Amazon Bulk Upload — {len(changes)} campaign budget updates",
        "⚠ Upload via Seller Central → Campaign Manager → Bulk Operations → Upload",
        C["header_dark"])

    headers = ["Product", "Entity", "Operation", "Campaign ID",
               "Campaign Name", "State", "Daily Budget",
               "Previous Budget", "Budget Change $", "Budget Change %", "Action Tag"]
    apply_header(ws, 3, headers, C["header_dark"])

    action_colors = {"INCREASE": C["positive"], "DECREASE": C["negative"]}

    for r_idx, (_, row) in enumerate(changes.iterrows()):
        er = 4 + r_idx
        action = row.get("action", "")
        light  = C["green_light"] if action == "INCREASE" else C["red_light"]
        fill   = hex_fill(light) if r_idx % 2 == 0 else hex_fill("FFFFFF")
        ws.row_dimensions[er].height = 15

        vals = {
            "Product":          "Sponsored Products",
            "Entity":           "Campaign",
            "Operation":        "update",
            "Campaign ID":      row.get("Campaign ID", ""),
            "Campaign Name":    row.get("Campaign Name", ""),
            "State":            row.get("State", "enabled"),
            "Daily Budget":     round(float(row.get("new_budget", 0)), 2),
            "Previous Budget":  round(float(row.get("Daily Budget", 0)), 2),
            "Budget Change $":  round(float(row.get("budget_change", 0)), 2),
            "Budget Change %":  row.get("budget_change_pct", 0),
            "Action Tag":       action,
        }
        fmts = {
            "Daily Budget":    '"$"#,##0.00',
            "Previous Budget": '"$"#,##0.00',
            "Budget Change $": '+$#,##0.00;-$#,##0.00;$-',
            "Budget Change %": '+0.0%;-0.0%;0.0%',
        }

        for c_idx, h in enumerate(headers, 1):
            val = vals.get(h, "")
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if h in fmts:
                cell.number_format = fmts[h]

        tag_cell = ws.cell(row=er, column=11)
        tag_cell.font = Font(name=FONT, bold=True, size=9,
                             color=action_colors.get(action, "1A1A1A"))

    for i, w in enumerate([18, 12, 12, 18, 36, 10, 18, 18, 16, 16, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def build_raw_sheet(wb, df):
    ws = wb.create_sheet("🗂 Raw Data")
    ws.sheet_view.showGridLines = False

    sheet_title(ws, "All Campaigns — Full Budget Analysis",
                "Every campaign with utilization, spend, and recommendation",
                C["header_dark"])

    all_cols = ["Portfolio Name", "Campaign Name", "State",
                "Daily Budget", "daily_spend", "budget_util",
                "Spend", "Sales", "Orders", "_acos",
                "action", "new_budget", "budget_change",
                "budget_change_pct", "reason"]
    all_hdrs = ["Portfolio", "Campaign", "State",
                "Daily Budget", "Avg Daily Spend", "Utilization",
                "30d Spend", "30d Sales", "30d Orders", "ACoS",
                "Action", "New Budget", "Change $", "Change %", "Reason"]
    fmts = {
        "Daily Budget": '"$"#,##0.00', "daily_spend": '"$"#,##0.00',
        "budget_util":  '0.0%',        "Spend": '"$"#,##0.00',
        "Sales": '"$"#,##0.00',        "_acos": '0.0%',
        "new_budget": '"$"#,##0.00',
        "budget_change": '+$#,##0.00;-$#,##0.00;$-',
        "budget_change_pct": '+0.0%;-0.0%;0.0%',
    }
    action_bg = {
        "INCREASE":    C["green_light"],
        "DECREASE":    C["red_light"],
        "INVESTIGATE": C["amber_light"],
        "ALERT":       "FFE4E4",
        "NO CHANGE":   "FFFFFF",
    }

    apply_header(ws, 3, all_hdrs, C["header_dark"])
    df_sorted = df.sort_values(
        ["action", "Spend"],
        key=lambda x: x.map(
            {"INCREASE": 0, "DECREASE": 1,
             "ALERT": 2, "INVESTIGATE": 3, "NO CHANGE": 4}
        ) if x.name == "action" else x,
        ascending=[True, False]
    )

    for r_idx, (_, row) in enumerate(df_sorted.iterrows()):
        er = 4 + r_idx
        act  = row.get("action", "NO CHANGE")
        fill = hex_fill(action_bg.get(act, "FFFFFF"))
        ws.row_dimensions[er].height = 14
        for c_idx, col in enumerate(all_cols, 1):
            val = row.get(col, "")
            if isinstance(val, float) and np.isnan(val):
                val = ""
            cell = ws.cell(row=er, column=c_idx, value=val)
            cell.font = bfont(size=9)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = lal()
            if col in fmts and val not in ("", None):
                cell.number_format = fmts[col]

    for i, w in enumerate(
        [22, 34, 10, 16, 16, 13, 14, 14, 12, 10,
         14, 14, 13, 13, 46], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Amazon PPC Budget Manager")
    parser.add_argument("--input",                   required=True)
    parser.add_argument("--output",                  required=True)
    parser.add_argument("--target-acos",             type=float, default=0.25)
    parser.add_argument("--days",                    type=int,   default=30)
    parser.add_argument("--constrained-threshold",   type=float, default=0.80)
    parser.add_argument("--underutilized-threshold", type=float, default=0.20)
    parser.add_argument("--increase-pct",            type=float, default=0.25)
    parser.add_argument("--decrease-pct",            type=float, default=0.20)
    parser.add_argument("--min-budget",              type=float, default=10.0)
    args = parser.parse_args()

    print(f"Loading bulk file: {args.input}")
    df = load_campaigns(args.input, args.days)
    print(f"  {len(df)} campaigns loaded")

    print("Classifying budgets...")
    df = classify(df, args.target_acos,
                  args.constrained_threshold, args.underutilized_threshold,
                  args.increase_pct, args.decrease_pct, args.min_budget)

    counts = df["action"].value_counts()
    print("\nResults:")
    for action, n in counts.items():
        print(f"  {action}: {n}")

    freed  = abs(df[df["action"] == "DECREASE"]["budget_change"].sum())
    needed = df[df["action"] == "INCREASE"]["budget_change"].sum()
    print(f"\n  Budget freed from cuts:    ${freed:,.2f}/day")
    print(f"  Budget needed for raises:  ${needed:,.2f}/day")
    print(f"  Net reallocation:          ${freed - needed:+,.2f}/day")

    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb, df, args.target_acos, args.days)
    inc_n = write_action_sheet(wb, df, "INCREASE", "📈 Increase Budget",
                               C["header_green"], C["green_light"],
                               sort_col="_acos", sort_asc=True)
    dec_n = write_action_sheet(wb, df, "DECREASE", "📉 Decrease Budget",
                               C["header_red"], C["red_light"],
                               sort_col="Spend", sort_asc=False)
    inv_n = write_action_sheet(wb, df, "INVESTIGATE", "🔍 Investigate",
                               C["header_amber"], C["amber_light"],
                               sort_col="Daily Budget", sort_asc=False)
    alt_n = write_action_sheet(wb, df, "ALERT", "🚨 Alerts",
                               C["header_alert"], "FFE4E4",
                               sort_col="Daily Budget", sort_asc=False)
    build_reallocation_sheet(wb, df)
    build_bulk_upload(wb, df)
    build_raw_sheet(wb, df)

    # ── Write cross-tool findings JSON ────────────────────────────────────
    constrained = df[(df["action"] == "INCREASE") | (
        (df["action"] == "NO CHANGE") & (df.get("budget_util", pd.Series(0, index=df.index)) >= 0.7)
    )]
    alerts = df[df["action"] == "ALERT"]
    findings = {
        "tool":             "budget_manager",
        "target_acos":      args.target_acos,
        "increase_count":   int(inc_n),
        "decrease_count":   int(dec_n),
        "investigate_count":int(inv_n),
        "alert_count":      int(alt_n),
        "budget_freed":     round(float(df[df["action"] == "DECREASE"]["Daily Budget"].sum() -
                                        df[df["action"] == "DECREASE"]["new_budget"].sum()), 2),
        "actions": (
            [{"priority": "HIGH",  "type": "ZERO_SPEND_ALERT",
              "subject":  r.get("Campaign Name", r.get("campaign","")),
              "campaign": r.get("Campaign Name", r.get("campaign","")),
              "impact_spend": 0.0,
              "detail": r.get("reason", "Enabled but zero spend — check bids and targeting")}
             for _, r in alerts.iterrows()]
            +
            [{"priority": "MEDIUM", "type": "BUDGET_CONSTRAINED",
              "subject":  r.get("Campaign Name", r.get("campaign","")),
              "campaign": r.get("Campaign Name", r.get("campaign","")),
              "impact_spend": round(float(r.get("Daily Budget", 0)), 2),
              "detail": r.get("reason","")}
             for _, r in df[df["action"] == "INCREASE"].iterrows()]
        ),
    }
    findings_path = args.output.replace(".xlsx", "_findings.json")
    with open(findings_path, "w") as _f:
        json.dump(findings, _f, indent=2, default=str)
    print(f"   Findings: {findings_path}")

    wb.save(args.output)
    print(f"\n✅ Saved: {args.output}")
    print(f"   📈 Increase:    {inc_n} campaigns")
    print(f"   📉 Decrease:    {dec_n} campaigns")
    print(f"   🔍 Investigate: {inv_n} campaigns")
    print(f"   🚨 Alert:       {alt_n} campaigns")
    print(f"   Bulk upload tab has {inc_n + dec_n} rows ready.")


if __name__ == "__main__":
    main()
