#!/usr/bin/env python3
"""
Amazon PPC Dashboard — Streamlit Web App
Manages multiple brands and orchestrates all 8 PPC optimization tools.

Prerequisites: pip3 install streamlit pandas numpy openpyxl
Run: streamlit run app.py
"""

import io
import json
import os
import sys
import glob
import subprocess
from datetime import date
from pathlib import Path

import streamlit as st

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Paths ────────────────────────────────────────────────────────────────────
APP_DIR = os.path.dirname(os.path.abspath(__file__))
BRANDS_FILE = os.path.join(APP_DIR, "brands.json")
UPLOAD_DIR = os.path.join(APP_DIR, "uploads")

# ── Default Global Settings ──────────────────────────────────────────────────
DEFAULT_GLOBALS = {
    "bid_floor": 0.20,
    "bid_ceiling": 5.00,
    "max_raise": 0.30,
    "max_lower": 0.40,
    "pause_spend": 15.0,
    "pause_clicks": 15,
    "min_clicks_bid": 10,
    "min_clicks_placement": 10,
    "neg_spend_threshold": 10.0,
    "min_orders": 1,
    "min_clicks_harvest": 3,
    "constrained_threshold": 0.80,
    "underutilized_threshold": 0.20,
    "increase_pct": 0.25,
    "decrease_pct": 0.20,
    "min_budget": 10.0,
    "max_increase_placement": 50,
    "max_decrease_placement": 100,
    "leakage_spend": 20,
    "skc_starting_bid": 0.50,
    "skc_bid_strategy": "inch-up",
    "skc_placement": "tos",
    "skc_daily_budget": 25.0,
    "skc_goal": "ranking",
    "days": 30,
}

DEFAULT_BRAND = {
    "name": "",
    "target_acos_profit": 0.25,
    "target_acos_ranking": 0.60,
    "target_acos_research": 0.35,
    "target_acos_reviews": 0.40,
    "target_acos_marketshare": 0.30,
    "brand_keywords": "",
    "primary_asin": "",
    "portfolio_names": "",
    "date_range": "",
}


# ── Persistence ──────────────────────────────────────────────────────────────

def load_config():
    if os.path.exists(BRANDS_FILE):
        with open(BRANDS_FILE) as f:
            data = json.load(f)
        if "global_defaults" not in data:
            data["global_defaults"] = DEFAULT_GLOBALS.copy()
        return data
    return {"brands": {}, "global_defaults": DEFAULT_GLOBALS.copy()}


def save_config(data):
    with open(BRANDS_FILE, "w") as f:
        json.dump(data, f, indent=2)


# ── File Management ──────────────────────────────────────────────────────────

def save_uploaded_file(uploaded_file, brand_name, file_type):
    brand_dir = os.path.join(UPLOAD_DIR, brand_name)
    os.makedirs(brand_dir, exist_ok=True)
    ext = Path(uploaded_file.name).suffix
    dest = os.path.join(brand_dir, f"{file_type}{ext}")
    with open(dest, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return dest


def get_uploaded_files(brand_name):
    brand_dir = os.path.join(UPLOAD_DIR, brand_name)
    result = {}
    for ft in ["bulk", "search_terms", "sqp"]:
        matches = glob.glob(os.path.join(brand_dir, f"{ft}.*"))
        result[ft] = matches[0] if matches else None
    return result


def get_output_dir(brand_name):
    d = os.path.expanduser(
        f"~/Desktop/PPC_Reports_{date.today().isoformat()}/{brand_name}"
    )
    os.makedirs(d, exist_ok=True)
    return d


# ── Argument Builders ────────────────────────────────────────────────────────

def _s(val):
    return str(val)


def build_campaign_strategist_args(brand, defaults, files, output_dir, output_file):
    return [
        "--input", files["bulk"],
        "--output", output_file,
        "--target-acos-profit", _s(brand.get("target_acos_profit", 0.25)),
        "--target-acos-ranking", _s(brand.get("target_acos_ranking", 0.60)),
        "--target-acos-research", _s(brand.get("target_acos_research", 0.35)),
        "--target-acos-reviews", _s(brand.get("target_acos_reviews", 0.40)),
        "--target-acos-marketshare", _s(brand.get("target_acos_marketshare", 0.30)),
        "--brand-keywords", brand.get("brand_keywords", ""),
        "--days", _s(defaults.get("days", 30)),
    ]


def build_harvester_args(brand, defaults, files, output_dir, output_file):
    args = [
        "--input", files["search_terms"],
        "--output", output_file,
        "--target-acos", _s(brand["target_acos_profit"]),
        "--min-clicks", _s(defaults["min_clicks_harvest"]),
        "--min-orders", _s(defaults["min_orders"]),
        "--neg-spend-threshold", _s(defaults["neg_spend_threshold"]),
    ]
    if files.get("bulk"):
        args += ["--bulk-file", files["bulk"]]
    return args


def build_bid_optimizer_args(brand, defaults, files, output_dir, output_file):
    return [
        "--input", files["bulk"],
        "--output", output_file,
        "--target-acos", _s(brand["target_acos_profit"]),
        "--min-clicks", _s(defaults["min_clicks_bid"]),
        "--max-raise", _s(defaults["max_raise"]),
        "--max-lower", _s(defaults["max_lower"]),
        "--bid-floor", _s(defaults["bid_floor"]),
        "--bid-ceiling", _s(defaults["bid_ceiling"]),
        "--pause-spend", _s(defaults["pause_spend"]),
        "--pause-clicks", _s(defaults["pause_clicks"]),
    ]


def build_budget_manager_args(brand, defaults, files, output_dir, output_file):
    return [
        "--input", files["bulk"],
        "--output", output_file,
        "--target-acos", _s(brand["target_acos_profit"]),
        "--days", _s(defaults["days"]),
        "--constrained-threshold", _s(defaults["constrained_threshold"]),
        "--underutilized-threshold", _s(defaults["underutilized_threshold"]),
        "--increase-pct", _s(defaults["increase_pct"]),
        "--decrease-pct", _s(defaults["decrease_pct"]),
        "--min-budget", _s(defaults["min_budget"]),
    ]


def build_placement_optimizer_args(brand, defaults, files, output_dir, output_file):
    return [
        "--input", files["bulk"],
        "--output", output_file,
        "--target-acos", _s(brand["target_acos_profit"]),
        "--min-clicks", _s(defaults["min_clicks_placement"]),
        "--max-increase", _s(defaults["max_increase_placement"]),
        "--max-decrease", _s(defaults["max_decrease_placement"]),
        "--leakage-spend", _s(defaults["leakage_spend"]),
    ]


def build_skc_builder_args(brand, defaults, files, output_dir, output_file):
    harvest_file = os.path.join(
        output_dir,
        f"2_Search_Term_Harvest_{date.today().isoformat()}.xlsx"
    )
    portfolio = brand.get("portfolio_names", "Default")
    if isinstance(portfolio, list):
        portfolio = portfolio[0] if portfolio else "Default"
    portfolio = portfolio.split(",")[0].strip() if portfolio else "Default"
    return [
        "--input", harvest_file,
        "--output", output_file,
        "--asin", brand.get("primary_asin", "B0XXXXXXXXX"),
        "--portfolio", portfolio,
        "--goal", defaults.get("skc_goal", "ranking"),
        "--target-acos", _s(brand.get("target_acos_ranking", 0.60)),
        "--starting-bid", _s(defaults["skc_starting_bid"]),
        "--bid-strategy", defaults["skc_bid_strategy"],
        "--placement", defaults["skc_placement"],
        "--daily-budget", _s(defaults["skc_daily_budget"]),
    ]


def build_rank_tracker_args(brand, defaults, files, output_dir, output_file):
    args = [
        "--bulk", files["bulk"],
        "--output", output_file,
        "--target-acos", _s(brand.get("target_acos_ranking", 0.60)),
        "--brand", brand["name"],
    ]
    if brand.get("primary_asin"):
        args += ["--asin", brand["primary_asin"]]
    if files.get("sqp"):
        args += ["--sqp", files["sqp"]]
    return args


def build_weekly_report_args(brand, defaults, files, output_dir, output_file):
    args = [
        "--bulk", files["bulk"],
        "--output", output_file,
        "--brand", brand["name"],
        "--target-acos", _s(brand["target_acos_profit"]),
        "--date-range", brand.get("date_range", ""),
    ]
    if files.get("search_terms"):
        args += ["--search-terms", files["search_terms"]]
    return args


# ── Tools Registry ───────────────────────────────────────────────────────────

TOOLS = [
    {"name": "Campaign Strategist", "key": "campaign_strategist",
     "script": "amazon-ppc-campaign-strategist/scripts/campaign_strategist.py",
     "prefix": "1_Campaign_Strategy", "requires": ["bulk"],
     "build": build_campaign_strategist_args},

    {"name": "Search Term Harvester", "key": "harvester",
     "script": "amazon-ppc-harvester/scripts/harvester.py",
     "prefix": "2_Search_Term_Harvest", "requires": ["search_terms"],
     "build": build_harvester_args},

    {"name": "Bid Optimizer", "key": "bid_optimizer",
     "script": "amazon-ppc-bid-optimizer/scripts/bid_optimizer.py",
     "prefix": "3_Bid_Optimization", "requires": ["bulk"],
     "build": build_bid_optimizer_args},

    {"name": "Budget Manager", "key": "budget_manager",
     "script": "amazon-ppc-budget-manager/scripts/budget_manager.py",
     "prefix": "4_Budget_Manager", "requires": ["bulk"],
     "build": build_budget_manager_args},

    {"name": "Placement Optimizer", "key": "placement_optimizer",
     "script": "amazon-ppc-placement-optimizer/scripts/placement_optimizer.py",
     "prefix": "5_Placement_Optimizer", "requires": ["bulk"],
     "build": build_placement_optimizer_args},

    {"name": "SKC Builder", "key": "skc_builder",
     "script": "amazon-ppc-skc-builder/scripts/skc_builder.py",
     "prefix": "6_SKC_Campaigns", "requires": ["bulk"],
     "build": build_skc_builder_args},

    {"name": "Rank Tracker", "key": "rank_tracker",
     "script": "amazon-ppc-rank-tracker/scripts/rank_tracker.py",
     "prefix": "7_Rank_Tracker", "requires": ["bulk"],
     "build": build_rank_tracker_args},

    {"name": "Weekly Report", "key": "weekly_report",
     "script": "amazon-ppc-weekly-report/scripts/weekly_report.py",
     "prefix": "8_Weekly_Report", "requires": ["bulk"],
     "build": build_weekly_report_args},
]


# ── Tool Execution ───────────────────────────────────────────────────────────

def run_tool(tool, brand, defaults, files, output_dir):
    script_path = os.path.join(APP_DIR, tool["script"])
    output_file = os.path.join(
        output_dir, f"{tool['prefix']}_{date.today().isoformat()}.xlsx"
    )
    args = tool["build"](brand, defaults, files, output_dir, output_file)
    cmd = [sys.executable, script_path] + args
    result = subprocess.run(
        cmd, capture_output=True, text=True, timeout=300, cwd=APP_DIR
    )
    return result, output_file


def build_summary_report(output_dir, brand_name):
    """Combine all tool output Excel files into one summary workbook."""
    if not HAS_OPENPYXL:
        return None, "openpyxl not installed"

    summary_path = os.path.join(
        output_dir, f"0_Summary_Report_{date.today().isoformat()}.xlsx"
    )
    summary_wb = openpyxl.Workbook()
    summary_wb.remove(summary_wb.active)  # remove default sheet

    xlsx_files = sorted([
        f for f in os.listdir(output_dir)
        if f.endswith(".xlsx") and not f.startswith("0_Summary")
    ])

    if not xlsx_files:
        return None, "No output files to combine"

    sheet_count = 0
    for xlsx_name in xlsx_files:
        xlsx_path = os.path.join(output_dir, xlsx_name)
        # Tool prefix like "1_Campaign_Strategy"
        tool_prefix = xlsx_name.split("_2")[0]  # strip date suffix
        # Shorter prefix: just the number + short name
        parts = xlsx_name.split("_")
        short_prefix = parts[0] if parts else ""

        try:
            src_wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception:
            continue

        for src_sheet_name in src_wb.sheetnames:
            src_ws = src_wb[src_sheet_name]
            # Build a sheet name that fits Excel's 31-char limit
            dest_name = f"{short_prefix}_{src_sheet_name}"
            if len(dest_name) > 31:
                dest_name = dest_name[:31]
            # Avoid duplicate sheet names
            base = dest_name
            counter = 2
            while dest_name in summary_wb.sheetnames:
                suffix = f"_{counter}"
                dest_name = base[:31 - len(suffix)] + suffix
                counter += 1

            dest_ws = summary_wb.create_sheet(title=dest_name)
            for row in src_ws.iter_rows(values_only=True):
                dest_ws.append(list(row))
            sheet_count += 1

        src_wb.close()

    if sheet_count == 0:
        return None, "No sheets found in output files"

    summary_wb.save(summary_path)
    return summary_path, None


def extract_bulk_sheet(filepath):
    """Extract the Amazon Bulk Upload sheet from a tool output file.

    Returns (sheet_name, rows_as_lists) or (None, None) if no bulk sheet found.
    """
    if not HAS_OPENPYXL:
        return None, None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None, None
    for name in wb.sheetnames:
        if "Bulk Upload" in name:
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
            return name, rows
    wb.close()
    return None, None


# Map file prefixes to friendly tab names for the combined bulk upload file
_BULK_TAB_NAMES = {
    "3": "Bid_Changes",
    "4": "Budget_Changes",
    "6": "New_SKCs",
}


def build_bulk_upload_bytes(rows, sheet_name="Bulk Upload"):
    """Create a single-sheet .xlsx in memory from a list of rows. Returns bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_combined_bulk_upload(output_dir):
    """Merge all bulk upload sheets from output files into one workbook.

    Returns (bytes, tab_count) or (None, 0) if no bulk sheets found.
    """
    if not HAS_OPENPYXL:
        return None, 0

    xlsx_files = sorted([
        f for f in os.listdir(output_dir)
        if f.endswith(".xlsx") and not f.startswith("0_Summary")
    ])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    tab_count = 0

    for xlsx_name in xlsx_files:
        filepath = os.path.join(output_dir, xlsx_name)
        prefix_num = xlsx_name.split("_")[0]  # e.g. "3", "4", "6"
        sheet_label = _BULK_TAB_NAMES.get(prefix_num, f"Tool_{prefix_num}")
        _, rows = extract_bulk_sheet(filepath)
        if rows and len(rows) > 1:  # has header + at least one data row
            ws = wb.create_sheet(title=sheet_label[:31])
            for row in rows:
                ws.append(row)
            tab_count += 1

    if tab_count == 0:
        return None, 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), tab_count


# ── Page: Brand Management ───────────────────────────────────────────────────

def page_brands():
    st.header("Brand Management")
    config = load_config()
    brands = config["brands"]

    # Add new brand
    st.subheader("Add New Brand")
    with st.form("add_brand", clear_on_submit=True):
        new_name = st.text_input("Brand Name", placeholder="e.g. RENUV")
        submitted = st.form_submit_button("Add Brand")
        if submitted and new_name.strip():
            name = new_name.strip().upper()
            if name not in brands:
                brands[name] = {**DEFAULT_BRAND, "name": name}
                save_config(config)
                st.success(f"Added brand: {name}")
                st.rerun()
            else:
                st.warning(f"Brand '{name}' already exists.")

    if not brands:
        st.info("No brands configured yet. Add one above.")
        return

    # Edit existing brands
    st.subheader("Edit Brands")
    for brand_name in sorted(brands.keys()):
        brand = brands[brand_name]
        with st.expander(f"{brand_name}", expanded=False):
            with st.form(f"edit_{brand_name}"):
                col1, col2 = st.columns(2)
                with col1:
                    profit = st.number_input("Profit ACoS %", 1, 100,
                        int(brand.get("target_acos_profit", 0.25) * 100),
                        key=f"profit_{brand_name}")
                    ranking = st.number_input("Ranking ACoS %", 1, 200,
                        int(brand.get("target_acos_ranking", 0.60) * 100),
                        key=f"ranking_{brand_name}")
                    research = st.number_input("Research ACoS %", 1, 100,
                        int(brand.get("target_acos_research", 0.35) * 100),
                        key=f"research_{brand_name}")
                with col2:
                    reviews = st.number_input("Reviews ACoS %", 1, 100,
                        int(brand.get("target_acos_reviews", 0.40) * 100),
                        key=f"reviews_{brand_name}")
                    marketshare = st.number_input("Market Share ACoS %", 1, 100,
                        int(brand.get("target_acos_marketshare", 0.30) * 100),
                        key=f"ms_{brand_name}")

                keywords = st.text_input("Brand Keywords (comma-separated)",
                    brand.get("brand_keywords", ""),
                    key=f"kw_{brand_name}")
                asin = st.text_input("Primary ASIN",
                    brand.get("primary_asin", ""),
                    key=f"asin_{brand_name}")
                portfolios = st.text_input("Portfolio Names (comma-separated)",
                    brand.get("portfolio_names", ""),
                    key=f"port_{brand_name}")
                date_range = st.text_input("Date Range Label",
                    brand.get("date_range", ""),
                    placeholder="e.g. Feb 1 - Feb 28, 2026",
                    key=f"dr_{brand_name}")

                c1, c2 = st.columns([3, 1])
                with c1:
                    save = st.form_submit_button("Save Changes")
                with c2:
                    delete = st.form_submit_button("Delete Brand")

                if save:
                    brand["target_acos_profit"] = profit / 100
                    brand["target_acos_ranking"] = ranking / 100
                    brand["target_acos_research"] = research / 100
                    brand["target_acos_reviews"] = reviews / 100
                    brand["target_acos_marketshare"] = marketshare / 100
                    brand["brand_keywords"] = keywords
                    brand["primary_asin"] = asin
                    brand["portfolio_names"] = portfolios
                    brand["date_range"] = date_range
                    save_config(config)
                    st.success(f"Saved {brand_name}")

                if delete:
                    del brands[brand_name]
                    save_config(config)
                    st.warning(f"Deleted {brand_name}")
                    st.rerun()


# ── Page: Run Analysis ───────────────────────────────────────────────────────

def page_run():
    st.header("Run PPC Analysis")
    config = load_config()
    brands = config["brands"]
    defaults = config["global_defaults"]

    if not brands:
        st.warning("No brands configured. Go to **Brands** page first.")
        return

    brand_name = st.selectbox("Select Brand", sorted(brands.keys()))
    brand = brands[brand_name]

    # ── File Upload ──
    st.subheader("Upload Files")
    col1, col2, col3 = st.columns(3)

    with col1:
        bulk = st.file_uploader("Bulk Operations (.xlsx)", type=["xlsx", "xls"],
                                key=f"up_bulk_{brand_name}")
        if bulk:
            path = save_uploaded_file(bulk, brand_name, "bulk")
            st.success(f"Saved: {os.path.basename(path)}")

    with col2:
        st_file = st.file_uploader("Search Term Report (.xlsx)", type=["xlsx", "xls"],
                                   key=f"up_st_{brand_name}")
        if st_file:
            path = save_uploaded_file(st_file, brand_name, "search_terms")
            st.success(f"Saved: {os.path.basename(path)}")

    with col3:
        sqp = st.file_uploader("Brand Analytics / SQP (optional, .csv)", type=["csv"],
                               key=f"up_sqp_{brand_name}")
        if sqp:
            path = save_uploaded_file(sqp, brand_name, "sqp")
            st.success(f"Saved: {os.path.basename(path)}")

    files = get_uploaded_files(brand_name)

    # Show upload status
    status_cols = st.columns(3)
    for i, (ft, label) in enumerate([
        ("bulk", "Bulk Ops"), ("search_terms", "Search Terms"), ("sqp", "SQP")
    ]):
        with status_cols[i]:
            if files.get(ft):
                st.markdown(f"**{label}:** {os.path.basename(files[ft])}")
            else:
                req = " (required)" if ft != "sqp" else " (optional)"
                st.markdown(f"**{label}:** not uploaded{req}")

    st.divider()

    # ── Run All ──
    st.subheader("Full Analysis")

    missing = []
    if not files.get("bulk"):
        missing.append("Bulk Operations file")
    if not files.get("search_terms"):
        missing.append("Search Term Report")

    if missing:
        st.warning(f"Upload required files first: {', '.join(missing)}")
    else:
        if st.button("Run All 8 Tools", type="primary", use_container_width=True):
            output_dir = get_output_dir(brand_name)
            with st.status("Running full PPC analysis...", expanded=True) as status:
                for i, tool in enumerate(TOOLS):
                    st.write(f"**{i+1}/8** {tool['name']}...")
                    try:
                        result, out_path = run_tool(
                            tool, brand, defaults, files, output_dir
                        )
                        if result.returncode == 0:
                            st.write(f"  {tool['name']} — done")
                        else:
                            err = result.stderr[-800:] if result.stderr else "Unknown error"
                            st.error(f"**{tool['name']} failed:**\n```\n{err}\n```")
                            status.update(
                                label=f"Failed at {tool['name']}",
                                state="error"
                            )
                            break
                    except subprocess.TimeoutExpired:
                        st.error(f"**{tool['name']}** timed out (5 min limit)")
                        status.update(
                            label=f"Timeout at {tool['name']}",
                            state="error"
                        )
                        break
                    except Exception as e:
                        st.error(f"**{tool['name']}** error: {e}")
                        status.update(
                            label=f"Error at {tool['name']}",
                            state="error"
                        )
                        break
                else:
                    # Build combined summary report
                    st.write("**Building summary report...**")
                    summary_path, err = build_summary_report(output_dir, brand_name)
                    if summary_path:
                        st.write("Summary report — done")
                    elif err:
                        st.warning(f"Summary report skipped: {err}")
                    status.update(
                        label="All 8 tools complete!",
                        state="complete"
                    )

    # Standalone summary report button
    output_dir_check = get_output_dir(brand_name)
    existing_outputs = [f for f in os.listdir(output_dir_check)
                        if f.endswith(".xlsx") and not f.startswith("0_Summary")] if os.path.exists(output_dir_check) else []
    if existing_outputs:
        if st.button("Rebuild Summary Report", key=f"rebuild_summary_{brand_name}"):
            summary_path, err = build_summary_report(output_dir_check, brand_name)
            if summary_path:
                st.success("Summary report rebuilt!")
            else:
                st.error(f"Failed: {err}")

    st.divider()

    # ── Individual Tools ──
    st.subheader("Run Individual Tools")
    for tool in TOOLS:
        col_name, col_btn, col_status = st.columns([4, 1, 2])
        with col_name:
            st.markdown(f"**{tool['prefix'].split('_', 1)[0]}. {tool['name']}**")
        with col_btn:
            can_run = all(files.get(r) for r in tool["requires"])
            if st.button("Run", key=f"run_{tool['key']}_{brand_name}",
                         disabled=not can_run):
                output_dir = get_output_dir(brand_name)
                try:
                    result, out_path = run_tool(
                        tool, brand, defaults, files, output_dir
                    )
                    st.session_state[f"result_{tool['key']}"] = (
                        result.returncode, result.stdout[-500:], result.stderr[-500:]
                    )
                except Exception as e:
                    st.session_state[f"result_{tool['key']}"] = (1, "", str(e))
        with col_status:
            r = st.session_state.get(f"result_{tool['key']}")
            if r is not None:
                if r[0] == 0:
                    st.success("Done")
                else:
                    st.error("Failed")

    st.divider()

    # ── Download Results ──
    st.subheader("Download Results")
    output_dir = get_output_dir(brand_name)
    if os.path.exists(output_dir):
        xlsx_files = sorted([f for f in os.listdir(output_dir) if f.endswith(".xlsx")])
        if xlsx_files:
            # ── Combined Bulk Upload button at the top ──
            combined_data, tab_count = build_combined_bulk_upload(output_dir)
            if combined_data and tab_count > 0:
                st.download_button(
                    label=f"⬆ Combined Bulk Upload — All Changes ({tab_count} tabs)",
                    data=combined_data,
                    file_name=f"Bulk_Upload_All_Changes_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_combined_bulk_{brand_name}",
                )
                st.caption("One file with all bid changes, budget changes, and new SKCs — ready for Seller Central.")
                st.divider()

            # ── Individual file downloads ──
            for f in xlsx_files:
                filepath = os.path.join(output_dir, f)
                _, bulk_rows = extract_bulk_sheet(filepath)
                has_bulk = bulk_rows is not None and len(bulk_rows) > 1

                if has_bulk:
                    col_full, col_bulk = st.columns([3, 2])
                else:
                    col_full, col_bulk = st.columns([3, 2])

                with col_full:
                    with open(filepath, "rb") as fh:
                        st.download_button(
                            label=f"📊 {f}",
                            data=fh.read(),
                            file_name=f,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{f}_{brand_name}",
                        )
                with col_bulk:
                    if has_bulk:
                        bulk_bytes = build_bulk_upload_bytes(bulk_rows)
                        bulk_fname = f.replace(".xlsx", "_BULK_UPLOAD.xlsx")
                        st.download_button(
                            label=f"⬆ Bulk Upload",
                            data=bulk_bytes,
                            file_name=bulk_fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dlbulk_{f}_{brand_name}",
                        )
        else:
            st.info("No output files yet. Run the analysis first.")
    else:
        st.info("No output files yet. Run the analysis first.")


# ── Page: Settings ───────────────────────────────────────────────────────────

def page_settings():
    st.header("Global Default Settings")
    st.caption("These defaults apply to all brands unless overridden at the brand level.")

    config = load_config()
    d = config["global_defaults"]

    with st.form("settings_form"):
        st.subheader("Bid Optimizer")
        col1, col2, col3 = st.columns(3)
        with col1:
            bid_floor = st.number_input("Bid Floor ($)", 0.01, 10.0,
                                        d.get("bid_floor", 0.20), step=0.05)
            max_raise = st.number_input("Max Raise %", 0.05, 1.0,
                                        d.get("max_raise", 0.30), step=0.05)
            pause_spend = st.number_input("Pause Spend ($)", 1.0, 100.0,
                                          d.get("pause_spend", 15.0), step=1.0)
        with col2:
            bid_ceiling = st.number_input("Bid Ceiling ($)", 1.0, 20.0,
                                          d.get("bid_ceiling", 5.00), step=0.50)
            max_lower = st.number_input("Max Lower %", 0.05, 1.0,
                                        d.get("max_lower", 0.40), step=0.05)
            pause_clicks = st.number_input("Pause Clicks", 1, 100,
                                           d.get("pause_clicks", 15))
        with col3:
            min_clicks_bid = st.number_input("Min Clicks (Bid)", 1, 50,
                                             d.get("min_clicks_bid", 10))

        st.subheader("Harvester")
        col1, col2, col3 = st.columns(3)
        with col1:
            min_clicks_harvest = st.number_input("Min Clicks (Harvest)", 1, 20,
                                                  d.get("min_clicks_harvest", 3))
        with col2:
            min_orders = st.number_input("Min Orders", 1, 10,
                                         d.get("min_orders", 1))
        with col3:
            neg_spend = st.number_input("Neg Spend Threshold ($)", 1.0, 100.0,
                                        d.get("neg_spend_threshold", 10.0), step=1.0)

        st.subheader("Budget Manager")
        col1, col2, col3 = st.columns(3)
        with col1:
            constrained = st.number_input("Constrained Threshold %", 0.5, 1.0,
                                          d.get("constrained_threshold", 0.80), step=0.05)
            increase_pct = st.number_input("Budget Increase %", 0.05, 1.0,
                                           d.get("increase_pct", 0.25), step=0.05)
        with col2:
            underutil = st.number_input("Underutilized Threshold %", 0.05, 0.5,
                                        d.get("underutilized_threshold", 0.20), step=0.05)
            decrease_pct = st.number_input("Budget Decrease %", 0.05, 1.0,
                                           d.get("decrease_pct", 0.20), step=0.05)
        with col3:
            min_budget = st.number_input("Min Daily Budget ($)", 1.0, 100.0,
                                         d.get("min_budget", 10.0), step=1.0)

        st.subheader("Placement Optimizer")
        col1, col2, col3 = st.columns(3)
        with col1:
            min_clicks_place = st.number_input("Min Clicks (Placement)", 1, 50,
                                               d.get("min_clicks_placement", 10))
        with col2:
            max_inc = st.number_input("Max Modifier Increase", 10, 200,
                                      d.get("max_increase_placement", 50))
        with col3:
            max_dec = st.number_input("Max Modifier Decrease", 10, 200,
                                      d.get("max_decrease_placement", 100))
        leakage = st.number_input("Leakage Spend Threshold ($)", 1.0, 100.0,
                                  d.get("leakage_spend", 20.0), step=1.0)

        st.subheader("SKC Builder")
        col1, col2, col3 = st.columns(3)
        with col1:
            skc_bid = st.number_input("Starting Bid ($)", 0.10, 5.0,
                                      d.get("skc_starting_bid", 0.50), step=0.10)
            skc_budget = st.number_input("Daily Budget ($)", 5.0, 100.0,
                                         d.get("skc_daily_budget", 25.0), step=5.0)
        with col2:
            skc_strategy = st.selectbox("Bid Strategy",
                                        ["inch-up", "revenue"],
                                        index=0 if d.get("skc_bid_strategy") == "inch-up" else 1)
        with col3:
            skc_place = st.selectbox("Default Placement",
                                     ["tos", "ros", "pp", "all"],
                                     index=["tos", "ros", "pp", "all"].index(
                                         d.get("skc_placement", "tos")))
            skc_goal = st.selectbox("Default Goal",
                                    ["ranking", "profit", "reviews", "marketshare", "research"],
                                    index=["ranking", "profit", "reviews", "marketshare", "research"].index(
                                        d.get("skc_goal", "ranking")))

        st.subheader("General")
        days = st.number_input("Lookback Days", 7, 90, d.get("days", 30))

        submitted = st.form_submit_button("Save Settings", type="primary",
                                          use_container_width=True)
        if submitted:
            config["global_defaults"] = {
                "bid_floor": bid_floor,
                "bid_ceiling": bid_ceiling,
                "max_raise": max_raise,
                "max_lower": max_lower,
                "pause_spend": pause_spend,
                "pause_clicks": pause_clicks,
                "min_clicks_bid": min_clicks_bid,
                "min_clicks_placement": min_clicks_place,
                "neg_spend_threshold": neg_spend,
                "min_orders": min_orders,
                "min_clicks_harvest": min_clicks_harvest,
                "constrained_threshold": constrained,
                "underutilized_threshold": underutil,
                "increase_pct": increase_pct,
                "decrease_pct": decrease_pct,
                "min_budget": min_budget,
                "max_increase_placement": max_inc,
                "max_decrease_placement": max_dec,
                "leakage_spend": leakage,
                "skc_starting_bid": skc_bid,
                "skc_bid_strategy": skc_strategy,
                "skc_placement": skc_place,
                "skc_daily_budget": skc_budget,
                "skc_goal": skc_goal,
                "days": days,
            }
            save_config(config)
            st.success("Settings saved.")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Amazon PPC Dashboard",
        page_icon="🎯",
        layout="wide",
    )

    st.sidebar.title("Amazon PPC Dashboard")
    st.sidebar.caption("AdsCrafted Method")

    page = st.sidebar.radio(
        "Navigation",
        ["Run Analysis", "Brands", "Settings"],
        index=0,
    )

    # Quick stats in sidebar
    config = load_config()
    brand_count = len(config.get("brands", {}))
    st.sidebar.divider()
    st.sidebar.metric("Brands Configured", brand_count)

    if brand_count > 0:
        for bn in sorted(config["brands"].keys()):
            files = get_uploaded_files(bn)
            has_bulk = bool(files.get("bulk"))
            has_st = bool(files.get("search_terms"))
            icon = "🟢" if (has_bulk and has_st) else "🟡" if (has_bulk or has_st) else "🔴"
            st.sidebar.caption(f"{icon} {bn}")

    st.sidebar.divider()
    st.sidebar.caption(
        f"Output: ~/Desktop/PPC_Reports_{date.today().isoformat()}/"
    )

    # Route to page
    if page == "Run Analysis":
        page_run()
    elif page == "Brands":
        page_brands()
    elif page == "Settings":
        page_settings()


if __name__ == "__main__":
    main()
